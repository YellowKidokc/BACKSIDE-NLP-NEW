from __future__ import annotations

import os
import re
from functools import lru_cache
from pathlib import Path
from typing import Iterable


DEFAULT_LEXICON_XLSX = Path(os.environ.get(
    "PAPER_GRADER_LEXICON_XLSX",
    r"\\dlowenas\HPWorkstation\Desktop\paper_grader_lexicons_master_enhanced.xlsx",
))


def normalize_term(value) -> str:
    term = str(value or "").strip().lower()
    return re.sub(r"\s+", " ", term)


@lru_cache(maxsize=4)
def load_workbook_cached(path_value: str):
    from openpyxl import load_workbook

    path = Path(path_value)
    if not path.exists():
        return None
    return load_workbook(path, read_only=True, data_only=True)

def workbook(path: Path | None = None):
    try:
        return load_workbook_cached(str(path or DEFAULT_LEXICON_XLSX))
    except Exception:
        return None


def sheet_terms(sheet_name: str, column_name: str = "term", path: Path | None = None) -> set[str]:
    wb = workbook(path)
    if wb is None or sheet_name not in wb.sheetnames:
        return set()
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    try:
        header = [str(cell or "").strip() for cell in next(rows)]
    except StopIteration:
        return set()
    if column_name not in header:
        return set()
    index = header.index(column_name)
    terms: set[str] = set()
    for row in rows:
        if index >= len(row):
            continue
        term = normalize_term(row[index])
        if term:
            terms.add(term)
    return terms

def semantic_terms(
    buckets: Iterable[str] | None = None,
    subbuckets: Iterable[str] | None = None,
    danger_levels: Iterable[str] | None = None,
    polarities: Iterable[str] | None = None,
    path: Path | None = None,
) -> set[str]:
    wb = workbook(path)
    if wb is None or "SEMANTIC_BUCKETS" not in wb.sheetnames:
        return set()
    ws = wb["SEMANTIC_BUCKETS"]
    rows = ws.iter_rows(values_only=True)
    try:
        header = [str(cell or "").strip() for cell in next(rows)]
    except StopIteration:
        return set()
    required = {"term", "bucket", "subbucket", "danger_level", "polarity"}
    if not required.issubset(set(header)):
        return set()

    term_i = header.index("term")
    bucket_i = header.index("bucket")
    subbucket_i = header.index("subbucket")
    danger_i = header.index("danger_level")
    polarity_i = header.index("polarity")
    bucket_filter = {normalize_term(item).upper() for item in buckets or []}
    subbucket_filter = {normalize_term(item) for item in subbuckets or []}
    danger_filter = {normalize_term(item) for item in danger_levels or []}
    polarity_filter = {normalize_term(item) for item in polarities or []}
    terms: set[str] = set()
    for row in rows:
        term = normalize_term(row[term_i] if term_i < len(row) else "")
        bucket = normalize_term(row[bucket_i] if bucket_i < len(row) else "").upper()
        subbucket = normalize_term(row[subbucket_i] if subbucket_i < len(row) else "")
        danger = normalize_term(row[danger_i] if danger_i < len(row) else "")
        polarity = normalize_term(row[polarity_i] if polarity_i < len(row) else "")
        if not term:
            continue
        if bucket_filter and bucket not in bucket_filter:
            continue
        if subbucket_filter and subbucket not in subbucket_filter:
            continue
        if danger_filter and danger not in danger_filter:
            continue
        if polarity_filter and polarity not in polarity_filter:
            continue
        terms.add(term)
    return terms


def regex_from_terms(terms: Iterable[str], fallback_pattern: str, flags: int = re.IGNORECASE) -> re.Pattern:
    clean = sorted({normalize_term(term) for term in terms if normalize_term(term)}, key=len, reverse=True)
    if not clean:
        return re.compile(fallback_pattern, flags)
    escaped = [re.escape(term).replace(r"\ ", r"\s+") for term in clean]
    return re.compile(r"\b(" + "|".join(escaped) + r")\b", flags)


def lexicon_status(path: Path | None = None) -> dict[str, object]:
    target = path or DEFAULT_LEXICON_XLSX
    wb = workbook(target)
    return {
        "source": "canonical workbook" if wb is not None else "embedded fallback",
        "path": str(target),
        "loaded": wb is not None,
    }
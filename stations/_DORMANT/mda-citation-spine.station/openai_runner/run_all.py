#!/usr/bin/env python3
"""
UNIFIED ASYNC RUNNER — Fire All OpenAI Tools in Parallel
=========================================================
Takes a paper (or folder of papers) and runs ALL applicable scoring
tools concurrently. What used to take 10 minutes sequentially now
takes ~30 seconds.

Usage:
    python run_all.py paper.md                    # All tools, one paper
    python run_all.py paper.md --tools ckg,domain # Only specific tools
    python run_all.py folder/                     # All papers in folder
    python run_all.py paper.md --dry-run          # Show plan, no API calls
    python run_all.py paper.md --batch            # Use Batch API (50% off, 24h)

Requires: pip install openai aiofiles
"""

import asyncio
import json
import os
import sys
import time
import argparse
from pathlib import Path
from datetime import datetime
from dataclasses import dataclass, field

# ---------------------------------------------------------------------------
# Pricing (USD per 1K tokens, 2025-Q2)
# ---------------------------------------------------------------------------
PRICING = {
    "gpt-4o":       (0.0025,  0.0100),
    "gpt-4o-mini":  (0.00015, 0.0006),
    "gpt-4-turbo":  (0.0100,  0.0300),
    "o1":           (0.0150,  0.0600),
    "o3-mini":      (0.0011,  0.0044),
}

SCRIPT_DIR = Path(__file__).resolve().parent

# ---------------------------------------------------------------------------
# Tool Registry — each tool is a prompt + model + parser
# ---------------------------------------------------------------------------
@dataclass
class Tool:
    name: str
    description: str
    prompt_file: str          # relative to SCRIPT_DIR
    model: str = "gpt-4o"
    temperature: float = 0.0
    max_tokens: int = 4096
    system_prompt: bool = True  # prompt goes as system msg (vs user msg)
    output_suffix: str = ".json"
    parse_json: bool = True    # try to parse response as JSON
    config_dir: str = ""       # where to find config.txt (if tool-specific)


TOOLS = {
    "ckg": Tool(
        name="ckg",
        description="CKG Epistemic Structure (5-tier, 0-100)",
        prompt_file="CKG/prompt.txt",
        model="gpt-4o-mini",
        temperature=0.0,
        output_suffix="_CKG.json",
        config_dir="CKG",
    ),
    "domain": Tool(
        name="domain",
        description="Domain Activation Vector (47 families)",
        prompt_file="DOMAIN/prompt.txt",
        model="gpt-4o-mini",
        temperature=0.0,
        output_suffix="_DOMAINS.json",
        config_dir="DOMAIN",
    ),
    "review": Tool(
        name="review",
        description="CDCM 44-Criteria Academic Review (0-100)",
        prompt_file="PAPER-REVIEW/prompt.txt",
        model="gpt-4o",
        temperature=0.0,
        output_suffix="_CDCM.json",
        config_dir="PAPER-REVIEW",
    ),
    "7q": Tool(
        name="7q",
        description="7Q Framework Forward Analysis",
        prompt_file="CALL/prompt_7q.txt",
        model="o3-mini",
        temperature=0.4,
        max_tokens=8192,
        output_suffix="_7Q.md",
        parse_json=False,
        config_dir="CALL",
    ),
    "decompress": Tool(
        name="decompress",
        description="Lossless Expansion (t:k|v to prose)",
        prompt_file="DECOMPRESS/prompt.txt",
        model="gpt-4o",
        temperature=0.0,
        max_tokens=16000,
        output_suffix="_EXPANDED.md",
        parse_json=False,
        config_dir="DECOMPRESS",
    ),
    "yaml": Tool(
        name="yaml",
        description="YAML Metadata Copilot (2-step classification)",
        prompt_file="YAML-COPILOT/prompt.txt",
        model="gpt-4o-mini",
        temperature=0.1,
        output_suffix="_YAML.json",
        config_dir="YAML-COPILOT",
    ),
    # ── Stage D/E/F (OPERATOR_2 pipeline) ──────────────────────
    "7q_judge": Tool(
        name="7q_judge",
        description="Stage D: 7Q+ adversarial analysis with cross-domain significance",
        prompt_file="CALL/prompt_stage_d_7q_judge.txt",
        model="gpt-4o",
        temperature=0.0,
        max_tokens=8192,
        output_suffix="_7Q_JUDGE.json",
        parse_json=True,
        config_dir="CALL",
    ),
    "academic": Tool(
        name="academic",
        description="Stage E: Academic review — citations, math, logic, novelty",
        prompt_file="CALL/prompt_stage_e_academic.txt",
        model="gpt-4o",
        temperature=0.0,
        max_tokens=6144,
        output_suffix="_ACADEMIC.json",
        parse_json=True,
        config_dir="CALL",
    ),
    "extractor": Tool(
        name="extractor",
        description="Stage F: Structural extraction — axiom mapping, theory resonance, imports",
        prompt_file="CALL/prompt_stage_f_extractor.txt",
        model="gpt-4o",
        temperature=0.0,
        max_tokens=8192,
        output_suffix="_EXTRACTOR.json",
        parse_json=True,
        config_dir="CALL",
    ),
}

# Default set when no --tools specified
DEFAULT_TOOLS = ["ckg", "domain", "review", "7q_judge", "academic", "extractor"]


# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
def load_api_key() -> str:
    """Find API key from any config.txt in the tree."""
    env_key = os.environ.get("OPENAI_API_KEY", "").strip()
    if env_key:
        return env_key

    # Check each tool dir, then root
    for subdir in ["CALL", "CKG", "DOMAIN", "PAPER-REVIEW", "VAULT-RATER",
                   "YAML-COPILOT", "DECOMPRESS"]:
        cfg_path = SCRIPT_DIR / subdir / "config.txt"
        if cfg_path.exists():
            for line in cfg_path.read_text().splitlines():
                line = line.strip()
                if line.startswith("OPENAI_API_KEY=") and "PASTE" not in line:
                    return line.split("=", 1)[1].strip()
    return ""


# ---------------------------------------------------------------------------
# Async OpenAI caller
# ---------------------------------------------------------------------------
async def call_tool(client, tool: Tool, paper_text: str, paper_name: str,
                    output_dir: Path, semaphore: asyncio.Semaphore,
                    ts: str) -> dict:
    """Run one tool against one paper asynchronously."""
    async with semaphore:
        prompt_path = SCRIPT_DIR / tool.prompt_file
        if not prompt_path.exists():
            return {"tool": tool.name, "status": "error",
                    "error": f"Prompt not found: {prompt_path}"}

        system_prompt = prompt_path.read_text(encoding="utf-8")

        if tool.system_prompt:
            messages = [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": paper_text},
            ]
        else:
            messages = [
                {"role": "user", "content": system_prompt + "\n\n" + paper_text},
            ]

        kwargs = {"model": tool.model, "messages": messages}

        if not tool.model.startswith("o"):
            kwargs["temperature"] = tool.temperature

        if tool.max_tokens:
            if tool.model.startswith("o"):
                kwargs["max_completion_tokens"] = tool.max_tokens
            else:
                kwargs["max_tokens"] = tool.max_tokens

        t0 = time.time()
        print(f"  [{tool.name}] Sending to {tool.model}...")

        try:
            response = await client.chat.completions.create(**kwargs)
        except Exception as e:
            elapsed = time.time() - t0
            print(f"  [{tool.name}] FAILED in {elapsed:.1f}s: {e}")
            return {"tool": tool.name, "status": "error", "error": str(e)}

        elapsed = time.time() - t0
        reply = response.choices[0].message.content
        usage = response.usage

        in_tok = usage.prompt_tokens if usage else 0
        out_tok = usage.completion_tokens if usage else 0

        # Cost
        cost = 0.0
        if tool.model in PRICING:
            in_rate, out_rate = PRICING[tool.model]
            cost = (in_tok / 1000) * in_rate + (out_tok / 1000) * out_rate

        print(f"  [{tool.name}] Done in {elapsed:.1f}s "
              f"({in_tok}+{out_tok} tokens, ${cost:.4f})")

        # Parse JSON if expected
        parsed = None
        if tool.parse_json:
            try:
                parsed = json.loads(reply)
            except json.JSONDecodeError:
                if "```json" in reply:
                    try:
                        block = reply.split("```json")[1].split("```")[0].strip()
                        parsed = json.loads(block)
                    except (json.JSONDecodeError, IndexError):
                        parsed = None

            if parsed:
                parsed["_metadata"] = {
                    "tool": tool.name,
                    "model": tool.model,
                    "input_file": paper_name,
                    "timestamp": datetime.now().isoformat(),
                    "tokens_in": in_tok,
                    "tokens_out": out_tok,
                    "cost_usd": round(cost, 6),
                    "elapsed_s": round(elapsed, 1),
                }

        # Save output
        stem = Path(paper_name).stem
        out_file = output_dir / f"{stem}{tool.output_suffix}"

        if parsed:
            out_file.write_text(json.dumps(parsed, indent=2), encoding="utf-8")
        else:
            out_file = output_dir / f"{stem}_{tool.name}.md"
            out_file.write_text(reply, encoding="utf-8")

        return {
            "tool": tool.name,
            "status": "ok",
            "output_file": str(out_file),
            "tokens_in": in_tok,
            "tokens_out": out_tok,
            "cost_usd": round(cost, 6),
            "elapsed_s": round(elapsed, 1),
            "parsed": parsed is not None,
        }


async def run_paper(api_key: str, paper_path: Path, tool_names: list,
                    output_dir: Path, max_concurrent: int = 8) -> dict:
    """Run all selected tools against one paper, in parallel."""
    from openai import AsyncOpenAI

    client = AsyncOpenAI(api_key=api_key)
    semaphore = asyncio.Semaphore(max_concurrent)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    paper_text = paper_path.read_text(encoding="utf-8")
    paper_name = paper_path.name

    # Truncate if needed
    if len(paper_text) > 60000:
        print(f"  WARNING: {paper_name} is {len(paper_text)} chars, truncating to 60K")
        paper_text = paper_text[:60000] + "\n\n[TRUNCATED]"

    tools_to_run = [TOOLS[name] for name in tool_names if name in TOOLS]

    print(f"\n{'='*60}")
    print(f"  {paper_name}")
    print(f"  Tools: {', '.join(t.name for t in tools_to_run)}")
    print(f"  Concurrent limit: {max_concurrent}")
    print(f"{'='*60}")

    t0 = time.time()

    tasks = [
        call_tool(client, tool, paper_text, paper_name, output_dir, semaphore, ts)
        for tool in tools_to_run
    ]
    results = await asyncio.gather(*tasks, return_exceptions=True)

    total_elapsed = time.time() - t0

    # Process results
    final_results = []
    total_cost = 0.0
    total_in = 0
    total_out = 0

    for r in results:
        if isinstance(r, Exception):
            final_results.append({"tool": "?", "status": "exception", "error": str(r)})
        else:
            final_results.append(r)
            if r.get("status") == "ok":
                total_cost += r.get("cost_usd", 0)
                total_in += r.get("tokens_in", 0)
                total_out += r.get("tokens_out", 0)

    print(f"\n{'-'*60}")
    print(f"  ALL DONE in {total_elapsed:.1f}s (wall clock)")
    print(f"  Total tokens: {total_in:,} in + {total_out:,} out")
    print(f"  Total cost:   ${total_cost:.4f}")

    # What sequential would have cost in time
    seq_time = sum(r.get("elapsed_s", 0) for r in final_results if isinstance(r, dict))
    if seq_time > 0:
        speedup = seq_time / total_elapsed if total_elapsed > 0 else 1
        print(f"  Sequential would have been: {seq_time:.1f}s ({speedup:.1f}x speedup)")
    print(f"{'-'*60}")

    # ── Post-processing: Annotation Injection ──────────────────
    # If Stage D/E/F ran, inject %%{...}%% blocks into the paper
    stage_tools = {"7q_judge", "academic", "extractor"}
    ran_stages = {r["tool"] for r in final_results
                  if isinstance(r, dict) and r.get("status") == "ok"}

    annotate_source = os.environ.get("OPENAI_RUNNER_ANNOTATE_SOURCE", "").strip().lower() in {"1", "true", "yes"}
    if annotate_source and ran_stages & stage_tools:
        try:
            from annotation_injector import inject_annotations
            stem = paper_path.stem
            d_path = output_dir / f"{stem}_7Q_JUDGE.json"
            e_path = output_dir / f"{stem}_ACADEMIC.json"
            f_path = output_dir / f"{stem}_EXTRACTOR.json"

            print(f"\n  [ANNOTATE] Injecting %%{{...}}%% blocks into {paper_name}...")
            inject_annotations(
                paper_path=paper_path,
                stage_d_path=str(d_path) if d_path.exists() else None,
                stage_e_path=str(e_path) if e_path.exists() else None,
                stage_f_path=str(f_path) if f_path.exists() else None,
            )
        except ImportError:
            print(f"  [ANNOTATE] annotation_injector.py not found, skipping injection")
        except Exception as e:
            print(f"  [ANNOTATE] Injection failed: {e}")

    return {
        "paper": paper_name,
        "tools": final_results,
        "total_cost_usd": round(total_cost, 6),
        "total_tokens_in": total_in,
        "total_tokens_out": total_out,
        "wall_clock_s": round(total_elapsed, 1),
        "sequential_s": round(seq_time, 1),
    }


async def run_folder(api_key: str, folder: Path, tool_names: list,
                     output_dir: Path, max_concurrent: int = 8) -> list:
    """Run all tools against all .md papers in a folder."""
    papers = sorted(f for f in folder.iterdir()
                    if f.is_file() and f.suffix.lower() == ".md"
                    and not f.name.lower().startswith(("readme", "index")))

    if not papers:
        print(f"No .md papers found in {folder}")
        return []

    print(f"\n{'='*60}")
    print(f"  BATCH MODE — {len(papers)} papers × {len(tool_names)} tools")
    print(f"  = {len(papers) * len(tool_names)} total API calls")
    print(f"{'='*60}")

    all_results = []
    for paper in papers:
        result = await run_paper(api_key, paper, tool_names, output_dir,
                                 max_concurrent)
        all_results.append(result)

    # Grand totals
    grand_cost = sum(r["total_cost_usd"] for r in all_results)
    grand_in = sum(r["total_tokens_in"] for r in all_results)
    grand_out = sum(r["total_tokens_out"] for r in all_results)
    grand_wall = sum(r["wall_clock_s"] for r in all_results)
    grand_seq = sum(r["sequential_s"] for r in all_results)

    print(f"\n{'='*60}")
    print(f"  BATCH COMPLETE")
    print(f"  Papers:     {len(papers)}")
    print(f"  Total cost: ${grand_cost:.4f}")
    print(f"  Tokens:     {grand_in:,} in + {grand_out:,} out")
    print(f"  Wall clock: {grand_wall:.1f}s")
    print(f"  Sequential: {grand_seq:.1f}s ({grand_seq/grand_wall:.1f}x speedup)" if grand_wall > 0 else "")
    print(f"{'='*60}")

    return all_results


# ---------------------------------------------------------------------------
# Excel Assembly (free — no API calls, just collects JSON into workbook)
# ---------------------------------------------------------------------------
def build_excel(output_dir: Path, folder_name: str = "results"):
    """Assemble all JSON results in output_dir into one Excel workbook.
    One sheet per paper, columns for each tool's scores."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  openpyxl not installed — skipping Excel. pip install openpyxl")
        return None

    # Collect all JSON files grouped by paper stem
    json_files = sorted(output_dir.glob("*.json"))
    if not json_files:
        print("  No JSON results found — skipping Excel")
        return None

    # Group by paper: strip the tool suffix to get the paper stem
    tool_suffixes = ["_CKG", "_DOMAINS", "_CDCM", "_7Q", "_YAML", "_SUMMARY"]
    papers = {}  # paper_stem -> {tool_name: parsed_json}

    for jf in json_files:
        stem = jf.stem
        # Skip summary files
        if stem.endswith("_SUMMARY") or stem.startswith("BATCH_SUMMARY"):
            continue

        # Figure out which tool produced this
        tool_name = None
        paper_stem = stem
        for suffix in tool_suffixes:
            if stem.endswith(suffix):
                tool_name = suffix.lstrip("_").lower()
                paper_stem = stem[: -len(suffix)]
                break

        if not tool_name:
            continue

        try:
            data = json.loads(jf.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, UnicodeDecodeError):
            continue

        if paper_stem not in papers:
            papers[paper_stem] = {}
        papers[paper_stem][tool_name] = data

    if not papers:
        print("  No parseable results — skipping Excel")
        return None

    # Build workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    # Style definitions
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    score_high = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    score_mid = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    score_low = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def score_fill(val):
        """Color-code a 0-100 score."""
        if isinstance(val, (int, float)):
            if val >= 70:
                return score_high
            elif val >= 40:
                return score_mid
            else:
                return score_low
        return None

    # --- OVERVIEW sheet (all papers, one row each) ---
    ws_overview = wb.create_sheet("Overview")
    overview_headers = ["Paper", "CKG Raw", "CKG Final", "CDCM Score",
                        "Dominant Domain", "Tokens In", "Tokens Out", "Cost USD"]
    for col, h in enumerate(overview_headers, 1):
        cell = ws_overview.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    for row_idx, (paper_stem, tool_data) in enumerate(sorted(papers.items()), 2):
        ws_overview.cell(row=row_idx, column=1, value=paper_stem).border = thin_border

        # CKG
        ckg = tool_data.get("ckg", {})
        agg = ckg.get("document_aggregate", {})
        raw = agg.get("raw_score", "")
        final = agg.get("final_score", "")
        c = ws_overview.cell(row=row_idx, column=2, value=raw)
        c.border = thin_border
        f = score_fill(raw)
        if f:
            c.fill = f
        ws_overview.cell(row=row_idx, column=3, value=final).border = thin_border

        # CDCM
        cdcm = tool_data.get("cdcm", {})
        cdcm_score = cdcm.get("weighted_average", cdcm.get("overall_score", ""))
        c = ws_overview.cell(row=row_idx, column=4, value=cdcm_score)
        c.border = thin_border
        f = score_fill(cdcm_score)
        if f:
            c.fill = f

        # Domain
        dom = tool_data.get("domains", {})
        ws_overview.cell(row=row_idx, column=5,
                         value=dom.get("dominant_domain", "")).border = thin_border

        # Metadata (from any tool that has it)
        meta = {}
        for td in tool_data.values():
            if isinstance(td, dict) and "_metadata" in td:
                meta = td["_metadata"]
                break
        ws_overview.cell(row=row_idx, column=6,
                         value=meta.get("tokens_in", "")).border = thin_border
        ws_overview.cell(row=row_idx, column=7,
                         value=meta.get("tokens_out", "")).border = thin_border
        ws_overview.cell(row=row_idx, column=8,
                         value=meta.get("cost_usd", "")).border = thin_border

    # Auto-width overview
    for col in range(1, len(overview_headers) + 1):
        ws_overview.column_dimensions[get_column_letter(col)].width = 18

    # --- Per-paper detail sheets ---
    for paper_stem, tool_data in sorted(papers.items()):
        # Sheet name max 31 chars
        sheet_name = paper_stem[:31]
        ws = wb.create_sheet(sheet_name)
        row = 1

        # Title
        ws.cell(row=row, column=1, value=paper_stem).font = Font(bold=True, size=14)
        row += 2

        # CKG section
        ckg = tool_data.get("ckg", {})
        if ckg:
            ws.cell(row=row, column=1, value="CKG EPISTEMIC STRUCTURE").font = header_font
            row += 1
            agg = ckg.get("document_aggregate", {})
            for key in ["raw_score", "final_score", "tier1_foundations",
                        "tier2_propositions", "tier3_constraints",
                        "tier4_evidence", "tier5_integration"]:
                val = agg.get(key, "")
                label = key.replace("_", " ").title()
                ws.cell(row=row, column=1, value=label).border = thin_border
                c = ws.cell(row=row, column=2, value=val)
                c.border = thin_border
                f = score_fill(val)
                if f:
                    c.fill = f
                row += 1
            row += 1

        # CDCM section
        cdcm = tool_data.get("cdcm", {})
        if cdcm:
            ws.cell(row=row, column=1, value="CDCM ACADEMIC REVIEW").font = header_font
            row += 1

            # Section scores (A-K)
            sections = cdcm.get("section_scores", cdcm.get("sections", {}))
            if isinstance(sections, dict):
                for sec_name, sec_val in sections.items():
                    ws.cell(row=row, column=1, value=sec_name).border = thin_border
                    if isinstance(sec_val, dict):
                        # Has sub-scores
                        for sub_key, sub_val in sec_val.items():
                            ws.cell(row=row, column=1,
                                    value=f"  {sub_key}").border = thin_border
                            c = ws.cell(row=row, column=2, value=sub_val)
                            c.border = thin_border
                            row += 1
                    else:
                        c = ws.cell(row=row, column=2, value=sec_val)
                        c.border = thin_border
                        f = score_fill(sec_val if isinstance(sec_val, (int, float)) else None)
                        if f:
                            c.fill = f
                        row += 1

            # Overall
            overall = cdcm.get("weighted_average", cdcm.get("overall_score", ""))
            if overall:
                ws.cell(row=row, column=1, value="OVERALL SCORE").font = header_font
                c = ws.cell(row=row, column=2, value=overall)
                c.font = Font(bold=True, size=12)
                f = score_fill(overall)
                if f:
                    c.fill = f
                row += 1
            row += 1

        # Domain section
        dom = tool_data.get("domains", {})
        if dom:
            ws.cell(row=row, column=1, value="DOMAIN ACTIVATIONS").font = header_font
            row += 1

            doc_act = dom.get("document_activation", {})
            if isinstance(doc_act, dict):
                # Sort by score descending, show top 15
                sorted_doms = sorted(doc_act.items(), key=lambda x: x[1]
                                     if isinstance(x[1], (int, float)) else 0,
                                     reverse=True)
                for d_name, d_score in sorted_doms[:15]:
                    if isinstance(d_score, (int, float)) and d_score > 0:
                        ws.cell(row=row, column=1, value=d_name).border = thin_border
                        ws.cell(row=row, column=2, value=round(d_score, 3)).border = thin_border
                        row += 1

            dominant = dom.get("dominant_domain", "")
            if dominant:
                ws.cell(row=row, column=1, value="Dominant").font = header_font
                ws.cell(row=row, column=2, value=dominant)
                row += 1

            tensions = dom.get("cross_domain_tensions", [])
            if tensions:
                row += 1
                ws.cell(row=row, column=1, value="Cross-Domain Tensions").font = header_font
                row += 1
                for t in tensions[:5]:
                    if isinstance(t, dict):
                        desc = t.get("tension_description",
                                     t.get("description", str(t)))
                        ws.cell(row=row, column=1, value=desc).border = thin_border
                        row += 1
            row += 1

        # Column widths
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 20

    # Save
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    xlsx_path = output_dir / f"{folder_name}_SCORES_{ts}.xlsx"
    wb.save(xlsx_path)
    print(f"\n  Excel workbook: {xlsx_path}")
    return xlsx_path


# ---------------------------------------------------------------------------
# Scored Paper Assembly — sandwich the original paper with 7Q rigor
# ---------------------------------------------------------------------------
import re

def parse_scores_block(text: str) -> dict:
    """Extract the ```scores block from 7Q output."""
    m = re.search(r'```scores\s*\n(.*?)```', text, re.DOTALL)
    if not m:
        return {}
    scores = {}
    for line in m.group(1).strip().splitlines():
        if ':' in line:
            key, val = line.split(':', 1)
            scores[key.strip()] = val.strip()
    return scores

def parse_7q_sections(text: str) -> dict:
    """Parse the 7Q markdown output into sections by # header."""
    # Strip the ```scores block before parsing
    text = re.sub(r'```scores\s*\n.*?```', '', text, flags=re.DOTALL)
    sections = {}
    current_key = None
    current_lines = []
    for line in text.splitlines():
        if re.match(r'^# [A-Za-z]', line):
            if current_key:
                sections[current_key] = '\n'.join(current_lines).strip()
            current_key = line[2:].strip()
            current_lines = []
        elif current_key:
            current_lines.append(line)
    if current_key:
        sections[current_key] = '\n'.join(current_lines).strip()
    return sections

def extract_kill_conditions(sections: dict) -> list:
    """Pull kill conditions from Primary Failure Points section."""
    text = sections.get('Primary Failure Points', '')
    kills = []
    for line in text.splitlines():
        line = line.strip()
        if line and (line[0].isdigit() or line.startswith('-') or line.startswith('*')):
            clean = re.sub(r'^[\d\.\-\*\s]+', '', line).strip()
            if clean:
                kills.append(clean)
    return kills

def _is_list_start(line: str) -> bool:
    """Check if a line starts a new top-level claim (not a sub-point like Kill Condition)."""
    s = line.strip()
    # Skip sub-points that start with "- " followed by known sub-labels
    if re.match(r'^[-\u2013\u2014\u2010]\s*(Kill|Independence|Dependency|Note|Mapping|Strength)',
                s, re.IGNORECASE):
        return False
    return bool(re.match(r'^(\d+\.|[\u2022]\s|Sub-?claim)', s, re.IGNORECASE))

def _clean_claim_line(line: str) -> str:
    """Strip bullets, numbering, 'Subclaim N:' prefixes from a claim line."""
    s = line.strip()
    # Strip leading bullets/numbers: "• ", "1. ", "- ", "* ", "**"
    s = re.sub(r'^[\d\.\-\*\u2022]+\s*', '', s)
    s = re.sub(r'^\*\*\s*', '', s)
    # Strip "Subclaim N:" or "Sub-claim N:" prefix
    s = re.sub(r'^Sub-?claim\s*\d*\s*[:\.]\s*', '', s, flags=re.IGNORECASE)
    # Strip leading "**text**" bold wrapper
    s = re.sub(r'^\*\*[^*]+\*\*\s*[:\.]\s*', '', s)
    return s.strip()

def extract_claims(sections: dict) -> list:
    """Pull decomposed claims from Claim Decomposition section."""
    text = sections.get('Claim Decomposition', '')
    claims = []
    current = []
    for line in text.splitlines():
        if _is_list_start(line):
            if current:
                claims.append('\n'.join(current).strip())
            current = [line]
        elif current:
            current.append(line)
    if current:
        claims.append('\n'.join(current).strip())
    # Clean the first line of each claim
    cleaned = []
    for c in claims:
        lines = c.splitlines()
        lines[0] = _clean_claim_line(lines[0])
        cleaned.append('\n'.join(lines))
    return cleaned

def extract_theories(sections: dict) -> list:
    """Pull theory mappings from Theory Resonance Map section."""
    text = sections.get('Theory Resonance Map', '')
    theories = []
    current = []
    for line in text.splitlines():
        if _is_list_start(line):
            if current:
                theories.append('\n'.join(current).strip())
            current = [line]
        elif current:
            current.append(line)
    if current:
        theories.append('\n'.join(current).strip())
    return theories

def build_scored_paper(paper_path: Path, output_dir: Path) -> Path:
    """Build the scored paper with YAML + tabbed callouts + paper + detail."""
    stem = paper_path.stem

    # Read the 7Q output
    q7_path = output_dir / f"{stem}_7q.md"
    if not q7_path.exists():
        q7_path = output_dir / f"{stem}_7Q.md"
    if not q7_path.exists():
        print(f"  No 7Q output found for {stem} -- skipping scored paper")
        return None

    q7_text = q7_path.read_text(encoding="utf-8")
    scores = parse_scores_block(q7_text)
    sections = parse_7q_sections(q7_text)
    kills = extract_kill_conditions(sections)
    claims = extract_claims(sections)
    theories = extract_theories(sections)

    # Read CKG score if available
    ckg_score = None
    ckg_path = output_dir / f"{stem}_CKG.json"
    if not ckg_path.exists():
        ckg_path = output_dir / f"{stem}_ckg.md"
    if ckg_path.exists():
        try:
            ckg_raw = ckg_path.read_text(encoding="utf-8")
            ckg_data = json.loads(ckg_raw)
            # Try document_aggregate first
            agg = ckg_data.get("document_aggregate", {})
            if agg.get("final_score"):
                ckg_score = float(agg["final_score"])
            elif agg.get("raw_score"):
                ckg_score = float(agg["raw_score"]) / 100.0 * 10.0
            else:
                # Compute from section scores: sum raw / (sections * 100) * 10
                secs = ckg_data.get("section_scores", [])
                if secs:
                    total_raw = sum(s.get("section_raw_score", 0) for s in secs)
                    max_possible = len(secs) * 100  # 5 tiers * 20 max each
                    ckg_score = round((total_raw / max_possible) * 10, 2) if max_possible > 0 else None
        except json.JSONDecodeError:
            # Truncated JSON — extract section_raw_score values with regex
            raw_scores = re.findall(r'"section_raw_score":\s*(\d+)', ckg_raw)
            if raw_scores:
                total_raw = sum(int(s) for s in raw_scores)
                max_possible = len(raw_scores) * 100
                ckg_score = round((total_raw / max_possible) * 10, 2) if max_possible > 0 else None
        except (ValueError, KeyError):
            pass

    # Read original paper
    paper_text = paper_path.read_text(encoding="utf-8")

    # Strip existing YAML frontmatter from paper if present
    paper_body = paper_text
    existing_yaml = ""
    if paper_text.startswith("---"):
        parts = paper_text.split("---", 2)
        if len(parts) >= 3:
            existing_yaml = parts[1].strip()
            paper_body = parts[2].strip()

    # Score values with defaults
    q0 = scores.get('Q0_POSTURE', '0.00')
    q1 = scores.get('Q1_IDENTITY', '0.00')
    q2 = scores.get('Q2_DOMAIN', '0.00')
    q3 = scores.get('Q3_ASSERTION', '0.00')
    q4 = scores.get('Q4_EVIDENCE', '0.00')
    q5 = scores.get('Q5_DEPENDENCIES', '0.00')
    q6 = scores.get('Q6_CONSEQUENCES', '0.00')
    q7 = scores.get('Q7_FALSIFICATION', '0.00')
    paper_type = scores.get('TYPE', 'unknown')
    confidence = scores.get('CONFIDENCE', 'unknown')
    strongest = scores.get('STRONGEST_Q', 'N/A')
    weakest = scores.get('WEAKEST_Q', 'N/A')
    claim_count = scores.get('CLAIM_COUNT', '0')
    kill_count = scores.get('KILL_COUNT', '0')
    iso_status = scores.get('ISO_STATUS', 'NOT_APPLICABLE')

    # Compute T_score (mean of Q0-Q7)
    try:
        q_vals = [float(q0), float(q1), float(q2), float(q3),
                  float(q4), float(q5), float(q6), float(q7)]
        t_score = sum(q_vals) / len(q_vals)
    except ValueError:
        t_score = 0.0

    # CKG display string
    ckg_display = f"{ckg_score:.2f}" if ckg_score is not None else "—"
    ckg_yaml_line = f"\nckg_score: {ckg_score:.2f}" if ckg_score is not None else ""

    # --- ZONE 0: YAML Frontmatter ---
    yaml_block = f"""---
title: "{stem.replace('_', ' ')}"
scored: true
scored_date: "{datetime.now().strftime('%Y-%m-%d')}"
type: "{paper_type}"
confidence: "{confidence}"
iso_status: "{iso_status}"
t_score: {t_score:.3f}{ckg_yaml_line}
claim_count: {claim_count}
kill_count: {kill_count}
strongest: "{strongest}"
weakest: "{weakest}"
7q_scores:
  Q0_POSTURE: {q0}
  Q1_IDENTITY: {q1}
  Q2_DOMAIN: {q2}
  Q3_ASSERTION: {q3}
  Q4_EVIDENCE: {q4}
  Q5_DEPENDENCIES: {q5}
  Q6_CONSEQUENCES: {q6}
  Q7_FALSIFICATION: {q7}
---"""

    # --- ZONE 1: Extract section data ---
    core_claim = sections.get('Core Claim', 'N/A')
    best_surviving = sections.get('Best Surviving Version', 'N/A')
    type_domain = sections.get('Type / Domain', 'N/A')
    robustness = sections.get('Robustness Estimate', 'N/A')
    dependencies = sections.get('Upstream Dependencies', 'N/A')
    evidence_audit = sections.get('Evidence Independence Audit', 'N/A')
    downstream = sections.get('Downstream Consequences', 'N/A')
    what_survives = sections.get('What Survives', 'N/A')
    what_dies = sections.get('What Dies', 'N/A')
    sci_extract = sections.get('Scientific Extraction', 'N/A')
    decisive = sections.get('Decisive Untested Prediction', 'N/A')
    exec_summary = sections.get('Executive Summary', 'N/A')
    kg_yaml_raw = sections.get('Knowledge Graph YAML', '')
    kg_yaml = re.sub(r'```ya?ml\s*\n?', '', kg_yaml_raw)
    kg_yaml = re.sub(r'```\s*$', '', kg_yaml).strip()

    q_names = ["Posture", "Identity", "Domain", "Assertion",
               "Evidence", "Dependencies", "Consequences", "Falsification"]
    q_scores_list = [q0, q1, q2, q3, q4, q5, q6, q7]

    # Indent helpers for nested callouts (2 levels: "> > ")
    def nest2(text):
        """Indent text for level-2 nesting inside a callout."""
        return '\n'.join('> > ' + l for l in text.strip().splitlines())

    def nest3(text):
        """Indent text for level-3 nesting."""
        return '\n'.join('> > > ' + l for l in text.strip().splitlines())

    # Build claims list for Q1
    claims_block = '\n'.join(
        f'> > {i+1}. {c.splitlines()[0].strip()}'
        for i, c in enumerate(claims[:6])
    )

    # Build kill conditions for Q7
    kill_block = '\n'.join(
        f'> > {i+1}. {k}' for i, k in enumerate(kills)
    )

    # Build theory table for nested callout (level 2)
    theory_table_nested = "> > | Theory | Mapping | Status |\n> > |--------|---------|--------|"
    for t in theories:
        lines = t.strip().splitlines()
        first = re.sub(r'^[\d\.\-\*\u2022]+\s*', '', lines[0].strip())
        # Format A: "Name (STATUS): description"
        m_t = re.match(r'\*?\*?([^(]+?)\*?\*?\s*\((\w+)\)\s*[:—\-–]\s*(.*)', first)
        if m_t:
            th_n = m_t.group(1).strip().rstrip('*')
            th_s = m_t.group(2).strip()
            th_d = m_t.group(3).strip()[:60]
            theory_table_nested += f"\n> > | {th_n} | {th_d} | {th_s} |"
        else:
            # Format B: look for Strength line
            th_n = first
            th_s = ''
            th_d = ''
            for sub in lines[1:]:
                sm = re.search(r'Strength:\s*(STRUCTURAL|ANALOGICAL|NOMINAL)\s*[—–\-]\s*(.*)', sub, re.IGNORECASE)
                if sm:
                    th_s = sm.group(1).upper()
                    th_d = sm.group(2).strip()[:60]
                    break
            if len(th_n) > 40:
                th_n = th_n[:40] + '...'
            theory_table_nested += f"\n> > | {th_n} | {th_d} | {th_s} |"

    # KG YAML for nested callout
    kg_nested = '\n'.join('> > ' + l for l in kg_yaml.splitlines()) if kg_yaml else '> > (none)'

    # --- ZONE 3: Nested Callout Groups (4 styled parents, Q's fold inside) ---
    # Parents use distinct callout types so colors don't collide with children:
    #   q1-define  (gold)    → Identity cluster  (Q0, Q1, Q2)
    #   q4-support (blue)    → Evidence cluster   (Q3, Q4)
    #   q6-propagate (emerald) → Consequences cluster (Q5, Q6)
    #   q7-destroy (red)     → Kill cluster       (Q7, theories, graph)
    detail_block = f"""---

%%
{'='*50}
  7Q DETAILED ANALYSIS -- ENGINE OUTPUT BELOW
{'='*50}
%%

## 7Q Detailed Analysis

> [!verdict] Executive Summary — T = {t_score:.3f} | CKG = {ckg_display}
> {exec_summary}

> [!q1-define]- Identity & Posture — Q0-Q2
> **Type:** {paper_type} | **ISO:** {iso_status} | **Claims:** {claim_count}
>
> > [!q0-arrive]- Q0 — ARRIVE ({q0})
> > **Posture Assessment**
{nest2(robustness)}
>
> > [!q1-define]- Q1 — DEFINE ({q1})
> > **Type:** {paper_type}
> > **Claims:** {claim_count} independently killable claims bundled
{claims_block}
>
> > [!q2-locate]- Q2 — LOCATE ({q2})
> > **Primary:** {type_domain}
> > **ISO Status:** {iso_status}

> [!q4-support]- Assertion & Evidence — Q3-Q4
> **Core Claim:** {core_claim[:200]}{'...' if len(core_claim) > 200 else ''}
>
> > [!q3-commit]- Q3 — COMMIT ({q3})
> > **Core Claim:** {core_claim}
> > **Best Surviving Version:** {best_surviving}
>
> > [!q4-support]- Q4 — SUPPORT ({q4})
> > **Evidence Assessment**
{nest2(evidence_audit)}

> [!q6-propagate]- Foundation & Consequences — Q5-Q6
> **Axiom Floor:** What it rests on and where it goes
>
> > [!q5-ground]- Q5 — GROUND ({q5})
> > **Upstream Dependencies**
{nest2(dependencies)}
>
> > [!q6-propagate]- Q6 — PROPAGATE ({q6})
> > **Downstream Consequences**
{nest2(downstream)}
> >
> > **Decisive Untested Prediction:**
{nest2(decisive)}

> [!q7-destroy]- Kill Analysis — Q7
> {kill_count} kill conditions | **What survives vs. what dies**
>
> > [!q7-destroy]- Q7 — DESTROY ({q7})
> > **Kill Conditions:**
{kill_block}
> > **What Survives:** {what_survives}
> > **What Dies:** {what_dies}
>
> > [!theory-map]- Theory Resonance Map
{theory_table_nested}
>
> > [!graph]- Knowledge Graph YAML
> > ```yaml
{kg_nested}
> > ```

---

*Scored by 7Q Engine v1.0 | o3-mini | {datetime.now().strftime('%B %d, %Y')} | ~$0.03*
*CSS: 7q-infobox.css + 7q-scored-callouts.css*
"""

    # --- ZONE 0.5: Wikipedia-style Infobox (floats right via CSS) ---
    # Build a clean title from the paper stem
    paper_title = stem.replace('_', ' ')
    # Remove common prefixes like "FP-008 " for display
    display_title = re.sub(r'^FP-?\d+\s*', '', paper_title).strip()
    if not display_title:
        display_title = paper_title

    # Confidence label formatting
    conf_label = confidence.replace('_', ' ').title()

    # Build theory resonance table for infobox
    # Supports two formats:
    #   A) "Name (STATUS): description"
    #   B) Multi-line: "1. Name\n   - Strength: STRUCTURAL – ..."
    infobox_theory_rows = ""
    for t in theories:
        lines = t.strip().splitlines()
        first = re.sub(r'^[\d\.\-\*\u2022]+\s*', '', lines[0].strip())
        # Try format A: "Name (STATUS): description"
        m_th = re.match(r'\*?\*?([^(]+?)\*?\*?\s*\((\w+)\)', first)
        if m_th:
            th_name = m_th.group(1).strip().rstrip('*')
            th_status = m_th.group(2).strip().upper()
        else:
            # Format B: look for "Strength: STRUCTURAL" in sub-lines
            th_name = first.strip()
            th_status = ''
            for sub in lines[1:]:
                sm = re.search(r'Strength:\s*(STRUCTURAL|ANALOGICAL|NOMINAL)', sub, re.IGNORECASE)
                if sm:
                    th_status = sm.group(1).upper()
                    break
        tag = f"**{th_status}**" if th_status == "STRUCTURAL" else th_status
        infobox_theory_rows += f"> | {th_name} | {tag} |\n"

    # Count structural vs analogical
    structural_count = 0
    analogical_count = 0
    for t in theories:
        if re.search(r'STRUCTURAL', t, re.IGNORECASE):
            structural_count += 1
        elif re.search(r'ANALOGICAL', t, re.IGNORECASE):
            analogical_count += 1
    theory_summary = ""
    if structural_count or analogical_count:
        parts = []
        if structural_count:
            parts.append(f"{structural_count} structural mapping{'s' if structural_count != 1 else ''} carry real predictive weight")
        if analogical_count:
            parts.append(f"{analogical_count} analogical mapping{'s' if analogical_count != 1 else ''} suggest pattern but don't transfer equations")
        theory_summary = ". ".join(parts) + "."

    # Strongest / weakest Q formatting
    q_labels = {
        'Q0': 'Posture', 'Q1': 'Identity', 'Q2': 'Domain', 'Q3': 'Assertion',
        'Q4': 'Evidence', 'Q5': 'Dependencies', 'Q6': 'Consequences', 'Q7': 'Falsification'
    }
    strongest_q_num = re.match(r'(Q\d)', strongest) if strongest else None
    weakest_q_num = re.match(r'(Q\d)', weakest) if weakest else None
    strongest_label = q_labels.get(strongest_q_num.group(1), '') if strongest_q_num else ''
    weakest_label = q_labels.get(weakest_q_num.group(1), '') if weakest_q_num else ''
    strongest_score = q_scores_list[int(strongest_q_num.group(1)[1])] if strongest_q_num else ''
    weakest_score = q_scores_list[int(weakest_q_num.group(1)[1])] if weakest_q_num else ''
    # Strip "Q3 — " prefix from the explanation text (we already show Q3 in the header)
    strongest_text = re.sub(r'^Q\d\s*[—\-–:]\s*', '', strongest).strip() if strongest else ''
    weakest_text = re.sub(r'^Q\d\s*[—\-–:]\s*', '', weakest).strip() if weakest else ''

    # Most dangerous kill condition (first one, or from the 7Q output)
    most_dangerous_kill = kills[0] if kills else "Not identified"
    if len(most_dangerous_kill) > 120:
        most_dangerous_kill = most_dangerous_kill[:120] + '...'

    # Extract domain count from Type/Domain section
    domain_text = sections.get('Type / Domain', '')
    domain_count_match = re.search(r'(\d+)\s*(?:domains?|fields?)', domain_text, re.IGNORECASE)
    domain_count = domain_count_match.group(1) if domain_count_match else '—'

    # Extract effective n from Evidence Independence Audit
    eff_n_match = re.search(r'effective\s+n\s*[=:≈~]\s*~?(\d+)', evidence_audit, re.IGNORECASE)
    effective_n = f"~{eff_n_match.group(1)}" if eff_n_match else '—'

    # Decisive test (truncated for infobox)
    decisive_short = decisive.strip().split('\n')[0] if decisive.strip() != 'N/A' else '—'
    if len(decisive_short) > 80:
        decisive_short = decisive_short[:80] + '...'

    # ISO upgrade hint
    iso_hints = {
        'ISO-PARALLEL': 'Pattern matches but mathematical mapping not formally demonstrated',
        'ISO-CONFIRMED': 'Structural identity proven — equations transfer',
        'ANALOGICAL': 'Qualitative pattern match; equations do not transfer directly',
        'STRUCTURAL': 'Same equations govern both domains',
        'NOT_APPLICABLE': 'No cross-domain mapping claimed',
    }
    iso_hint = iso_hints.get(iso_status, 'Classification pending')

    # Core claim truncated for infobox
    core_claim_short = core_claim.strip().split('\n')[0]
    if len(core_claim_short) > 200:
        core_claim_short = core_claim_short[:200] + '...'

    infobox_block = f"""> [!7q-infobox]- 7Q Rigor Card
>
> ## {display_title}
>
> **{conf_label}** | T = **{t_score:.3f}** | CKG = **{ckg_display}**
>
> ### Core Claim
>
> {core_claim_short}
>
> ### Theory Resonance
>
> | . | . |
> |---|---|
{infobox_theory_rows}>
> {theory_summary}
>
> ### Strongest
>
> **{strongest_q_num.group(1) if strongest_q_num else '—'} {strongest_label}** ({strongest_score}) — {strongest_text}
>
> ### Weakest
>
> **{weakest_q_num.group(1) if weakest_q_num else '—'} {weakest_label}** ({weakest_score}) — {weakest_text}
>
> ### Kill Threat
>
> {kill_count} kill conditions | **Most dangerous:**
> {most_dangerous_kill}
>
> ### ISO Status
>
> **{iso_status}** — {iso_hint}
>
> ### Bundled Claims
>
> **{claim_count}** independently killable claims in one paper.
>
> ### At A Glance
>
> | . | . |
> |---|---|
> | Type | {paper_type.replace('_', ' ').title()} |
> | Domains | {domain_count} |
> | Effective n | {effective_n} |
> | Decisive test | {decisive_short} |
>
> ### Scores
>
> | . | . |
> |---|---|
> | Q0 Posture | {q0} |
> | Q1 Identity | {q1} |
> | Q2 Domain | {q2} |
> | Q3 Assertion | {q3} |
> | Q4 Evidence | {q4} |
> | Q5 Dependencies | {q5} |
> | Q6 Consequences | {q6} |
> | Q7 Falsification | {q7} |
> | **T (7Q)** | **{t_score:.3f}** |
> | **CKG** | **{ckg_display}** |
"""

    # --- ZONE 0.75: Extract FACTS At A Glance from paper body ---
    # Pull the FACTS table out of the body and place it as a top callout
    facts_callout = ""
    facts_pattern = re.search(
        r'(##\s*FACTS\s+At\s+A\s+Glance.*?)(?=\n---|\n##\s+[^F]|\Z)',
        paper_body, re.DOTALL | re.IGNORECASE
    )
    if facts_pattern:
        facts_raw = facts_pattern.group(1).strip()
        # Extract just the table rows from the FACTS section
        table_lines = re.findall(r'(\|.*\|)', facts_raw)
        if len(table_lines) >= 3:  # header + separator + at least 1 row
            facts_table = '\n'.join(f'> {line}' for line in table_lines)
            facts_callout = f"""> [!faq]- FACTS At A Glance
>
{facts_table}
"""
        # Remove the FACTS section from paper_body so it doesn't appear twice
        paper_body = paper_body[:facts_pattern.start()] + paper_body[facts_pattern.end():]
        # Clean up any resulting double --- separators
        paper_body = re.sub(r'\n---\s*\n---', '\n---', paper_body)

    # --- ASSEMBLE THE SANDWICH ---
    scored_paper = f"""{yaml_block}

{infobox_block}

{facts_callout}

{paper_body}

{detail_block}
"""

    # Write scored paper
    scored_path = output_dir / f"{stem}_SCORED.md"
    scored_path.write_text(scored_paper, encoding="utf-8")
    print(f"\n  Scored paper: {scored_path}")
    return scored_path


# ---------------------------------------------------------------------------
# Batch API (50% discount, 24h turnaround)
# ---------------------------------------------------------------------------
async def submit_batch(api_key: str, paper_path: Path, tool_names: list,
                       output_dir: Path):
    """Submit all calls via OpenAI Batch API for 50% cost reduction."""
    from openai import AsyncOpenAI

    client = AsyncOpenAI(api_key=api_key)
    paper_text = paper_path.read_text(encoding="utf-8")
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    # Build JSONL batch file
    batch_requests = []
    for tool_name in tool_names:
        if tool_name not in TOOLS:
            continue
        tool = TOOLS[tool_name]
        prompt_path = SCRIPT_DIR / tool.prompt_file
        if not prompt_path.exists():
            continue

        system_prompt = prompt_path.read_text(encoding="utf-8")

        request = {
            "custom_id": f"{paper_path.stem}_{tool_name}",
            "method": "POST",
            "url": "/v1/chat/completions",
            "body": {
                "model": tool.model,
                "messages": [
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": paper_text},
                ],
                "max_tokens": tool.max_tokens,
            }
        }
        if not tool.model.startswith("o"):
            request["body"]["temperature"] = tool.temperature

        batch_requests.append(request)

    # Write JSONL
    jsonl_path = output_dir / f"batch_{paper_path.stem}_{ts}.jsonl"
    with open(jsonl_path, "w", encoding="utf-8") as f:
        for req in batch_requests:
            f.write(json.dumps(req) + "\n")

    print(f"  Batch JSONL written: {jsonl_path.name} ({len(batch_requests)} requests)")

    # Upload and submit
    with open(jsonl_path, "rb") as f:
        batch_file = await client.files.create(file=f, purpose="batch")

    batch = await client.batches.create(
        input_file_id=batch_file.id,
        endpoint="/v1/chat/completions",
        completion_window="24h",
        metadata={"paper": paper_path.stem, "tools": ",".join(tool_names)},
    )

    print(f"  Batch submitted: {batch.id}")
    print(f"  Status: {batch.status}")
    print(f"  Check with: python run_all.py --check-batch {batch.id}")

    # Save batch ID for later retrieval
    batch_record = output_dir / f"batch_{paper_path.stem}_{ts}_ID.txt"
    batch_record.write_text(f"{batch.id}\n{jsonl_path.name}\n{','.join(tool_names)}")
    print(f"  Batch ID saved: {batch_record.name}")

    return batch.id


async def check_batch(api_key: str, batch_id: str, output_dir: Path):
    """Check batch status and download results if complete."""
    from openai import AsyncOpenAI

    client = AsyncOpenAI(api_key=api_key)
    batch = await client.batches.retrieve(batch_id)

    print(f"  Batch: {batch_id}")
    print(f"  Status: {batch.status}")
    print(f"  Created: {batch.created_at}")

    if batch.status == "completed":
        print(f"  Completed! Downloading results...")

        result_file = await client.files.content(batch.output_file_id)
        results_path = output_dir / f"batch_results_{batch_id[:12]}.jsonl"
        results_path.write_bytes(result_file.content)

        # Parse results
        count = 0
        for line in results_path.read_text().splitlines():
            result = json.loads(line)
            custom_id = result["custom_id"]
            response = result["response"]["body"]
            reply = response["choices"][0]["message"]["content"]

            out_path = output_dir / f"{custom_id}_BATCH.md"
            out_path.write_text(reply, encoding="utf-8")
            count += 1

        print(f"  Downloaded {count} results to {output_dir}")

    elif batch.status == "failed":
        print(f"  FAILED")
        if batch.errors:
            for err in batch.errors.data:
                print(f"    {err.code}: {err.message}")

    else:
        completed = batch.request_counts.completed if batch.request_counts else 0
        total = batch.request_counts.total if batch.request_counts else 0
        print(f"  Progress: {completed}/{total}")
        print(f"  (Check again later)")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(
        description="Run all OpenAI scoring tools in parallel"
    )
    parser.add_argument("input", nargs="?",
                        help="Paper file (.md) or folder of papers")
    parser.add_argument("--tools", type=str, default=None,
                        help=f"Comma-separated tool list. Available: {','.join(TOOLS.keys())}. "
                             f"Default: {','.join(DEFAULT_TOOLS)}")
    parser.add_argument("--output", type=str, default=None,
                        help="Output directory (default: OpenAI_DATA/ next to input)")
    parser.add_argument("--concurrent", type=int, default=8,
                        help="Max concurrent API calls (default: 8)")
    parser.add_argument("--dry-run", action="store_true",
                        help="Show plan without making API calls")
    parser.add_argument("--batch", action="store_true",
                        help="Submit via Batch API (50%% off, 24h turnaround)")
    parser.add_argument("--check-batch", type=str, default=None,
                        help="Check status of a batch job by ID")
    parser.add_argument("--list-tools", action="store_true",
                        help="List available tools and exit")

    args = parser.parse_args()

    # List tools
    if args.list_tools:
        print(f"\nAvailable tools:")
        for name, tool in TOOLS.items():
            default = " (default)" if name in DEFAULT_TOOLS else ""
            print(f"  {name:12s}  {tool.model:15s}  {tool.description}{default}")
        return

    # API key
    api_key = load_api_key()

    # Check batch status
    if args.check_batch:
        if not api_key:
            sys.exit("ERROR: No API key found in any config.txt")
        out_dir = Path(args.output) if args.output else Path(".")
        asyncio.run(check_batch(api_key, args.check_batch, out_dir))
        return

    if not args.input:
        parser.print_help()
        return

    input_path = Path(args.input)
    if not input_path.exists():
        sys.exit(f"ERROR: Not found: {input_path}")

    # Tools
    tool_names = args.tools.split(",") if args.tools else DEFAULT_TOOLS
    bad = [t for t in tool_names if t not in TOOLS]
    if bad:
        sys.exit(f"ERROR: Unknown tools: {bad}. Available: {list(TOOLS.keys())}")

    # Output dir
    if args.output:
        output_dir = Path(args.output)
    elif input_path.is_file():
        output_dir = input_path.parent / "OpenAI_DATA"
    else:
        output_dir = input_path / "OpenAI_DATA"
    output_dir.mkdir(parents=True, exist_ok=True)

    # Dry run
    if args.dry_run:
        if input_path.is_file():
            papers = [input_path]
        else:
            papers = sorted(f for f in input_path.iterdir()
                            if f.is_file() and f.suffix.lower() == ".md"
                            and not f.name.lower().startswith(("readme", "index")))

        print(f"\nDRY RUN")
        print(f"  Papers: {len(papers)}")
        print(f"  Tools:  {tool_names}")
        print(f"  Calls:  {len(papers) * len(tool_names)}")
        print(f"  Output: {output_dir}")

        # Estimate cost
        total_chars = sum(p.read_text(encoding="utf-8").__len__() for p in papers)
        est_in_tokens = total_chars // 4 * len(tool_names)
        est_out_tokens = 4096 * len(tool_names) * len(papers)

        print(f"\n  Estimated input tokens:  ~{est_in_tokens:,}")
        print(f"  Estimated output tokens: ~{est_out_tokens:,}")

        # Per-model cost
        models_used = set(TOOLS[t].model for t in tool_names)
        total_cost = 0
        for m in models_used:
            if m in PRICING:
                tools_for_model = [t for t in tool_names if TOOLS[t].model == m]
                m_in = total_chars // 4 * len(tools_for_model)
                m_out = 4096 * len(tools_for_model) * len(papers)
                in_rate, out_rate = PRICING[m]
                cost = (m_in / 1000) * in_rate + (m_out / 1000) * out_rate
                total_cost += cost
                print(f"  {m}: ${cost:.4f}")

        print(f"\n  ESTIMATED TOTAL: ${total_cost:.4f}")
        if args.batch:
            print(f"  WITH BATCH API:  ${total_cost * 0.5:.4f} (50% off)")

        for p in papers:
            print(f"\n  {p.name}")
            for t in tool_names:
                tool = TOOLS[t]
                print(f"    [{t}] -> {tool.model} -> {p.stem}{tool.output_suffix}")
        return

    # Need API key for real runs
    if not api_key:
        sys.exit("ERROR: No API key found. Add OPENAI_API_KEY to any config.txt")

    # Batch mode
    if args.batch:
        if input_path.is_file():
            asyncio.run(submit_batch(api_key, input_path, tool_names, output_dir))
        else:
            papers = sorted(f for f in input_path.iterdir()
                            if f.is_file() and f.suffix.lower() == ".md"
                            and not f.name.lower().startswith(("readme", "index")))
            for paper in papers:
                asyncio.run(submit_batch(api_key, paper, tool_names, output_dir))
        return

    # Real async run
    if input_path.is_file():
        result = asyncio.run(
            run_paper(api_key, input_path, tool_names, output_dir, args.concurrent)
        )
        # Save summary
        summary_path = output_dir / f"{input_path.stem}_SUMMARY.json"
        summary_path.write_text(json.dumps(result, indent=2), encoding="utf-8")
        print(f"\n  Summary: {summary_path}")

        # Build Excel (free, no API)
        build_excel(output_dir, input_path.stem)

        # Build scored paper if 7Q was run
        if "7q" in tool_names:
            build_scored_paper(input_path, output_dir)
    else:
        results = asyncio.run(
            run_folder(api_key, input_path, tool_names, output_dir, args.concurrent)
        )
        # Save batch summary
        summary_path = output_dir / f"BATCH_SUMMARY_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        summary_path.write_text(json.dumps(results, indent=2), encoding="utf-8")
        print(f"\n  Batch summary: {summary_path}")

        # Build Excel (free, no API)
        build_excel(output_dir, input_path.name)

        # Build scored papers if 7Q was run
        if "7q" in tool_names:
            papers = sorted(f for f in input_path.iterdir()
                            if f.is_file() and f.suffix.lower() == ".md"
                            and not f.name.lower().startswith(("readme", "index")))
            for paper in papers:
                build_scored_paper(paper, output_dir)


if __name__ == "__main__":
    main()

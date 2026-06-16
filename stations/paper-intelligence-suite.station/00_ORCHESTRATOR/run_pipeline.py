"""
MASTER ORCHESTRATOR
===================
Runs all 7 layers on a paper or series.
Outputs: Excel workbook + L4 7Q files into vault.

Usage:
  python run_pipeline.py --paper "path/to/paper.md"
  python run_pipeline.py --series "path/to/series/" --output "path/to/output/"
"""
import sys, json, re, argparse, hashlib
from pathlib import Path
from datetime import datetime

# Bump when scoring formulas, layer columns, or aggregation logic change.
# Lets you tell two runs apart even if file names match.
SCHEMA_VERSION = "2026.04.07-B"  # B: L6 upgraded to truth_coherence_scanner (fruits + anti-fruits + character)

# Add backend to path
BACKEND = Path(r"O:\999_IGNORE\Obsidian Programs\Python_Backend")
sys.path.insert(0, str(BACKEND))
sys.path.insert(0, str(BACKEND / "analytics"))
sys.path.insert(0, str(BACKEND / "core"))

SUITE_DIR = Path(__file__).resolve().parent.parent

# Suite root on sys.path for `from lib.snapshot_*` imports (L13 + snapshot save)
sys.path.insert(0, str(SUITE_DIR))

sys.path.insert(0, str(SUITE_DIR / "01_TEXT_ANALYTICS"))
sys.path.insert(0, str(SUITE_DIR / "02_ACADEMIC_STANDARD"))
sys.path.insert(0, str(SUITE_DIR / "03_THEOPHYSICS_METRICS"))
sys.path.insert(0, str(SUITE_DIR / "04_OPENAI_7Q"))
sys.path.insert(0, str(SUITE_DIR / "05_NLP_DEEP"))
sys.path.insert(0, str(SUITE_DIR / "06_TRUTH_ENGINE"))
sys.path.insert(0, str(SUITE_DIR / "07_KNOWLEDGE_GRAPHS"))
sys.path.insert(0, str(SUITE_DIR / "08_EMOTION_PROFILE"))
sys.path.insert(0, str(SUITE_DIR / "09_LINGUISTIC_DEPTH"))
sys.path.insert(0, str(SUITE_DIR / "10_IDEA_DENSITY"))

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False
    print("WARNING: openpyxl not installed. Excel output disabled.")


def log_status(ok: bool, label: str, detail: str = ""):
    prefix = "[OK]" if ok else "[ERROR]"
    if detail:
        print(f"  {prefix} {label}: {detail}")
    else:
        print(f"  {prefix} {label}")


def run_layer(name, fn, *args, **kwargs):
    try:
        result = fn(*args, **kwargs)
        log_status(True, name)
        return result
    except Exception as e:
        log_status(False, name, str(e))
        return {}


def _check_layer_result(name: str, result: dict) -> bool:
    """Detect silent layer failures: a layer returning a dict with an *_error
    field is broken, even if no exception was raised. Returns True if healthy."""
    if not isinstance(result, dict):
        return True
    err_keys = [k for k in result.keys() if k.endswith("_error") or k == "error"]
    if err_keys:
        msgs = "; ".join(f"{k}={result[k]}" for k in err_keys)
        log_status(False, name, msgs)
        return False
    return True


def _stable_paper_id(paper: Path) -> str:
    """Hash the resolved path so the same paper gets the same ID across runs,
    even when re-run from a different cwd."""
    h = hashlib.sha1(str(paper.resolve()).lower().encode("utf-8")).hexdigest()
    return f"P-{h[:12]}"


def _slugify(s: str) -> str:
    s = re.sub(r"[^A-Za-z0-9]+", "_", s).strip("_")
    return s[:60] or "series"


def analyze_paper(paper_path, run_openai=False, vault_output=None,
                  series_id="", run_id="", snapshot_dir=None,
                  identity_overrides=None):
    paper = Path(paper_path)
    print(f"\n{'='*60}")
    print(f"ANALYZING: {paper.name}")
    print(f"{'='*60}")

    row = {
        # ── Identity columns (stable across runs / reorderings) ──
        'paper_id':       _stable_paper_id(paper),
        'file':           paper.name,
        'series_id':      series_id,
        'run_id':         run_id,
        'schema_version': SCHEMA_VERSION,
        'source_path':    str(paper.resolve()),
        'analyzed_at':    datetime.now().isoformat(),
    }
    # Track per-layer health so we know which rows are trustworthy.
    row['_layer_status'] = {}

    def _record(layer_key, name, result):
        ok = _check_layer_result(name, result) if isinstance(result, dict) else False
        row['_layer_status'][layer_key] = "ok" if ok else "error"
        return ok

    # PA: Ground-up paper analyzer (claims, flow, link topology; local only)
    try:
        import paper_analyzer as PA
        r = PA.analyze(str(paper))
        row.update({f"PA_{k}": v for k, v in r.items() if k != 'file'})
        if _record("PA", "PA Paper Analyzer", r):
            log_status(True, "PA Paper Analyzer")
    except Exception as e:
        log_status(False, "PA Paper Analyzer", str(e))
        row['_layer_status']['PA'] = "error"

    # L1: Text Analytics
    try:
        import text_analyzer as L1
        r = L1.analyze(str(paper))
        r.pop('_keyword_list', None)
        r.pop('_bigram_list', None)
        row.update({f"L1_{k}": v for k, v in r.items() if k != 'file'})
        if _record("L1", "L1 Text Analytics", r):
            log_status(True, "L1 Text Analytics")
    except Exception as e:
        log_status(False, "L1 Text Analytics", str(e))
        row['_layer_status']['L1'] = "error"

    # L2: Academic Standard
    try:
        import academic_scorer as L2
        r = L2.analyze(str(paper))
        row.update({f"L2_{k}": v for k, v in r.items() if k != 'file'})
        if _record("L2", "L2 Academic Standard", r):
            log_status(True, "L2 Academic Standard")
    except Exception as e:
        log_status(False, "L2 Academic Standard", str(e))
        row['_layer_status']['L2'] = "error"

    # L3: Theophysics Metrics
    try:
        import theophysics_scorer as L3
        r = L3.analyze(str(paper))
        row.update({f"L3_{k}": v for k, v in r.items() if k != 'file'})
        if _record("L3", "L3 Theophysics Metrics", r):
            log_status(True, "L3 Theophysics Metrics")
    except Exception as e:
        log_status(False, "L3 Theophysics Metrics", str(e))
        row['_layer_status']['L3'] = "error"

    # L4: OpenAI 7Q (optional, costs money)
    if run_openai:
        try:
            import seven_q_runner as L4
            r = L4.run_paper(str(paper), vault_output)
            verdict = r.get('reverse_7q', {}).get('verdict', '')
            confidence = r.get('reverse_7q', {}).get('confidence_score', '')
            row['L4_7q_verdict'] = verdict
            row['L4_7q_confidence'] = confidence
            row['L4_7q_file'] = 'see vault _7Q_ANALYSIS folder'
            log_status(True, "L4 OpenAI 7Q")
            row['_layer_status']['L4'] = "ok"
        except Exception as e:
            log_status(False, "L4 OpenAI 7Q", str(e))
            row['_layer_status']['L4'] = "error"
    else:
        row['L4_7q_verdict'] = 'not run (use --openai flag)'
        row['_layer_status']['L4'] = "skipped"
        print(f"  - L4 OpenAI 7Q: skipped")

    # L5: NLP Deep
    try:
        import nlp_analyzer as L5
        r = L5.analyze(str(paper))
        row.update({f"L5_{k}": v for k, v in r.items() if k != 'file'})
        if _record("L5", "L5 NLP Deep", r):
            log_status(True, "L5 NLP Deep")
    except Exception as e:
        log_status(False, "L5 NLP Deep", str(e))
        row['_layer_status']['L5'] = "error"

    # L6: Truth Engine (requires sentence-transformers, slow first run)
    try:
        import truth_runner as L6
        r = L6.score_paper(str(paper))
        row.update({f"L6_{k}": v for k, v in r.items() if k != 'file'})
        if _record("L6", "L6 Truth Engine", r):
            log_status(True, "L6 Truth Engine")
    except Exception as e:
        log_status(False, "L6 Truth Engine", str(e))
        row['_layer_status']['L6'] = "error"

    # L8: Emotion Profile (NRCLex + GoEmotions 27-emotion → Fruits mapping)
    try:
        import emotion_analyzer as L8
        r = L8.analyze(str(paper))
        row.update({f"L8_{k}": v for k, v in r.items()})
        if _record("L8", "L8 Emotion Profile", r):
            log_status(True, "L8 Emotion Profile")
    except Exception as e:
        log_status(False, "L8 Emotion Profile", str(e))
        row['_layer_status']['L8'] = "error"

    # L9: Linguistic Depth (textdescriptives + lexicalrichness)
    try:
        import linguistic_analyzer as L9
        r = L9.analyze(str(paper))
        row.update({f"L9_{k}": v for k, v in r.items()})
        if _record("L9", "L9 Linguistic Depth", r):
            log_status(True, "L9 Linguistic Depth")
    except Exception as e:
        log_status(False, "L9 Linguistic Depth", str(e))
        row['_layer_status']['L9'] = "error"

    # L10: Idea Density (propositional density)
    try:
        import idea_density_analyzer as L10
        r = L10.analyze(str(paper))
        row.update({f"L10_{k}": v for k, v in r.items()})
        if _record("L10", "L10 Idea Density", r):
            log_status(True, "L10 Idea Density")
    except Exception as e:
        log_status(False, "L10 Idea Density", str(e))
        row['_layer_status']['L10'] = "error"

    # L13: OpenAI Peer-Review Prompt Library (10 prompts → snapshot sections)
    # Honors --openai flag for cost control; skips gracefully without API key.
    sections = None
    if run_openai:
        if not __import__('os').environ.get('OPENAI_API_KEY'):
            log_status(False, "L13 Peer Review", "OPENAI_API_KEY not set — skipped")
            row['_layer_status']['L13'] = "skipped"
        else:
            try:
                from openai import OpenAI
                from prompts import run_all as _run_all
                content = paper.read_text(encoding='utf-8', errors='ignore')
                client = OpenAI()
                sections = _run_all(content, client)
                ok_count = sum(1 for v in sections.values()
                               if isinstance(v, dict) and 'error' not in v)
                err_count = len(sections) - ok_count
                detail = f"{ok_count}/{len(sections)} sections ok"
                if err_count:
                    detail += f", {err_count} errored"
                log_status(err_count == 0, "L13 Peer Review", detail)
                row['_layer_status']['L13'] = "ok" if err_count == 0 else "partial"
                row['L13_sections_ok'] = ok_count
                row['L13_sections_err'] = err_count
            except Exception as e:
                log_status(False, "L13 Peer Review", str(e))
                row['_layer_status']['L13'] = "error"
    else:
        row['_layer_status']['L13'] = "skipped"
        print(f"  - L13 Peer Review: skipped (use --openai to enable)")

    # Build the unified snapshot and save it next to the run output.
    # Always emits, even when L13 was skipped — sections will just be empty.
    try:
        from lib.snapshot_merge import build_snapshot
        pm = {k: v for k, v in row.items()
              if not k.startswith('_') and k not in ('paper_id',)}
        snap = build_snapshot(
            paper_id=row['paper_id'],
            paper_path=str(paper),
            pipeline_metrics=pm,
            sections=sections,
            identity_overrides=identity_overrides,
        )
        # Resolve output dir: explicit arg > paper folder fallback.
        out_dir = Path(snapshot_dir) if snapshot_dir else paper.parent
        out_dir.mkdir(parents=True, exist_ok=True)
        snap_path = out_dir / f"{row['paper_id']}_snapshot.json"
        snap_path.write_text(
            json.dumps(snap.to_dict(), indent=2, ensure_ascii=False, default=str),
            encoding='utf-8',
        )
        row['snapshot_path'] = str(snap_path)
        log_status(True, "Snapshot JSON", snap_path.name)
    except Exception as e:
        log_status(False, "Snapshot JSON", str(e))

    return row


def analyze_series(series_path, output_path=None, run_openai=False):
    series = Path(series_path)
    papers = sorted([f for f in series.glob("*.md")
                     if re.match(r'^\d{2}', f.name) and not f.name.startswith('00')])

    if output_path:
        out_dir = Path(output_path)
    else:
        out_dir = series / "_PAPER_INTELLIGENCE"
    out_dir.mkdir(parents=True, exist_ok=True)

    series_slug = _slugify(series.name)
    run_id = datetime.now().strftime('%Y%m%d_%H%M%S')
    series_id = f"S-{hashlib.sha1(str(series.resolve()).lower().encode()).hexdigest()[:10]}"

    print(f"\nSeries: {series.name}")
    print(f"Papers: {len(papers)}")
    print(f"Output: {out_dir}")
    print(f"Run ID: {run_id}  Schema: {SCHEMA_VERSION}")

    snapshot_dir = out_dir / "snapshots"
    snapshot_dir.mkdir(parents=True, exist_ok=True)

    all_rows = []
    for p in papers:
        vault_out = str(p.parent)
        row = analyze_paper(str(p), run_openai=run_openai, vault_output=vault_out,
                            series_id=series_id, run_id=run_id,
                            snapshot_dir=str(snapshot_dir),
                            identity_overrides={"series": series.name})
        all_rows.append(row)

    # Build Knowledge Graph
    try:
        import graph_builder as L7
        graph_result = L7.build_graph(all_rows, str(out_dir))
        for i, row in enumerate(all_rows):
            g = graph_result.get('node_data', {}).get(row['file'], {})
            row['L7_centrality_within_series'] = g.get('centrality', '')
            row['L7_cluster'] = g.get('cluster', '')
            row['_layer_status']['L7'] = "ok"
        print()
        log_status(True, "L7 Knowledge Graph built")
    except Exception as e:
        print()
        log_status(False, "L7 Knowledge Graph", str(e))
        for row in all_rows:
            row['_layer_status']['L7'] = "error"

    # Export to Excel
    excel_path = None
    if HAS_EXCEL and all_rows:
        excel_path = out_dir / f"{series_slug}_PAPER_INTELLIGENCE_{run_id}.xlsx"
        write_excel(all_rows, excel_path)
        print(f"\n  Excel: {excel_path.name}")
        _write_deep_workbook_safe(all_rows, out_dir / f"{series_slug}_DEEP_TABS_{run_id}.xlsx")

    # JSON backup
    json_path = out_dir / f"{series_slug}_pipeline_results_{run_id}.json"
    json_path.write_text(json.dumps(all_rows, indent=2, default=str), encoding='utf-8')
    print(f"  JSON:  {json_path.name}")

    # Run summary — one-glance health snapshot for this run
    summary = {
        "schema_version": SCHEMA_VERSION,
        "run_id":         run_id,
        "series_id":      series_id,
        "series_name":    series.name,
        "series_path":    str(series.resolve()),
        "output_dir":     str(out_dir.resolve()),
        "paper_count":    len(all_rows),
        "openai_enabled": bool(run_openai),
        "excel":          excel_path.name if excel_path else None,
        "layer_health":   _aggregate_layer_health(all_rows),
        "papers":         [{"paper_id": r.get("paper_id"),
                            "file":     r.get("file"),
                            "status":   r.get("_layer_status", {})}
                           for r in all_rows],
    }
    (out_dir / f"{series_slug}_run_summary_{run_id}.json").write_text(
        json.dumps(summary, indent=2), encoding="utf-8"
    )

    print(f"\n{'='*60}")
    print(f"COMPLETE — {len(all_rows)} papers analyzed")
    print(f"Layer health: {summary['layer_health']}")
    print(f"{'='*60}")
    return all_rows


def _aggregate_layer_health(rows):
    """Count ok/error/skipped per layer across all papers in this run."""
    counts = {}
    for r in rows:
        for layer, status in r.get("_layer_status", {}).items():
            counts.setdefault(layer, {"ok": 0, "error": 0, "skipped": 0})
            counts[layer][status] = counts[layer].get(status, 0) + 1
    return counts


LAYER_COLORS = {
    'PA_': 'E6D6F7',  # lavender
    'L1_': 'D6E4F7',  # blue
    'L2_': 'D6F7D6',  # green
    'L3_': 'F7F0D6',  # gold
    'L4_': 'F7D6D6',  # red/pink
    'L5_': 'EBD6F7',  # purple
    'L6_': 'D6F7F0',  # teal
    'L7_': 'F7E6D6',  # orange
    'L8_': 'F7D6E8',  # pink (emotion)
    'L9_': 'D6E8F7',  # light steel blue (linguistic)
    'L10_': 'E8F7D6', # light lime (idea density)
}

IDENTITY_COLUMNS = [
    'paper_id',
    'file',
    'series_id',
    'run_id',
    'schema_version',
    'source_path',
    'analyzed_at',
    'snapshot_path',
    'layer_status',
]

LAYER_ORDER = ['PA_', 'L1_', 'L2_', 'L3_', 'L4_', 'L5_', 'L6_', 'L7_', 'L8_', 'L9_', 'L10_', 'L13_']

def _flatten_for_excel(rows):
    """openpyxl can't write dicts/lists. Flatten nested values into strings."""
    out = []
    for r in rows:
        r2 = dict(r)
        ls = r2.pop("_layer_status", None)
        if isinstance(ls, dict):
            r2["layer_status"] = ",".join(f"{k}:{v}" for k, v in sorted(ls.items()))
        for key, value in list(r2.items()):
            if isinstance(value, dict):
                r2[key] = json.dumps(value, ensure_ascii=False, default=str)
            elif isinstance(value, (list, tuple, set)):
                r2[key] = " | ".join(str(item) for item in value)
        out.append(r2)
    return out


def _collect_all_keys(rows):
    ordered = []
    seen = set()
    for row in rows:
        for key in row.keys():
            if key not in seen:
                seen.add(key)
                ordered.append(key)
    return ordered


def _ordered_excel_columns(rows):
    all_keys = _collect_all_keys(rows)
    ordered = []
    seen = set()

    def add(key):
        if key in all_keys and key not in seen:
            seen.add(key)
            ordered.append(key)

    for key in IDENTITY_COLUMNS:
        add(key)

    for prefix in LAYER_ORDER:
        for key in all_keys:
            if key.startswith(prefix):
                add(key)

    for key in all_keys:
        add(key)

    return ordered


def _header_fill_for_key(key):
    for prefix, color in LAYER_COLORS.items():
        if key.startswith(prefix):
            return PatternFill(start_color=color, end_color=color, fill_type='solid')
    return None


def _write_deep_workbook_safe(rows, path):
    """Emit the ~20-tab deep workbook. Never fatal — it's an extra view."""
    try:
        from deep_workbook import write_deep_workbook
        write_deep_workbook(rows, path)
        print(f"  Deep tabs: {Path(path).name}")
    except Exception as e:
        print(f"  Deep tabs skipped: {e}")


def write_excel(rows, path):
    rows = _flatten_for_excel(rows)
    if not rows:
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ALL_METRICS"
    columns = _ordered_excel_columns(rows)

    for col, key in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col, value=key)
        fill = _header_fill_for_key(key)
        if fill:
            cell.fill = fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical='center')

    for row_idx, row in enumerate(rows, 2):
        for col, key in enumerate(columns, 1):
            val = row.get(key, '')
            if isinstance(val, float):
                val = round(val, 4)
            ws.cell(row=row_idx, column=col, value=val)

    freeze_col = openpyxl.utils.get_column_letter(min(len(IDENTITY_COLUMNS) + 1, len(columns)))
    ws.freeze_panes = f"{freeze_col}2"
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(columns))}{len(rows) + 1}"

    for idx, key in enumerate(columns, 1):
        sample_values = [str(row.get(key, '')) for row in rows[:5] if row.get(key, '') not in ('', None)]
        width = max([len(key)] + [len(value) for value in sample_values]) + 2
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = min(max(width, 12), 42)

    wb.save(str(path))


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description="Theophysics Paper Intelligence Pipeline")
    parser.add_argument('--paper', help='Single paper .md path')
    parser.add_argument('--series', help='Series folder path')
    parser.add_argument('--output', help='Output directory for Excel and JSON')
    parser.add_argument('--openai', action='store_true', help='Run OpenAI 7Q layer (costs ~$0.02/paper)')
    args = parser.parse_args()

    if args.paper:
        paper = Path(args.paper).expanduser().resolve()
        out_dir = Path(args.output).expanduser() if args.output else paper.parent / "_PAPER_INTELLIGENCE"
        out_dir.mkdir(parents=True, exist_ok=True)
        snapshot_dir = out_dir / "snapshots"
        snapshot_dir.mkdir(parents=True, exist_ok=True)
        run_id = datetime.now().strftime('%Y%m%d_%H%M%S')
        series_id = f"SINGLE-{hashlib.sha1(str(paper).lower().encode('utf-8')).hexdigest()[:10]}"

        row = analyze_paper(
            str(paper),
            run_openai=args.openai,
            vault_output=str(paper.parent),
            series_id=series_id,
            run_id=run_id,
            snapshot_dir=str(snapshot_dir),
            identity_overrides={"series": paper.parent.name},
        )
        all_rows = [row]

        try:
            import graph_builder as L7
            graph_result = L7.build_graph(all_rows, str(out_dir))
            node_data = graph_result.get('node_data', {}).get(row.get('file', ''), {})
            row['L7_centrality_within_series'] = node_data.get('centrality', '')
            row['L7_cluster'] = node_data.get('cluster', '')
            row.setdefault('_layer_status', {})['L7'] = 'ok'
        except Exception as e:
            row.setdefault('_layer_status', {})['L7'] = 'error'
            row['L7_error'] = str(e)

        if HAS_EXCEL:
            excel_path = out_dir / f"{paper.stem}_PAPER_INTELLIGENCE_{run_id}.xlsx"
            write_excel(all_rows, excel_path)
            print(f"Excel: {excel_path}")
            _write_deep_workbook_safe(all_rows, out_dir / f"{paper.stem}_DEEP_TABS_{run_id}.xlsx")

        json_path = out_dir / f"{paper.stem}_pipeline_results_{run_id}.json"
        json_path.write_text(json.dumps(all_rows, indent=2, default=str), encoding='utf-8')
        print(f"JSON:  {json_path}")

        summary = {
            "schema_version": SCHEMA_VERSION,
            "run_id": run_id,
            "series_id": series_id,
            "paper": paper.name,
            "paper_path": str(paper),
            "output_dir": str(out_dir.resolve()),
            "openai_enabled": bool(args.openai),
            "layer_health": _aggregate_layer_health(all_rows),
        }
        summary_path = out_dir / f"{paper.stem}_run_summary_{run_id}.json"
        summary_path.write_text(json.dumps(summary, indent=2), encoding='utf-8')
        print(f"Summary: {summary_path}")
    elif args.series:
        analyze_series(args.series, args.output, run_openai=args.openai)
    else:
        print("Usage:")
        print("  python run_pipeline.py --paper path/to/paper.md")
        print("  python run_pipeline.py --series path/to/series/ [--output path/] [--openai]")

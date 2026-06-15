"""Deep multi-tab Excel writer for the Paper Intelligence Suite.

Takes the list of per-paper result rows assembled by run_pipeline.analyze_paper
(one wide dict per paper, with layer-prefixed keys such as ``L1_*``, ``L8_nrc_*``)
and lays them out across ~20 themed worksheet tabs instead of one mega-sheet.

Design notes
------------
* Tabs are carved by KEY PREFIX, not by hard-coded field names, so new metrics
  added to any analyzer flow in automatically and nothing breaks if a layer is
  missing (its tab just renders empty with a note).
* Nested values are expanded: a dict value becomes ``key.subkey`` columns; a
  list becomes a single joined cell. So L3's ``fruits`` dict fans out to one
  column per fruit.
* One row per paper on every themed tab, with stable identity columns frozen on
  the left. ``Overview`` leads, ``Layer_Health`` and ``ALL_METRICS`` trail.

Standalone (for testing against a results JSON):
    python deep_workbook.py <pipeline_results.json> [out.xlsx]
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

IDENTITY = ["paper_id", "file", "series_id", "run_id"]

# (tab title, header color, key predicate). Order defines tab order.
TAB_SPEC: list[tuple[str, str, object]] = [
    ("PA_Structure",          "E6D6F7", lambda k: k.startswith("PA_")),
    ("L1_Readability",        "D6E4F7", lambda k: k.startswith("L1_")),
    ("L2_Academic",           "D6F7D6", lambda k: k.startswith("L2_")),
    ("L3_CHI_WK_Cross",       "F7F0D6", lambda k: k.startswith("L3_") and not (
        k.startswith("L3_fruit") or k.startswith("L3_anti")
        or k.startswith("L3_me") or k in ("L3_fruits", "L3_anti_fruits"))),
    ("L3_Fruits",             "F7E8C0", lambda k: k.startswith("L3_fruit")
        or k.startswith("L3_anti") or k in ("L3_fruits", "L3_anti_fruits",
        "L3_dominant_fruit", "L3_dominant_anti_fruit")),
    ("L3_MasterEquation",     "F7E0A0", lambda k: k.startswith("L3_me")),
    ("L4_OpenAI_7Q",          "F7D6D6", lambda k: k.startswith("L4_")),
    ("L5_NLP_Deep",           "EBD6F7", lambda k: k.startswith("L5_")),
    ("L6_Truth_Coherence",    "D6F7F0", lambda k: k.startswith("L6_")),
    ("L7_Knowledge_Graph",    "F7E6D6", lambda k: k.startswith("L7_")),
    ("L8_NRC_Emotion",        "F7D6E8", lambda k: k.startswith("L8_nrc")),
    ("L8_GoEmotions",         "F7C6DE", lambda k: k.startswith("L8_emo")),
    ("L8_Fruits_Emotion",     "F7B6D4", lambda k: k.startswith("L8_fruit_emo")),
    ("L8_AntiFruits_Emotion", "E8A6C4", lambda k: k.startswith("L8_anti_emo")),
    ("L9_TextDescriptives",   "D6E8F7", lambda k: k.startswith("L9_td")),
    ("L9_Lexical_Richness",   "C6DEF7", lambda k: k.startswith("L9_lr")),
    ("L10_Idea_Density",      "E8F7D6", lambda k: k.startswith("L10_")),
    ("L13_PeerReview",        "F0E0D0", lambda k: k.startswith("L13_")),
]

# Headline metrics shown on Overview when present (guarded by .get).
HEADLINES = [
    "L1_flesch_reading_ease", "L1_text_standard",
    "L2_academic_score", "L2_rigor_score",
    "L3_chi", "L3_wk_ratio", "L3_fruit_net", "L3_me_avg",
    "L6_coherence", "L6_coherence_score",
    "L8_fruit_emo_net", "L8_emo_dominant",
    "L10_idea_density_mean", "L10_idea_density_level",
]

_INVALID = set(r'[]:*?/\\')
_HEADER_FONT = Font(bold=True)
_WRAP = Alignment(wrap_text=True, vertical="center")


def _sanitize_title(title: str, used: set[str]) -> str:
    clean = "".join("_" if c in _INVALID else c for c in title)[:31] or "Sheet"
    candidate, n = clean, 1
    while candidate.lower() in used:
        n += 1
        suffix = f"_{n}"
        candidate = clean[: 31 - len(suffix)] + suffix
    used.add(candidate.lower())
    return candidate


def _cell(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, float):
        return round(value, 4)
    if isinstance(value, (int, str, bool)):
        return value
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False, default=str)
    if isinstance(value, (list, tuple, set)):
        return " | ".join(str(v) for v in value)
    return str(value)


def _matching_keys(rows: list[dict], predicate) -> list[str]:
    ordered, seen = [], set()
    for row in rows:
        for key in row:
            if key in seen or key.startswith("_") or key in IDENTITY:
                continue
            if predicate(key):
                seen.add(key)
                ordered.append(key)
    return ordered


def _expand_columns(rows: list[dict], keys: list[str]) -> list[tuple[str, object]]:
    """Return [(column_label, getter(row))]; dict-valued keys fan out to subkeys."""
    cols: list[tuple[str, object]] = []
    for key in keys:
        is_dict = any(isinstance(r.get(key), dict) for r in rows)
        if is_dict:
            subkeys: list[str] = []
            sseen = set()
            for r in rows:
                v = r.get(key)
                if isinstance(v, dict):
                    for sk in v:
                        if sk not in sseen:
                            sseen.add(sk)
                            subkeys.append(sk)
            for sk in subkeys:
                cols.append((
                    f"{key}.{sk}",
                    (lambda r, k=key, s=sk: (r.get(k) or {}).get(s, "")),
                ))
        else:
            cols.append((key, (lambda r, k=key: r.get(k, ""))))
    return cols


def _write_sheet(wb, title, columns, rows, color, used_titles, note=""):
    ws = wb.create_sheet(_sanitize_title(title, used_titles))
    id_cols = [(c, (lambda r, k=c: r.get(k, ""))) for c in IDENTITY]
    all_cols = id_cols + columns

    if not columns:
        ws.cell(row=1, column=1, value=note or "no metrics (layer not run / dependency missing)")
        for col, (label, _) in enumerate(id_cols, 1):
            ws.cell(row=3, column=col, value=label).font = _HEADER_FONT
        for ridx, r in enumerate(rows, 4):
            for col, (_, get) in enumerate(id_cols, 1):
                ws.cell(row=ridx, column=col, value=_cell(get(r)))
        return ws

    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    for col, (label, _) in enumerate(all_cols, 1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = _HEADER_FONT
        cell.alignment = _WRAP
        if col > len(id_cols):
            cell.fill = fill
    for ridx, r in enumerate(rows, 2):
        for col, (_, get) in enumerate(all_cols, 1):
            ws.cell(row=ridx, column=col, value=_cell(get(r)))

    ws.freeze_panes = f"{get_column_letter(len(id_cols) + 1)}2"
    last = get_column_letter(len(all_cols))
    ws.auto_filter.ref = f"A1:{last}{len(rows) + 1}"
    for idx, (label, get) in enumerate(all_cols, 1):
        sample = [str(get(r)) for r in rows[:6] if get(r) not in ("", None)]
        width = max([len(str(label))] + [len(s) for s in sample] + [10]) + 2
        ws.column_dimensions[get_column_letter(idx)].width = min(max(width, 12), 44)
    return ws


def _overview(wb, rows, used_titles):
    headline_cols = [(h, (lambda r, k=h: r.get(k, ""))) for h in HEADLINES
                     if any(h in r for r in rows)]
    # populated-field count + status string per layer
    def _status_str(r):
        ls = r.get("_layer_status", {}) or {}
        return ", ".join(f"{k}:{v}" for k, v in sorted(ls.items()))
    extra = [("layer_status", _status_str)]
    _write_sheet(wb, "Overview", headline_cols + extra, rows, "C8E6C9", used_titles)


def _layer_health(wb, rows, used_titles):
    layers, seen = [], set()
    for r in rows:
        for layer in (r.get("_layer_status", {}) or {}):
            if layer not in seen:
                seen.add(layer)
                layers.append(layer)
    layers.sort()
    cols = [(layer, (lambda r, L=layer: (r.get("_layer_status", {}) or {}).get(L, "")))
            for layer in layers]
    _write_sheet(wb, "Layer_Health", cols, rows, "FFE0B2", used_titles)


def _all_metrics(wb, rows, used_titles):
    keys = _matching_keys(rows, lambda k: True)
    cols = _expand_columns(rows, keys)
    _write_sheet(wb, "ALL_METRICS", cols, rows, "ECEFF1", used_titles)


def write_deep_workbook(rows: list[dict], path) -> Path:
    """Write the ~20-tab workbook. Returns the path written."""
    path = Path(path)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # drop default empty sheet
    used: set[str] = set()

    _overview(wb, rows, used)
    for title, color, predicate in TAB_SPEC:
        keys = _matching_keys(rows, predicate)
        cols = _expand_columns(rows, keys)
        _write_sheet(wb, title, cols, rows, color, used)
    _layer_health(wb, rows, used)
    _all_metrics(wb, rows, used)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    return path


def main(argv: list[str] | None = None) -> int:
    argv = list(sys.argv[1:] if argv is None else argv)
    if not argv:
        print("usage: python deep_workbook.py <pipeline_results.json> [out.xlsx]")
        return 2
    src = Path(argv[0])
    rows = json.loads(src.read_text(encoding="utf-8"))
    if isinstance(rows, dict):
        rows = [rows]
    out = Path(argv[1]) if len(argv) > 1 else src.with_name(src.stem + "_DEEP.xlsx")
    written = write_deep_workbook(rows, out)
    print(f"[deep_workbook] {len(rows)} paper(s) -> {written}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

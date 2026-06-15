"""
WORD-LEVEL ALIGNMENT MATRIX
============================
Maps every scoring channel to the same word position in Excel.
Each column = one content word (no spaces, no punctuation-only tokens).
Each row group = a scoring channel.

Channels:
  A. GoEmotions 27 (BERT sentence-level, attributed to word positions)
  B. Fruit Blend (9 fruits from GoEmotions, blended formula)
  C. Anti-Fruit (9 anti-fruits from GoEmotions)
  D. SBERT Fruit (9 semantic anchor similarities)
  E. SBERT Chi (10 Master Equation variables)
  F. Structural v2 (11 pattern categories, word-level attribution)
  G. SBERT 24 Properties (6 CORE + 18 DERIVED)
  H. SBERT 10 Laws Constructive
  I. SBERT 10 Laws Destructive
  J. SBERT Fruits-as-Physics (9 + 9 anti)
  K. SBERT Armor of God (6)
  L. SBERT Beatitudes (8)
  M. SBERT Gifts of Spirit (9)
  N. SBERT Couplings (8)

Layout in Excel:
  - Words across columns (auto-rotated headers)
  - Channel groups down rows, separated by 2 blank rows
  - Conditional formatting: green=positive, red=negative, gold=structural hits
"""

import re
import json
import numpy as np
from pathlib import Path
from collections import defaultdict

# ── Imports from sibling modules ──
import sys
_parent = str(Path(__file__).resolve().parent.parent)
if _parent not in sys.path:
    sys.path.insert(0, _parent)

from heartbeat_analyzer import (
    FRUIT_ANCHORS, FRUIT_ANTI_ANCHORS, CHI_ANCHORS, CHI_ANTI_ANCHORS,
    STRUCTURAL_PATTERNS, _get_model, _cosine_sim
)
from canonical_anchors import (
    PROPERTIES_24, LAWS_CONSTRUCTIVE, LAWS_DESTRUCTIVE,
    FRUITS_PHYSICS, ANTI_FRUITS_PHYSICS, ARMOR, BEATITUDES, GIFTS, COUPLINGS
)
from openai_paper_intel import (
    analyze_paper_strategic, score_sentences_openai, write_strategic_markdown,
    SENTENCE_TYPE_ENCODING, HAS_OPENAI
)

# GoEmotions labels and mappings (from emotion_analyzer)
GOEMOTIONS_LABELS = [
    'admiration', 'amusement', 'anger', 'annoyance', 'approval', 'caring',
    'confusion', 'curiosity', 'desire', 'disappointment', 'disapproval',
    'disgust', 'embarrassment', 'excitement', 'fear', 'gratitude', 'grief',
    'joy', 'love', 'nervousness', 'optimism', 'pride', 'realization',
    'relief', 'remorse', 'sadness', 'surprise', 'neutral'
]

FRUIT_BLEND = {
    'love': {'positive': ['love', 'caring', 'admiration', 'gratitude'],
             'negative': ['anger', 'disgust', 'disapproval']},
    'joy': {'positive': ['joy', 'optimism', 'excitement', 'pride', 'gratitude'],
            'negative': ['sadness', 'grief', 'disappointment']},
    'peace': {'positive': ['relief', 'approval', 'realization', 'gratitude'],
              'negative': ['anger', 'fear', 'nervousness', 'annoyance']},
    'patience': {'positive': ['caring', 'approval', 'relief', 'optimism'],
                 'negative': ['anger', 'annoyance', 'disappointment', 'disgust']},
    'kindness': {'positive': ['caring', 'approval', 'gratitude', 'love'],
                 'negative': ['disgust', 'disapproval', 'anger', 'annoyance']},
    'goodness': {'positive': ['admiration', 'approval', 'gratitude', 'optimism'],
                 'negative': ['disgust', 'disapproval', 'anger']},
    'faithfulness': {'positive': ['realization', 'approval', 'admiration', 'optimism'],
                     'negative': ['confusion', 'nervousness', 'fear', 'embarrassment']},
    'gentleness': {'positive': ['caring', 'love', 'relief', 'approval'],
                   'negative': ['anger', 'annoyance', 'disgust', 'pride']},
    'self_control': {'positive': ['realization', 'approval', 'relief'],
                     'negative': ['desire', 'anger', 'annoyance', 'disgust', 'excitement']},
}

ANTI_FRUIT_MAP = {
    'hatred':     ['anger', 'disgust', 'disapproval'],
    'despair':    ['sadness', 'grief', 'disappointment'],
    'conflict':   ['anger', 'annoyance', 'fear', 'nervousness'],
    'impatience': ['anger', 'annoyance', 'disappointment'],
    'cruelty':    ['disgust', 'disapproval', 'anger', 'annoyance'],
    'corruption': ['disgust', 'disapproval', 'anger'],
    'betrayal':   ['confusion', 'nervousness', 'disappointment', 'fear'],
    'harshness':  ['anger', 'annoyance', 'disgust'],
    'indulgence': ['desire', 'excitement', 'anger', 'annoyance'],
}


# ═══════════════════════════════════════════════════════════════
# WORD TOKENIZATION
# ═════════════════════════════════════════════��═════════════════

def _tokenize_to_words(text):
    """
    Split text into content words with position tracking.
    Returns list of dicts: {word, char_start, char_end, sentence_idx}
    Skips pure punctuation, numbers-only, and very short tokens.
    """
    # Strip frontmatter
    text = re.sub(r'^---.*?---', '', text, flags=re.DOTALL)
    # Strip markdown headers (keep text)
    text = re.sub(r'#+\s+', '', text)
    # Strip tables
    text = re.sub(r'\|[^\n]+\|', '', text)
    # Strip markdown formatting
    text = re.sub(r'[*_`~]{1,3}', '', text)
    # Strip link syntax
    text = re.sub(r'\[([^\]]*)\]\([^)]*\)', r'\1', text)

    # Split into sentences first (for sentence_idx tracking)
    raw_sents = re.split(r'(?<=[.!?])\s+', text)
    sentences = []
    sent_words = []
    global_word_idx = 0

    for sent_idx, sent in enumerate(raw_sents):
        sent = sent.strip()
        if len(sent) < 10:
            continue

        # Tokenize this sentence into words
        words_in_sent = []
        for m in re.finditer(r"[A-Za-zÀ-ÿ'\-]+", sent):
            w = m.group()
            if len(w) < 2:
                continue
            words_in_sent.append({
                'word': w,
                'sentence_idx': len(sentences),
                'word_idx': global_word_idx,
                'sent_text': sent,
            })
            global_word_idx += 1

        if words_in_sent:
            sentences.append(sent)
            sent_words.extend(words_in_sent)

    return sent_words, sentences


# ════════════════════════════════════════════��══════════════════
# SCORING ENGINES
# ══════════════════��════════════════════════════════════════════

def _run_goemotions(sentences):
    """Run GoEmotions on sentences, return per-sentence score dicts."""
    try:
        from transformers import pipeline
        pipe = pipeline(
            'text-classification',
            model='monologg/bert-base-cased-goemotions-original',
            top_k=None, device=-1, batch_size=16,
        )
    except Exception as e:
        print(f"  GoEmotions load error: {e}")
        return None

    # Chunk sentences to max 350 chars
    chunks = []
    chunk_to_sent = []  # maps chunk index → list of sentence indices
    buf = ""
    buf_sents = []

    for i, s in enumerate(sentences):
        if len(buf) + len(s) > 350:
            if buf:
                chunks.append(buf.strip())
                chunk_to_sent.append(list(buf_sents))
            buf = s
            buf_sents = [i]
        else:
            buf = (buf + " " + s) if buf else s
            buf_sents.append(i)
    if buf:
        chunks.append(buf.strip())
        chunk_to_sent.append(list(buf_sents))

    chunks = [c for c in chunks if len(c) > 20]

    # Run GoEmotions
    per_chunk_scores = []
    for batch_start in range(0, len(chunks), 16):
        batch = chunks[batch_start:batch_start + 16]
        results = pipe(batch)
        for result in results:
            scores = {item['label']: item['score'] for item in result}
            per_chunk_scores.append(scores)

    # Map chunk scores back to sentences
    per_sent_scores = [{label: 0.0 for label in GOEMOTIONS_LABELS} for _ in sentences]
    for chunk_idx, sent_indices in enumerate(chunk_to_sent):
        if chunk_idx < len(per_chunk_scores):
            for si in sent_indices:
                if si < len(per_sent_scores):
                    per_sent_scores[si] = per_chunk_scores[chunk_idx]

    return per_sent_scores


def _encode_anchor_group(model, anchor_dict):
    """Encode an anchor dict {name: description} into keys + vectors."""
    keys = list(anchor_dict.keys())
    vecs = model.encode([anchor_dict[k] for k in keys],
                         convert_to_numpy=True, show_progress_bar=False)
    return keys, vecs


def _score_sentences_vs_anchors(sent_vecs, keys, pos_vecs, neg_vecs=None):
    """Score each sentence against anchors. If neg_vecs given, returns net."""
    results = []
    for svec in sent_vecs:
        pos = _cosine_sim(svec, pos_vecs)
        if neg_vecs is not None:
            neg = _cosine_sim(svec, neg_vecs)
            net = pos - neg
        else:
            net = pos
        results.append({k: float(net[i]) for i, k in enumerate(keys)})
    return results


def _run_sbert(sentences):
    """Run SBERT scoring on sentences for ALL anchor groups."""
    model = _get_model()

    print(f"    Encoding {len(sentences)} sentences...")
    sent_vecs = model.encode(sentences, convert_to_numpy=True,
                              show_progress_bar=True, batch_size=64)

    # ── Original channels (Fruit + Chi net scores) ──
    fruit_keys, fruit_vecs = _encode_anchor_group(model, FRUIT_ANCHORS)
    _, fruit_anti_vecs = _encode_anchor_group(model, FRUIT_ANTI_ANCHORS)
    chi_keys, chi_vecs = _encode_anchor_group(model, CHI_ANCHORS)
    _, chi_anti_vecs = _encode_anchor_group(model, CHI_ANTI_ANCHORS)

    per_sent_fruit = _score_sentences_vs_anchors(sent_vecs, fruit_keys, fruit_vecs, fruit_anti_vecs)
    per_sent_chi = _score_sentences_vs_anchors(sent_vecs, chi_keys, chi_vecs, chi_anti_vecs)

    # ── Canonical anchor groups (cosine similarity, no net) ──
    print(f"    Encoding canonical anchors (107 anchors)...")

    prop_keys, prop_vecs = _encode_anchor_group(model, PROPERTIES_24)
    per_sent_props = _score_sentences_vs_anchors(sent_vecs, prop_keys, prop_vecs)

    lawc_keys, lawc_vecs = _encode_anchor_group(model, LAWS_CONSTRUCTIVE)
    per_sent_lawc = _score_sentences_vs_anchors(sent_vecs, lawc_keys, lawc_vecs)

    lawd_keys, lawd_vecs = _encode_anchor_group(model, LAWS_DESTRUCTIVE)
    per_sent_lawd = _score_sentences_vs_anchors(sent_vecs, lawd_keys, lawd_vecs)

    # Fruits-as-physics has natural anti pairs
    fp_keys, fp_vecs = _encode_anchor_group(model, FRUITS_PHYSICS)
    _, afp_vecs = _encode_anchor_group(model, ANTI_FRUITS_PHYSICS)
    per_sent_fruitphys = _score_sentences_vs_anchors(sent_vecs, fp_keys, fp_vecs, afp_vecs)

    armor_keys, armor_vecs = _encode_anchor_group(model, ARMOR)
    per_sent_armor = _score_sentences_vs_anchors(sent_vecs, armor_keys, armor_vecs)

    beat_keys, beat_vecs = _encode_anchor_group(model, BEATITUDES)
    per_sent_beat = _score_sentences_vs_anchors(sent_vecs, beat_keys, beat_vecs)

    gift_keys, gift_vecs = _encode_anchor_group(model, GIFTS)
    per_sent_gifts = _score_sentences_vs_anchors(sent_vecs, gift_keys, gift_vecs)

    coup_keys, coup_vecs = _encode_anchor_group(model, COUPLINGS)
    per_sent_coup = _score_sentences_vs_anchors(sent_vecs, coup_keys, coup_vecs)

    return {
        'fruit': per_sent_fruit,
        'chi': per_sent_chi,
        'props': per_sent_props,
        'laws_c': per_sent_lawc,
        'laws_d': per_sent_lawd,
        'fruit_phys': per_sent_fruitphys,
        'armor': per_sent_armor,
        'beatitudes': per_sent_beat,
        'gifts': per_sent_gifts,
        'couplings': per_sent_coup,
    }


def _run_structural_wordlevel(words):
    """Check each word's surrounding context for structural pattern hits."""
    per_word_struct = []
    for w in words:
        sent = w['sent_text']
        hits = {}
        for category, patterns in STRUCTURAL_PATTERNS.items():
            count = 0
            for pat in patterns:
                count += len(re.findall(pat, sent, re.IGNORECASE))
            hits[category] = count
        per_word_struct.append(hits)
    return per_word_struct


# ══��═══════════════════════��════════════════════════════════════
# EXCEL WRITER
# ════���═════════════���════════════════════════════════════════════

def write_word_matrix(paper_path: str, output_path: str = None, run_openai: bool = False):
    """
    Build and write the word-level alignment matrix to Excel.
    If run_openai=True, adds channels O-V via OpenAI API + strategic report.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.formatting.rule import CellIsRule
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("ERROR: pip install openpyxl")
        return

    text = Path(paper_path).read_text(encoding='utf-8', errors='replace')
    paper_name = Path(paper_path).stem

    if output_path is None:
        output_path = str(Path(paper_path).parent / f"{paper_name}_WORD_MATRIX.xlsx")

    print(f"  Tokenizing words...")
    words, sentences = _tokenize_to_words(text)
    n_words = len(words)
    n_sents = len(sentences)
    print(f"  {n_words} content words, {n_sents} sentences")

    if n_words == 0:
        print("  ERROR: No words found")
        return

    # ── Run all scoring channels ──
    print(f"  Running GoEmotions (27 emotions)...")
    goemo_scores = _run_goemotions(sentences)

    print(f"  Running SBERT (Fruits + Chi + 107 canonical anchors)...")
    sbert_all = _run_sbert(sentences)
    sbert_fruit = sbert_all['fruit']
    sbert_chi = sbert_all['chi']

    print(f"  Running Structural v2 patterns...")
    struct_scores = _run_structural_wordlevel(words)

    # ── OpenAI sentence scoring (channels O-V) ──
    openai_scores = None
    strategic_analysis = None
    if run_openai and HAS_OPENAI:
        cache_dir = str(Path(output_path or paper_path).parent / '_openai_cache')
        print(f"  Running OpenAI sentence scoring...")
        openai_scores = score_sentences_openai(sentences, paper_name, cache_dir=cache_dir)
        print(f"  Running OpenAI strategic analysis...")
        strategic_analysis = analyze_paper_strategic(paper_path, cache_dir=cache_dir)
    elif run_openai:
        print(f"  WARNING: run_openai=True but OpenAI not configured (set OPENAI_API_KEY)")

    # ── Compute fruit blends and anti-fruits from GoEmotions ──
    def _blend_fruit(goemo, fruit_name):
        ch = FRUIT_BLEND[fruit_name]
        pos_avg = sum(goemo.get(s, 0) for s in ch['positive']) / len(ch['positive'])
        neg_avg = sum(goemo.get(s, 0) for s in ch['negative']) / len(ch['negative'])
        return (pos_avg * 0.6) + ((1.0 - neg_avg) * 0.4)

    def _anti_fruit(goemo, anti_name):
        sources = ANTI_FRUIT_MAP[anti_name]
        return sum(goemo.get(s, 0) for s in sources) / len(sources)

    # ── Build Excel ──
    print(f"  Writing Excel...")
    wb = Workbook()
    ws = wb.active
    ws.title = "Word Matrix"

    # Styles
    gold = Font(color="D4AF37", bold=True, size=9)
    white = Font(color="E8E8E8", size=8)
    dim = Font(color="888888", size=7)
    header_fill = PatternFill("solid", fgColor="0A0A0A")
    label_fill = PatternFill("solid", fgColor="141414")
    group_fill = PatternFill("solid", fgColor="1A1A1A")
    rotated = Alignment(textRotation=90, horizontal='center', vertical='bottom')
    normal_align = Alignment(horizontal='left', vertical='center')
    center_align = Alignment(horizontal='center', vertical='center')

    # Column widths
    ws.column_dimensions['A'].width = 22  # Row labels
    ws.column_dimensions['B'].width = 14  # Sub-labels

    # Word columns start at C (col 3)
    WORD_COL_START = 3
    for i in range(n_words):
        col_letter = get_column_letter(WORD_COL_START + i)
        ws.column_dimensions[col_letter].width = 3.5

    # ── HEADER ROWS ──
    row = 1

    # Row 1: Word index
    ws.cell(row=row, column=1, value="INDEX").font = gold
    ws.cell(row=row, column=1).fill = header_fill
    for i, w in enumerate(words):
        c = ws.cell(row=row, column=WORD_COL_START + i, value=i)
        c.font = dim
        c.fill = header_fill
        c.alignment = center_align

    # Row 2: Actual words (rotated)
    row = 2
    ws.cell(row=row, column=1, value="WORD").font = gold
    ws.cell(row=row, column=1).fill = header_fill
    ws.row_dimensions[row].height = 80
    for i, w in enumerate(words):
        c = ws.cell(row=row, column=WORD_COL_START + i, value=w['word'])
        c.font = Font(color="D4AF37", size=7, bold=True)
        c.fill = header_fill
        c.alignment = rotated

    # Row 3: Sentence index
    row = 3
    ws.cell(row=row, column=1, value="SENTENCE #").font = dim
    ws.cell(row=row, column=1).fill = header_fill
    for i, w in enumerate(words):
        c = ws.cell(row=row, column=WORD_COL_START + i, value=w['sentence_idx'])
        c.font = dim
        c.fill = header_fill
        c.alignment = center_align

    row = 4  # blank separator

    # ══════════════════════════════════════════════════════════
    # CHANNEL A: GoEmotions 27
    # ════��═══════════════════════��═════════════════════════════
    row = 5
    ws.cell(row=row, column=1, value="═ GoEmotions 27 ═").font = Font(color="FF6B6B", bold=True, size=10)
    ws.cell(row=row, column=1).fill = group_fill
    row += 1

    for emo in GOEMOTIONS_LABELS:
        ws.cell(row=row, column=1, value=f"  {emo}").font = white
        ws.cell(row=row, column=1).fill = label_fill
        ws.cell(row=row, column=1).alignment = normal_align

        if goemo_scores:
            for i, w in enumerate(words):
                si = w['sentence_idx']
                val = goemo_scores[si].get(emo, 0) if si < len(goemo_scores) else 0
                c = ws.cell(row=row, column=WORD_COL_START + i, value=round(val, 4))
                c.font = dim
                c.alignment = center_align
                c.number_format = '0.00'
        row += 1

    row += 2  # skip 2

    # ════��══════════════���═════════════════════════════════���════
    # CHANNEL B: Fruit Blend (from GoEmotions)
    # ════════���═════════════════════════════════════════════════
    ws.cell(row=row, column=1, value="═ Fruit Blend (GoEmo) ═").font = Font(color="6BCB77", bold=True, size=10)
    ws.cell(row=row, column=1).fill = group_fill
    row += 1

    for fruit_name in FRUIT_BLEND:
        ws.cell(row=row, column=1, value=f"  {fruit_name}").font = Font(color="6BCB77", size=8)
        ws.cell(row=row, column=1).fill = label_fill

        if goemo_scores:
            for i, w in enumerate(words):
                si = w['sentence_idx']
                goemo = goemo_scores[si] if si < len(goemo_scores) else {}
                val = _blend_fruit(goemo, fruit_name)
                c = ws.cell(row=row, column=WORD_COL_START + i, value=round(val, 4))
                c.font = dim
                c.alignment = center_align
                c.number_format = '0.00'
        row += 1

    row += 2  # skip 2

    # ═══════════���══════════════════════════════════════════════
    # CHANNEL C: Anti-Fruits (from GoEmotions)
    # ═══════���═══════════════════���══════════════════════════════
    ws.cell(row=row, column=1, value="═ Anti-Fruits (GoEmo) ��").font = Font(color="FF6B6B", bold=True, size=10)
    ws.cell(row=row, column=1).fill = group_fill
    row += 1

    for anti_name in ANTI_FRUIT_MAP:
        ws.cell(row=row, column=1, value=f"  {anti_name}").font = Font(color="FF6B6B", size=8)
        ws.cell(row=row, column=1).fill = label_fill

        if goemo_scores:
            for i, w in enumerate(words):
                si = w['sentence_idx']
                goemo = goemo_scores[si] if si < len(goemo_scores) else {}
                val = _anti_fruit(goemo, anti_name)
                c = ws.cell(row=row, column=WORD_COL_START + i, value=round(val, 4))
                c.font = dim
                c.alignment = center_align
                c.number_format = '0.00'
        row += 1

    row += 2  # skip 2

    # ═══════════════════════════════════��══════════════════════
    # CHANNEL D: SBERT Fruit (semantic anchors)
    # ═══════════════���═════════════════════════════���════════════
    ws.cell(row=row, column=1, value="═ SBERT Fruit Net ═").font = Font(color="DDA0DD", bold=True, size=10)
    ws.cell(row=row, column=1).fill = group_fill
    row += 1

    for fruit_name in FRUIT_ANCHORS:
        ws.cell(row=row, column=1, value=f"  {fruit_name}").font = Font(color="DDA0DD", size=8)
        ws.cell(row=row, column=1).fill = label_fill

        for i, w in enumerate(words):
            si = w['sentence_idx']
            val = sbert_fruit[si].get(fruit_name, 0) if si < len(sbert_fruit) else 0
            c = ws.cell(row=row, column=WORD_COL_START + i, value=round(val, 4))
            c.font = dim
            c.alignment = center_align
            c.number_format = '0.000'
        row += 1

    row += 2  # skip 2

    # ══════════��════════════════════════════════════���══════════
    # CHANNEL E: SBERT Chi (Master Equation)
    # ════════���═════════════════════════════════════════════════
    ws.cell(row=row, column=1, value="═ SBERT Chi Net ═").font = Font(color="4D96FF", bold=True, size=10)
    ws.cell(row=row, column=1).fill = group_fill
    row += 1

    for chi_name in CHI_ANCHORS:
        ws.cell(row=row, column=1, value=f"  {chi_name} ({CHI_ANCHORS[chi_name].split(',')[0]})").font = Font(color="4D96FF", size=8)
        ws.cell(row=row, column=1).fill = label_fill

        for i, w in enumerate(words):
            si = w['sentence_idx']
            val = sbert_chi[si].get(chi_name, 0) if si < len(sbert_chi) else 0
            c = ws.cell(row=row, column=WORD_COL_START + i, value=round(val, 4))
            c.font = dim
            c.alignment = center_align
            c.number_format = '0.000'
        row += 1

    row += 2  # skip 2

    # ══��═════════���═════════════════════════════════��═══════════
    # CHANNEL F: Structural v2 (pattern hits)
    # ════════════════════════════════════════��═════════════════
    ws.cell(row=row, column=1, value="═ Structural v2 ═").font = Font(color="D4AF37", bold=True, size=10)
    ws.cell(row=row, column=1).fill = group_fill
    row += 1

    for category in STRUCTURAL_PATTERNS:
        is_neg = category == 'overclaim'
        color = "FF6B6B" if is_neg else "D4AF37"
        ws.cell(row=row, column=1, value=f"  {category}{'  ⚠' if is_neg else ''}").font = Font(color=color, size=8)
        ws.cell(row=row, column=1).fill = label_fill

        for i in range(n_words):
            val = struct_scores[i].get(category, 0)
            c = ws.cell(row=row, column=WORD_COL_START + i, value=val if val > 0 else '')
            if val > 0:
                c.font = Font(color=color, bold=True, size=9)
                c.fill = PatternFill("solid", fgColor="2A2A00" if not is_neg else "2A0000")
            else:
                c.font = dim
            c.alignment = center_align
        row += 1

    # ══════════════════════════════════════════════════════════
    # CHANNEL G: 24 Properties (SBERT cosine similarity)
    # ══════════════════════════════════════════════════════════
    row += 2
    ws.cell(row=row, column=1, value="═ 24 Properties (SBERT) ═").font = Font(color="00CED1", bold=True, size=10)
    ws.cell(row=row, column=1).fill = group_fill
    row += 1

    prop_keys = list(PROPERTIES_24.keys())
    for pk in prop_keys:
        # Mark CORE properties
        is_core = pk in ('P04_simple', 'P05_consistent', 'P12_true', 'P13_rational', 'P23_generative', 'P24_judging')
        label = f"  {'★ ' if is_core else ''}{pk}"
        ws.cell(row=row, column=1, value=label).font = Font(color="00CED1" if is_core else "5F9EA0", size=8, bold=is_core)
        ws.cell(row=row, column=1).fill = label_fill

        for i, w in enumerate(words):
            si = w['sentence_idx']
            val = sbert_all['props'][si].get(pk, 0) if si < len(sbert_all['props']) else 0
            c = ws.cell(row=row, column=WORD_COL_START + i, value=round(val, 4))
            c.font = dim
            c.alignment = center_align
            c.number_format = '0.000'
        row += 1

    # ══════════════════════════════════════════════════════════
    # CHANNEL H: 10 Laws Constructive (SBERT)
    # ══════════════════════════════════════════════════════════
    row += 2
    ws.cell(row=row, column=1, value="═ 10 Laws CONSTRUCTIVE ═").font = Font(color="32CD32", bold=True, size=10)
    ws.cell(row=row, column=1).fill = group_fill
    row += 1

    lawc_keys = list(LAWS_CONSTRUCTIVE.keys())
    for lk in lawc_keys:
        ws.cell(row=row, column=1, value=f"  {lk}").font = Font(color="32CD32", size=8)
        ws.cell(row=row, column=1).fill = label_fill
        for i, w in enumerate(words):
            si = w['sentence_idx']
            val = sbert_all['laws_c'][si].get(lk, 0) if si < len(sbert_all['laws_c']) else 0
            c = ws.cell(row=row, column=WORD_COL_START + i, value=round(val, 4))
            c.font = dim
            c.alignment = center_align
            c.number_format = '0.000'
        row += 1

    # ══════════════════════════════════════════════════════════
    # CHANNEL I: 10 Laws Destructive (SBERT)
    # ══════════════════════════════════════════════════════════
    row += 2
    ws.cell(row=row, column=1, value="═ 10 Laws DESTRUCTIVE ═").font = Font(color="DC143C", bold=True, size=10)
    ws.cell(row=row, column=1).fill = group_fill
    row += 1

    lawd_keys = list(LAWS_DESTRUCTIVE.keys())
    for lk in lawd_keys:
        ws.cell(row=row, column=1, value=f"  {lk}").font = Font(color="DC143C", size=8)
        ws.cell(row=row, column=1).fill = label_fill
        for i, w in enumerate(words):
            si = w['sentence_idx']
            val = sbert_all['laws_d'][si].get(lk, 0) if si < len(sbert_all['laws_d']) else 0
            c = ws.cell(row=row, column=WORD_COL_START + i, value=round(val, 4))
            c.font = dim
            c.alignment = center_align
            c.number_format = '0.000'
        row += 1

    # ══════════════════════════════════════════════════════════
    # CHANNEL J: Fruits-as-Physics NET (SBERT)
    # ══════════════════════════════════════════════════════════
    row += 2
    ws.cell(row=row, column=1, value="═ Fruits as PHYSICS Net ═").font = Font(color="FF69B4", bold=True, size=10)
    ws.cell(row=row, column=1).fill = group_fill
    row += 1

    fp_keys = list(FRUITS_PHYSICS.keys())
    for fk in fp_keys:
        ws.cell(row=row, column=1, value=f"  {fk}").font = Font(color="FF69B4", size=8)
        ws.cell(row=row, column=1).fill = label_fill
        for i, w in enumerate(words):
            si = w['sentence_idx']
            val = sbert_all['fruit_phys'][si].get(fk, 0) if si < len(sbert_all['fruit_phys']) else 0
            c = ws.cell(row=row, column=WORD_COL_START + i, value=round(val, 4))
            c.font = dim
            c.alignment = center_align
            c.number_format = '0.000'
        row += 1

    # ══════════════════════════════════════════════════════════
    # CHANNEL K: Armor of God (SBERT)
    # ══════════════════════════════════════════════════════════
    row += 2
    ws.cell(row=row, column=1, value="═ Armor of God (SBERT) ═").font = Font(color="FFD700", bold=True, size=10)
    ws.cell(row=row, column=1).fill = group_fill
    row += 1

    armor_keys = list(ARMOR.keys())
    for ak in armor_keys:
        ws.cell(row=row, column=1, value=f"  {ak}").font = Font(color="FFD700", size=8)
        ws.cell(row=row, column=1).fill = label_fill
        for i, w in enumerate(words):
            si = w['sentence_idx']
            val = sbert_all['armor'][si].get(ak, 0) if si < len(sbert_all['armor']) else 0
            c = ws.cell(row=row, column=WORD_COL_START + i, value=round(val, 4))
            c.font = dim
            c.alignment = center_align
            c.number_format = '0.000'
        row += 1

    # ══════════════════════════════════════════════════════════
    # CHANNEL L: Beatitudes (SBERT)
    # ══════════════════════════════════════════════════════════
    row += 2
    ws.cell(row=row, column=1, value="═ Beatitudes (SBERT) ═").font = Font(color="87CEEB", bold=True, size=10)
    ws.cell(row=row, column=1).fill = group_fill
    row += 1

    beat_keys = list(BEATITUDES.keys())
    for bk in beat_keys:
        ws.cell(row=row, column=1, value=f"  {bk}").font = Font(color="87CEEB", size=8)
        ws.cell(row=row, column=1).fill = label_fill
        for i, w in enumerate(words):
            si = w['sentence_idx']
            val = sbert_all['beatitudes'][si].get(bk, 0) if si < len(sbert_all['beatitudes']) else 0
            c = ws.cell(row=row, column=WORD_COL_START + i, value=round(val, 4))
            c.font = dim
            c.alignment = center_align
            c.number_format = '0.000'
        row += 1

    # ══════════════════════════════════════════════════════════
    # CHANNEL M: Gifts of the Spirit (SBERT)
    # ══════════════════════════════════════════════════════════
    row += 2
    ws.cell(row=row, column=1, value="═ Gifts of Spirit (SBERT) ═").font = Font(color="DA70D6", bold=True, size=10)
    ws.cell(row=row, column=1).fill = group_fill
    row += 1

    gift_keys = list(GIFTS.keys())
    for gk in gift_keys:
        ws.cell(row=row, column=1, value=f"  {gk}").font = Font(color="DA70D6", size=8)
        ws.cell(row=row, column=1).fill = label_fill
        for i, w in enumerate(words):
            si = w['sentence_idx']
            val = sbert_all['gifts'][si].get(gk, 0) if si < len(sbert_all['gifts']) else 0
            c = ws.cell(row=row, column=WORD_COL_START + i, value=round(val, 4))
            c.font = dim
            c.alignment = center_align
            c.number_format = '0.000'
        row += 1

    # ══════════════════════════════════════════════════════════
    # CHANNEL N: Couplings (SBERT)
    # ══════════════════════════════════════════════════════════
    row += 2
    ws.cell(row=row, column=1, value="═ Couplings (SBERT) ═").font = Font(color="FFA07A", bold=True, size=10)
    ws.cell(row=row, column=1).fill = group_fill
    row += 1

    coup_keys = list(COUPLINGS.keys())
    for ck in coup_keys:
        ws.cell(row=row, column=1, value=f"  {ck}").font = Font(color="FFA07A", size=8)
        ws.cell(row=row, column=1).fill = label_fill
        for i, w in enumerate(words):
            si = w['sentence_idx']
            val = sbert_all['couplings'][si].get(ck, 0) if si < len(sbert_all['couplings']) else 0
            c = ws.cell(row=row, column=WORD_COL_START + i, value=round(val, 4))
            c.font = dim
            c.alignment = center_align
            c.number_format = '0.000'
        row += 1

    # ══════════════════════════════════════════════════════════
    # CHANNELS O-V: OpenAI Sentence-Level Intelligence
    # ══════════════════════════════════════════════════════════
    if openai_scores:
        row += 2
        ws.cell(row=row, column=1, value="═ OpenAI: Sentence Type ═").font = Font(color="FF8C00", bold=True, size=10)
        ws.cell(row=row, column=1).fill = group_fill
        row += 1

        # O. Sentence Type (encoded as numeric)
        ws.cell(row=row, column=1, value="  sentence_type").font = Font(color="FF8C00", size=8)
        ws.cell(row=row, column=1).fill = label_fill
        for i, w in enumerate(words):
            si = w['sentence_idx']
            if si < len(openai_scores):
                stype = openai_scores[si].get('type', 'narrative')
                val = SENTENCE_TYPE_ENCODING.get(stype, 0.3)
                c = ws.cell(row=row, column=WORD_COL_START + i, value=round(val, 2))
                c.font = dim
                c.alignment = center_align
                c.number_format = '0.00'
        row += 1

        # P. Bridge Score
        ws.cell(row=row, column=1, value="  bridge_score").font = Font(color="FF8C00", size=8)
        ws.cell(row=row, column=1).fill = label_fill
        for i, w in enumerate(words):
            si = w['sentence_idx']
            if si < len(openai_scores):
                val = openai_scores[si].get('bridge_score', 0)
                c = ws.cell(row=row, column=WORD_COL_START + i, value=round(val, 3))
                c.font = dim
                c.alignment = center_align
                c.number_format = '0.000'
        row += 1

        # Q. Logic Score
        ws.cell(row=row, column=1, value="  logic_score").font = Font(color="FF8C00", size=8)
        ws.cell(row=row, column=1).fill = label_fill
        for i, w in enumerate(words):
            si = w['sentence_idx']
            if si < len(openai_scores):
                val = openai_scores[si].get('logic_score', 0)
                c = ws.cell(row=row, column=WORD_COL_START + i, value=round(val, 3))
                c.font = dim
                c.alignment = center_align
                c.number_format = '0.000'
        row += 1

        # R. Falsifiability
        ws.cell(row=row, column=1, value="  falsifiability").font = Font(color="FF8C00", size=8)
        ws.cell(row=row, column=1).fill = label_fill
        for i, w in enumerate(words):
            si = w['sentence_idx']
            if si < len(openai_scores):
                val = openai_scores[si].get('falsifiable', 0)
                c = ws.cell(row=row, column=WORD_COL_START + i, value=round(val, 3))
                c.font = dim
                c.alignment = center_align
                c.number_format = '0.000'
        row += 1

        # S. Axiom dependency count
        ws.cell(row=row, column=1, value="  axiom_dep_count").font = Font(color="FF8C00", size=8)
        ws.cell(row=row, column=1).fill = label_fill
        for i, w in enumerate(words):
            si = w['sentence_idx']
            if si < len(openai_scores):
                val = len(openai_scores[si].get('axiom_deps', []))
                c = ws.cell(row=row, column=WORD_COL_START + i, value=val if val > 0 else '')
                if val > 0:
                    c.font = Font(color="FF8C00", bold=True, size=8)
                else:
                    c.font = dim
                c.alignment = center_align
        row += 1

        # T. Theory count
        ws.cell(row=row, column=1, value="  theory_count").font = Font(color="FF8C00", size=8)
        ws.cell(row=row, column=1).fill = label_fill
        for i, w in enumerate(words):
            si = w['sentence_idx']
            if si < len(openai_scores):
                val = len(openai_scores[si].get('theories', []))
                c = ws.cell(row=row, column=WORD_COL_START + i, value=val if val > 0 else '')
                if val > 0:
                    c.font = Font(color="FF8C00", bold=True, size=8)
                else:
                    c.font = dim
                c.alignment = center_align
        row += 1

        # U. Objection present (binary)
        ws.cell(row=row, column=1, value="  has_objection").font = Font(color="FF4444", size=8)
        ws.cell(row=row, column=1).fill = label_fill
        for i, w in enumerate(words):
            si = w['sentence_idx']
            if si < len(openai_scores):
                obj = openai_scores[si].get('strongest_objection', '')
                val = 1 if obj and len(obj) > 5 else 0
                c = ws.cell(row=row, column=WORD_COL_START + i, value=val if val > 0 else '')
                if val:
                    c.font = Font(color="FF4444", bold=True, size=8)
                    c.fill = PatternFill("solid", fgColor="2A0000")
                else:
                    c.font = dim
                c.alignment = center_align
        row += 1

    # ══════════════════════════════════════════════════════════
    # STRATEGIC INTEL SHEET (if OpenAI ran)
    # ══════════════════════════════════════════════════════════
    if strategic_analysis and 'error' not in strategic_analysis:
        ws_strat = wb.create_sheet("Strategic Intel")
        ws_strat.column_dimensions['A'].width = 25
        ws_strat.column_dimensions['B'].width = 80

        sr = 1
        ws_strat.cell(row=sr, column=1, value="STRATEGIC INTELLIGENCE").font = Font(color="FF8C00", bold=True, size=14)
        ws_strat.cell(row=sr, column=1).fill = header_fill
        sr += 2

        # Overall grade
        grade = strategic_analysis.get('overall_grade', {})
        if grade:
            ws_strat.cell(row=sr, column=1, value="OVERALL GRADE").font = Font(color="D4AF37", bold=True, size=11)
            sr += 1
            for k in ('rigor', 'originality', 'engagement', 'clarity', 'citation_completeness'):
                ws_strat.cell(row=sr, column=1, value=f"  {k}").font = Font(color="E8E8E8", size=9)
                ws_strat.cell(row=sr, column=2, value=grade.get(k, '')).font = Font(color="D4AF37", size=9)
                sr += 1
            ws_strat.cell(row=sr, column=1, value="  VERDICT").font = Font(color="D4AF37", bold=True, size=9)
            ws_strat.cell(row=sr, column=2, value=grade.get('one_line_verdict', '')).font = Font(color="E8E8E8", size=9)
            sr += 2

        # Competing theories
        theories = strategic_analysis.get('competing_theories', [])
        if theories:
            ws_strat.cell(row=sr, column=1, value="COMPETING THEORIES").font = Font(color="00CED1", bold=True, size=11)
            sr += 1
            for t in theories:
                ws_strat.cell(row=sr, column=1, value=f"  {t.get('theory', '')}").font = Font(color="00CED1", size=9)
                ws_strat.cell(row=sr, column=2, value=f"{t.get('relationship', '')} | {t.get('relevance', '')}").font = Font(color="E8E8E8", size=8)
                sr += 1
            sr += 1

        # Citation gaps
        gaps = strategic_analysis.get('citation_gaps', [])
        if gaps:
            ws_strat.cell(row=sr, column=1, value="CITATION GAPS").font = Font(color="FF6B6B", bold=True, size=11)
            sr += 1
            for g in gaps:
                ws_strat.cell(row=sr, column=1, value=f"  {g.get('author', '')}").font = Font(color="FF6B6B", size=9)
                ws_strat.cell(row=sr, column=2, value=f"{g.get('work', '')} — {g.get('why_needed', '')}").font = Font(color="E8E8E8", size=8)
                sr += 1
            sr += 1

        # Hook analysis
        hook = strategic_analysis.get('hook_analysis', {})
        if hook:
            ws_strat.cell(row=sr, column=1, value="HOOK ANALYSIS").font = Font(color="32CD32", bold=True, size=11)
            sr += 1
            ws_strat.cell(row=sr, column=1, value="  Current").font = Font(color="888888", size=9)
            ws_strat.cell(row=sr, column=2, value=hook.get('current_hook', '')).font = Font(color="E8E8E8", size=8)
            sr += 1
            ws_strat.cell(row=sr, column=1, value="  BETTER HOOK").font = Font(color="32CD32", bold=True, size=9)
            ws_strat.cell(row=sr, column=2, value=hook.get('better_hook', '')).font = Font(color="32CD32", size=8)
            sr += 1
            ws_strat.cell(row=sr, column=1, value="  Score").font = Font(color="888888", size=9)
            ws_strat.cell(row=sr, column=2, value=hook.get('hook_score', '')).font = Font(color="D4AF37", size=9)
            sr += 2

        # Argument map
        arg = strategic_analysis.get('argument_map', {})
        if arg:
            ws_strat.cell(row=sr, column=1, value="ARGUMENT MAP").font = Font(color="DDA0DD", bold=True, size=11)
            sr += 1
            for k in ('strongest_link', 'weakest_link', 'missing_step', 'tightening_suggestion'):
                ws_strat.cell(row=sr, column=1, value=f"  {k}").font = Font(color="DDA0DD", size=9)
                ws_strat.cell(row=sr, column=2, value=arg.get(k, '')).font = Font(color="E8E8E8", size=8)
                sr += 1
            sr += 1

        # Emotional appeal
        emo = strategic_analysis.get('emotional_appeal', {})
        if emo:
            ws_strat.cell(row=sr, column=1, value="EMOTIONAL APPEAL").font = Font(color="FF69B4", bold=True, size=11)
            sr += 1
            ws_strat.cell(row=sr, column=1, value="  Reader connection").font = Font(color="FF69B4", size=9)
            ws_strat.cell(row=sr, column=2, value=emo.get('reader_connection', '')).font = Font(color="E8E8E8", size=8)
            sr += 1
            ws_strat.cell(row=sr, column=1, value="  SUGGESTED ADD").font = Font(color="FF69B4", bold=True, size=9)
            ws_strat.cell(row=sr, column=2, value=emo.get('suggested_addition', '')).font = Font(color="FF69B4", size=8)
            sr += 1
            missed = emo.get('missed_opportunities', [])
            for m in missed:
                ws_strat.cell(row=sr, column=1, value="  Missed opp").font = Font(color="888888", size=8)
                ws_strat.cell(row=sr, column=2, value=str(m)).font = Font(color="E8E8E8", size=8)
                sr += 1
            sr += 1

        # Killer objection
        killer = strategic_analysis.get('killer_objection', {})
        if killer:
            ws_strat.cell(row=sr, column=1, value="KILLER OBJECTION").font = Font(color="FF0000", bold=True, size=11)
            sr += 1
            ws_strat.cell(row=sr, column=1, value="  Objection").font = Font(color="FF0000", size=9)
            ws_strat.cell(row=sr, column=2, value=killer.get('objection', '')).font = Font(color="FF0000", size=9)
            sr += 1
            ws_strat.cell(row=sr, column=1, value="  Why dangerous").font = Font(color="FF6B6B", size=9)
            ws_strat.cell(row=sr, column=2, value=killer.get('why_dangerous', '')).font = Font(color="E8E8E8", size=8)
            sr += 1
            ws_strat.cell(row=sr, column=1, value="  Best defense").font = Font(color="32CD32", size=9)
            ws_strat.cell(row=sr, column=2, value=killer.get('best_defense', '')).font = Font(color="32CD32", size=8)
            sr += 1
            ws_strat.cell(row=sr, column=1, value="  ADD TO PAPER").font = Font(color="D4AF37", bold=True, size=9)
            ws_strat.cell(row=sr, column=2, value=killer.get('preemptive_addition', '')).font = Font(color="D4AF37", size=8)
            sr += 2

        # Falsifiable predictions
        preds = strategic_analysis.get('falsifiable_predictions', [])
        if preds:
            ws_strat.cell(row=sr, column=1, value="PREDICTIONS").font = Font(color="FFD700", bold=True, size=11)
            sr += 1
            for p in preds:
                ws_strat.cell(row=sr, column=1, value=f"  {p.get('current_status', '?')}").font = Font(color="FFD700", size=9)
                ws_strat.cell(row=sr, column=2, value=f"{p.get('prediction', '')} — Test: {p.get('test_method', '')}").font = Font(color="E8E8E8", size=8)
                sr += 1

    # ══════════════════════════════════════════════════════════
    # SENTENCE DETAIL SHEET (if OpenAI ran)
    # ══════════════════════════════════════════════════════════
    if openai_scores:
        ws_sent = wb.create_sheet("Sentence Detail")
        ws_sent.column_dimensions['A'].width = 8
        ws_sent.column_dimensions['B'].width = 80
        ws_sent.column_dimensions['C'].width = 12
        ws_sent.column_dimensions['D'].width = 10
        ws_sent.column_dimensions['E'].width = 10
        ws_sent.column_dimensions['F'].width = 10
        ws_sent.column_dimensions['G'].width = 40
        ws_sent.column_dimensions['H'].width = 60
        ws_sent.column_dimensions['I'].width = 40

        # Header
        headers = ['#', 'Sentence', 'Type', 'Bridge', 'Logic', 'Falsif.', 'Theories', 'Objection', 'Missing Cites']
        for ci, h in enumerate(headers, 1):
            c = ws_sent.cell(row=1, column=ci, value=h)
            c.font = Font(color="D4AF37", bold=True, size=9)
            c.fill = header_fill

        for si, (sent, score) in enumerate(zip(sentences, openai_scores)):
            r = si + 2
            ws_sent.cell(row=r, column=1, value=si).font = dim
            ws_sent.cell(row=r, column=2, value=sent[:200]).font = Font(color="E8E8E8", size=8)
            ws_sent.cell(row=r, column=3, value=score.get('type', '')).font = Font(color="FF8C00", size=8)
            ws_sent.cell(row=r, column=4, value=score.get('bridge_score', 0)).font = dim
            ws_sent.cell(row=r, column=4).number_format = '0.00'
            ws_sent.cell(row=r, column=5, value=score.get('logic_score', 0)).font = dim
            ws_sent.cell(row=r, column=5).number_format = '0.00'
            ws_sent.cell(row=r, column=6, value=score.get('falsifiable', 0)).font = dim
            ws_sent.cell(row=r, column=6).number_format = '0.00'
            ws_sent.cell(row=r, column=7, value=', '.join(score.get('theories', []))).font = Font(color="00CED1", size=8)
            ws_sent.cell(row=r, column=8, value=score.get('strongest_objection', '')).font = Font(color="FF6B6B", size=8)
            ws_sent.cell(row=r, column=9, value=', '.join(score.get('missing_citations', []))).font = Font(color="FFD700", size=8)

    # ── Conditional formatting for numeric ranges ──
    max_col = get_column_letter(WORD_COL_START + n_words - 1)
    data_range = f"C5:{max_col}{row}"

    ws.conditional_formatting.add(data_range, CellIsRule(
        operator='greaterThan', formula=['0.05'],
        fill=PatternFill("solid", fgColor="0D2B0D")))  # dark green
    ws.conditional_formatting.add(data_range, CellIsRule(
        operator='lessThan', formula=['-0.05'],
        fill=PatternFill("solid", fgColor="2B0D0D")))  # dark red

    # Freeze panes
    ws.freeze_panes = 'C4'

    # ── Save ──
    wb.save(output_path)
    print(f"  Excel: {output_path}")
    print(f"  {n_words} words × {row} rows = {n_words * row:,} cells")
    if openai_scores:
        print(f"  + Strategic Intel sheet + Sentence Detail sheet")

    # Write strategic markdown alongside Excel
    if strategic_analysis and 'error' not in strategic_analysis:
        md_path = str(Path(output_path).with_suffix('.STRATEGIC.md'))
        write_strategic_markdown(strategic_analysis, md_path)

    return output_path


# ═══════════════════════════════════════════════════════════════
# CLI
# ══════════════════════════════════════════════════════════════��

if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser(description="Word-Level Alignment Matrix")
    parser.add_argument('paper', help='Path to paper .md file')
    parser.add_argument('--output', '-o', help='Output Excel path')
    parser.add_argument('--openai', action='store_true',
                        help='Run OpenAI scoring (channels O-V + strategic intel)')
    args = parser.parse_args()

    print(f"Word-Level Alignment Matrix: {Path(args.paper).name}")
    print("=" * 60)
    if args.openai:
        print("  OpenAI mode: ON (sentence scoring + strategic analysis)")
    write_word_matrix(args.paper, args.output, run_openai=args.openai)

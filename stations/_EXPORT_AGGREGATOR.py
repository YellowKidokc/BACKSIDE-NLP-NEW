"""
EXPORT AGGREGATOR
=================
Walks all 71 station _outbox directories, collects every JSON artifact,
and writes a rich multi-sheet Excel workbook to 10_EXPORTS.

Sheets:
  1. Summary          — one row per artifact (station, status, timing, word count)
  2. Classifications  — every label/score from classify + zero_shot calls
  3. Contradictions   — contradiction pairs with scores
  4. Claims           — extracted claims (axioms, citations, claim-spine)
  5. Summaries        — NLP summaries + key bullets per document
  6. Embeddings       — cluster assignments, similarity edges, node previews
  7. NER & Metadata   — named entities, authors, organizations, dates
  8. All Raw          — every data key/value flattened (fallback / debug)
"""
import json, sys, traceback
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed — run: pip install openpyxl")
    sys.exit(1)

# ── paths ─────────────────────────────────────────────────────────────────────
STATIONS_DIR = Path(__file__).parent
NAS_EXPORTS = Path(r"\\192.168.2.50\brain\10_EXPORTS")
LOCAL_EXPORTS = STATIONS_DIR.parent / "exports"
EXPORTS = NAS_EXPORTS if NAS_EXPORTS.exists() else LOCAL_EXPORTS
EXPORTS.mkdir(parents=True, exist_ok=True)
PROJECT_ROOT = STATIONS_DIR.parent
AAA_ROOT = PROJECT_ROOT / "AAA"
AAA_TEMPLATES = AAA_ROOT / "templates"
AAA_EXPORTS = AAA_ROOT / "exports"
AAA_TEMPLATE = AAA_TEMPLATES / "BRAIN_EXPORT_TEMPLATE.xlsx"
AAA_ROOT.mkdir(parents=True, exist_ok=True)
AAA_TEMPLATES.mkdir(parents=True, exist_ok=True)
AAA_EXPORTS.mkdir(parents=True, exist_ok=True)

# ── style helpers ─────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="2F4F8F")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL    = PatternFill("solid", fgColor="EEF2FB")
PASS_FILL   = PatternFill("solid", fgColor="D4EDDA")
FAIL_FILL   = PatternFill("solid", fgColor="F8D7DA")
WRAP_ALIGN  = Alignment(wrap_text=True, vertical="top")
TOP_ALIGN   = Alignment(vertical="top")
THIN_SIDE   = Side(style="thin", color="CCCCCC")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)


def style_header(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 18


def autofit(ws, min_w=10, max_w=60):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        w = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col_letter].width = max(min_w, min(max_w, w + 2))


def alt_rows(ws, start=2, col_count=None):
    for i, row in enumerate(ws.iter_rows(min_row=start)):
        fill = ALT_FILL if i % 2 == 0 else None
        for cell in row:
            if fill:
                cell.fill = fill
            cell.alignment = TOP_ALIGN
            cell.border = THIN_BORDER


def _safe_str(v):
    if v is None:
        return ""
    if isinstance(v, (dict, list)):
        try:
            return json.dumps(v, ensure_ascii=False, default=str)
        except Exception:
            return str(v)
    return str(v)


def _cap(v, max_len=32000):
    s = _safe_str(v)
    return s if len(s) <= max_len else s[:max_len - 1] + "…"


def _first_non_empty(*vals):
    for v in vals:
        s = _safe_str(v).strip()
        if s:
            return s
    return ""


def _word_count(v):
    return len(_safe_str(v).split())


def _pick_version_text(versions, aliases):
    if not isinstance(versions, dict):
        return ""
    for alias in aliases:
        if alias in versions:
            candidate = versions[alias]
            if isinstance(candidate, dict):
                for key in ("text", "content", "markdown"):
                    if candidate.get(key):
                        return _safe_str(candidate[key])
            else:
                s = _safe_str(candidate).strip()
                if s:
                    return s
    return ""


def _claim_count(data):
    keys = ("claims", "axioms", "extracted_answers", "claim_extraction", "claim_items")
    total = 0
    for key in keys:
        val = data.get(key)
        if isinstance(val, list):
            total += len(val)
    return total


def _evidence_count(data):
    keys = ("evidence", "evidence_map", "evidence_matches", "supporting_evidence")
    total = 0
    for key in keys:
        val = data.get(key)
        if isinstance(val, list):
            total += len(val)
    if isinstance(data.get("evidence_map"), dict):
        total += len(data.get("evidence_map", {}))
    return total


def _contradiction_count(data):
    keys = ("tensions", "contradictions", "contradictions_found", "flow_issues", "incoherent_transitions")
    total = 0
    for key in keys:
        val = data.get(key)
        if isinstance(val, list):
            total += len(val)
        elif key == "contradictions_found":
            try:
                total += int(val)
            except Exception:
                pass
    return total


def _flatten_sections(data):
    sections = data.get("sections")
    if isinstance(sections, list) and sections:
        for i, sec in enumerate(sections):
            if not isinstance(sec, dict):
                continue
            sec_id = (
                _safe_str(sec.get("id") or sec.get("section_id") or sec.get("name") or f"section_{i+1}")
            )
            sec_text = _first_non_empty(sec.get("text"), sec.get("content"), sec.get("markdown"), sec.get("body"))
            sec_heading = _safe_str(sec.get("heading") or sec.get("title")).strip()
            if sec_heading:
                if sec_text:
                    sec_text = f"## {sec_heading}\n\n{sec_text}"
                else:
                    sec_text = f"## {sec_heading}"
            yield (sec_id, sec_text, sec.get("versions", {}), sec.get("data", {}))
    else:
        yield ("section_0", "", {}, {})


def _iter_payload_rows(artifacts):
    for art in artifacts:
        data = art.get("data", {})
        station = art.get("station_name", art.get("_station_dir", ""))
        source_file = art.get("input_file", art.get("_source_file", ""))
        article_id = _first_non_empty(
            data.get("article_id"), data.get("id"), data.get("title"), source_file
        )
        source_text = _first_non_empty(
            data.get("source_markdown"),
            data.get("markdown"),
            data.get("text_clean"),
            data.get("text"),
            data.get("content"),
            data.get("raw_text"),
            source_file,
        )
        versions = data.get("versions", {})
        academic_text = _first_non_empty(
            _pick_version_text(versions, ["academic"]),
            data.get("academic"),
            data.get("academic_text"),
            source_text
        )
        grade8_text = _first_non_empty(
            _pick_version_text(versions, ["grade_8", "easy"]),
            data.get("grade_8"),
            data.get("grade_8_text"),
            _pick_version_text(versions, ["standard"]),
            source_text,
        )
        lossless_text = _first_non_empty(
            data.get("lossless_summary"),
            data.get("lossless"),
            data.get("summary"),
            data.get("document_summary"),
            data.get("article_summary"),
            data.get("combined_summary"),
        )
        source_wc = _word_count(source_text)
        grade8_wc = _word_count(grade8_text)
        academic_wc = _word_count(academic_text)
        lossless_wc = _word_count(lossless_text)
        claims_count = _claim_count(data)
        evidence_count = _evidence_count(data)
        contradiction_count = _contradiction_count(data)
        section_count = len(list(_flatten_sections(data)))
        section_rows = list(_flatten_sections(data))

        for section_id, section_text, section_versions, section_data in section_rows:
            if section_text:
                section_source = section_text
            else:
                section_source = source_text

            section_grade8 = _first_non_empty(
                _pick_version_text(section_versions, ["grade_8", "easy"]),
                grade8_text
            )
            section_academic = _first_non_empty(
                _pick_version_text(section_versions, ["academic"]),
                academic_text
            )

            section_stats = {
                "source_tokens": _word_count(section_source),
                "grade8_tokens": _word_count(section_grade8),
                "academic_tokens": _word_count(section_academic),
                "lossless_tokens": lossless_wc,
                "compression_ratio": round(lossless_wc / max(_word_count(section_source), 1), 4),
                "claims_count": claims_count,
                "evidence_count": evidence_count,
                "contradiction_count": contradiction_count,
                "section_count": section_count,
            }
            section_stats["section_count"] = data.get("section_count", section_stats["section_count"])

            payload = {
                "article_id": article_id,
                "station": station,
                "input_file": source_file,
                "section_id": section_id,
                "source_markdown": section_source,
                "rewrites": {
                    "grade_8_markdown": section_grade8,
                    "academic_markdown": section_academic,
                },
                "lossless_summary": lossless_text,
                "stats": section_stats,
            }

            payload_md = [
                f"# {article_id}" if article_id else "# Section Payload",
                f"**Station:** {station}",
                f"**Input File:** {source_file}",
                f"**Section:** {section_id}",
                "",
                "## Source",
                section_source,
                "",
                "## Grade 8 rewrite",
                section_grade8,
                "",
                "## Academic rewrite",
                section_academic,
                "",
                "## Lossless summary",
                lossless_text,
            ]
            payload_json = json.dumps(payload, ensure_ascii=False, default=str)
            markdown_payload = "\n".join(payload_md)

            yield [
                article_id,
                station,
                source_file,
                section_id,
                _cap(grade8_text),
                _cap(academic_text),
                _cap(payload_json, 20000),
                _cap(markdown_payload, 30000),
            ]
def add_freeze(ws, cell="A2"):
    ws.freeze_panes = cell


# ── artifact loader ───────────────────────────────────────────────────────────

def load_artifacts():
    artifacts = []
    for station_dir in sorted(STATIONS_DIR.iterdir()):
        if not station_dir.is_dir() or station_dir.name.startswith("_"):
            continue
        outbox = station_dir / "_outbox"
        if not outbox.exists():
            continue
        for art_file in sorted(outbox.glob("*.json"), key=lambda f: f.stat().st_mtime):
            try:
                raw = json.loads(art_file.read_text(encoding="utf-8", errors="replace"))
                raw["_source_file"] = art_file.name
                raw["_station_dir"] = station_dir.name
                artifacts.append(raw)
            except Exception:
                pass
    return artifacts


# ── helper: flatten nested dict to depth 2 ───────────────────────────────────

def flatten(d, prefix="", sep=".", depth=2):
    out = {}
    for k, v in d.items():
        key = f"{prefix}{sep}{k}" if prefix else k
        if isinstance(v, dict) and depth > 0:
            out.update(flatten(v, key, sep, depth - 1))
        elif isinstance(v, list) and depth > 0:
            out[key] = "; ".join(str(x) if not isinstance(x, dict) else json.dumps(x, ensure_ascii=False)[:120] for x in v[:10])
        else:
            out[key] = v
    return out


def trunc(v, n=400):
    s = str(v) if v is not None else ""
    return s[:n] + ("…" if len(s) > n else "")


# ── Sheet 1: Summary ──────────────────────────────────────────────────────────

def build_summary(wb, artifacts):
    ws = wb.create_sheet("Summary")
    headers = ["Station", "Input File", "Success", "Worker Used", "Processed At",
               "Data Keys", "Error Count", "Station Errors"]
    ws.append(headers)
    style_header(ws)

    for art in artifacts:
        data = art.get("data", {})
        errors = art.get("errors", [])
        success = art.get("success", True)
        row = [
            art.get("station_name", art.get("_station_dir", "")),
            art.get("input_file", art.get("_source_file", "")),
            "YES" if success else "NO",
            art.get("nlp_used", ""),
            art.get("processed_at", ""),
            ", ".join(list(data.keys())[:8]) if data else "",
            len(errors),
            "; ".join(str(e)[:100] for e in errors[:3]),
        ]
        ws.append(row)
        r = ws.max_row
        ws.cell(r, 3).fill = PASS_FILL if success else FAIL_FILL

    add_freeze(ws)
    autofit(ws)
    alt_rows(ws)
    ws.sheet_properties.tabColor = "2F4F8F"
    return ws


# ── Sheet 2: Classifications ──────────────────────────────────────────────────

def build_classifications(wb, artifacts):
    ws = wb.create_sheet("Classifications")
    headers = ["Station", "Input File", "Top Label", "Top Score", "2nd Label", "2nd Score",
               "3rd Label", "3rd Score", "All Labels (Top 5)", "Text Preview"]
    ws.append(headers)
    style_header(ws)

    for art in artifacts:
        data = art.get("data", {})
        station = art.get("station_name", "")

        # hunt for classification data in various shapes
        label_lists = []
        if "seven_q_scores" in data:
            label_lists.append(data["seven_q_scores"])
        if "classification" in data and "labels" in data["classification"]:
            z = data["classification"]
            zipped = list(zip(z.get("labels", []), z.get("scores", [])))
            label_lists.append([{"label": l, "score": s} for l, s in zipped])
        if "classified_sentences" in data:
            for item in data["classified_sentences"][:5]:
                label_lists.append([{"label": item.get("type", item.get("argument_type", "")),
                                      "score": item.get("confidence", 0)}])
        if "fcard" in data:
            fc = data["fcard"]
            label_lists.append([{"label": fc.get("category", ""), "score": fc.get("confidence", 0)}])
        if "file_type" in data:
            label_lists.append([{"label": data.get("file_type", ""), "score": data.get("confidence", 0)}])
        if "dominant_fruit" in data:
            present = data.get("fruits_present", {})
            label_lists.append([{"label": f, "score": v.get("score", 0) if isinstance(v, dict) else (max(x.get("score",0) for x in v) if isinstance(v, list) and v else 0)} for f, v in present.items()])
        if "dominant_operator" in data:
            dist = data.get("type_distribution", {})
            total = max(sum(dist.values()), 1)
            label_lists.append([{"label": k, "score": v / total} for k, v in sorted(dist.items(), key=lambda x: -x[1])])
        if "persons_addressed" in data:
            pm = data.get("person_map", {})
            label_lists.append([{"label": k, "score": v if isinstance(v, (int, float)) else len(v)} for k, v in pm.items()])

        for labels in label_lists:
            if not labels:
                continue
            labels_sorted = sorted(labels, key=lambda x: float(x.get("score", 0)), reverse=True)[:5]
            preview = trunc(art.get("input_file", ""), 60)
            row = [station, art.get("input_file", "")]
            for i in range(3):
                if i < len(labels_sorted):
                    row += [labels_sorted[i].get("label", ""), round(float(labels_sorted[i].get("score", 0)), 4)]
                else:
                    row += ["", ""]
            row += ["; ".join(f"{l.get('label','')} ({round(float(l.get('score',0)),3)})" for l in labels_sorted), preview]
            ws.append(row)

    add_freeze(ws)
    autofit(ws)
    alt_rows(ws)
    ws.sheet_properties.tabColor = "4472C4"
    return ws


# ── Sheet 3: Contradictions ───────────────────────────────────────────────────

def build_contradictions(wb, artifacts):
    ws = wb.create_sheet("Contradictions")
    headers = ["Station", "Input File", "Text A (preview)", "Text B (preview)",
               "Contradiction Score", "Entailment Score", "Neutral Score", "Severity"]
    ws.append(headers)
    style_header(ws)

    HIGH = PatternFill("solid", fgColor="F8D7DA")
    MED  = PatternFill("solid", fgColor="FFF3CD")

    for art in artifacts:
        data = art.get("data", {})
        station = art.get("station_name", "")
        src = art.get("input_file", "")

        pairs = []
        if "tensions" in data:
            for t in data["tensions"]:
                pairs.append({"text_a": t.get("premise_a", t.get("axiom_a", "")),
                               "text_b": t.get("premise_b", t.get("axiom_b", "")),
                               "contradiction": t.get("contradiction_score", t.get("contradiction", 0)),
                               "entailment": t.get("entailment", 0)})
        if "incoherent_transitions" in data:
            for t in data["incoherent_transitions"]:
                pairs.append({"text_a": t.get("para_a", ""), "text_b": t.get("para_b", ""),
                               "contradiction": t.get("contradiction_score", 0), "entailment": 0})
        if "flow_issues" in data:
            for t in data["flow_issues"]:
                pairs.append({"text_a": str(t.get("from_id", "")), "text_b": str(t.get("to_id", "")),
                               "contradiction": t.get("contradiction", 0), "entailment": t.get("entailment", 0),
                               "severity": t.get("severity", "")})
        if "contradictions_found" in data and "dimensions" in data:
            pairs.append({"text_a": f"{data.get('contradictions_found', 0)} internal contradictions",
                           "text_b": f"TSR Score: {data.get('tsr_score', '')}",
                           "contradiction": min(1.0, data.get("contradictions_found", 0) / 10.0),
                           "entailment": 0})

        for p in pairs:
            con = float(p.get("contradiction", 0))
            ent = float(p.get("entailment", 0))
            neu = round(1.0 - con - ent, 4)
            severity = p.get("severity") or ("HIGH" if con > 0.7 else ("MEDIUM" if con > 0.4 else "LOW"))
            row = [station, src, trunc(p.get("text_a", ""), 200), trunc(p.get("text_b", ""), 200),
                   round(con, 4), round(ent, 4), max(0, neu), severity]
            ws.append(row)
            r_num = ws.max_row
            ws.cell(r_num, 5).fill = HIGH if con > 0.6 else (MED if con > 0.35 else PatternFill())

    add_freeze(ws)
    autofit(ws)
    ws.sheet_properties.tabColor = "C00000"
    return ws


# ── Sheet 4: Claims ───────────────────────────────────────────────────────────

def build_claims(wb, artifacts):
    ws = wb.create_sheet("Claims")
    headers = ["Station", "Input File", "Claim ID", "Claim Text", "Claim Type",
               "Confidence", "Entities", "Source"]
    ws.append(headers)
    style_header(ws)

    for art in artifacts:
        data = art.get("data", {})
        station = art.get("station_name", "")
        src = art.get("input_file", "")

        claim_lists = []
        if "claims" in data:
            claim_lists.extend(data["claims"] if isinstance(data["claims"], list) else [])
        if "axioms" in data:
            for a in data["axioms"]:
                claim_lists.append({"text": a.get("text", ""), "claim_type": "axiom",
                                     "confidence": a.get("confidence", 0)})
        if "extracted_answers" in data:
            for a in data["extracted_answers"]:
                claim_lists.append({"text": a.get("answer", ""), "claim_type": "extracted",
                                     "confidence": a.get("confidence", 0), "question": a.get("question", "")})

        for c in claim_lists:
            if not c.get("text"):
                continue
            ws.append([
                station, src,
                c.get("claim_id", c.get("id", "")),
                trunc(c.get("text", ""), 300),
                c.get("claim_type", c.get("type", "")),
                round(float(c.get("confidence", c.get("score", 0))), 4),
                "; ".join(c.get("entities", [])[:5]) if isinstance(c.get("entities"), list) else "",
                c.get("question", c.get("source", "")),
            ])

    add_freeze(ws)
    autofit(ws)
    alt_rows(ws)
    ws.sheet_properties.tabColor = "70AD47"
    return ws


# ── Sheet 5: Summaries ────────────────────────────────────────────────────────

def build_summaries(wb, artifacts):
    ws = wb.create_sheet("Summaries")
    headers = ["Station", "Input File", "Document Summary", "Section Count",
               "Word Count", "Reading Grade", "Key Bullets (preview)"]
    ws.append(headers)
    style_header(ws)

    for art in artifacts:
        data = art.get("data", {})
        station = art.get("station_name", "")
        src = art.get("input_file", "")

        summary = (data.get("article_summary") or data.get("input_summary") or
                   data.get("combined_summary") or data.get("full_summary") or
                   (data.get("fcard", {}) or {}).get("summary") or "")
        if not summary:
            sec_sums = data.get("section_summaries", [])
            if sec_sums:
                summary = "; ".join(s.get("summary", "") for s in sec_sums[:3])
        if not summary:
            publication = data.get("publication_path", "")
            if publication:
                summary = f"→ {publication}"

        bullets = data.get("lossless_summary", data.get("summary_bullets", []))
        bullets_str = "; ".join(str(b) for b in bullets[:5]) if isinstance(bullets, list) else ""

        sec_count = data.get("section_count", len(data.get("sections", [])) or len(data.get("section_summaries", [])))
        word_count = data.get("word_count", (data.get("fcard", {}) or {}).get("word_count", ""))
        grade = data.get("reading_grade", data.get("grade", ""))

        if summary or bullets_str:
            ws.append([station, src, trunc(summary, 400), sec_count or "", word_count or "", grade, trunc(bullets_str, 300)])
            ws.cell(ws.max_row, 3).alignment = WRAP_ALIGN

    add_freeze(ws)
    ws.column_dimensions["C"].width = 60
    autofit(ws, min_w=8, max_w=80)
    ws.column_dimensions["C"].width = 60
    alt_rows(ws)
    ws.sheet_properties.tabColor = "ED7D31"
    return ws


# ── Sheet 6: Embeddings & Clusters ───────────────────────────────────────────

def build_embeddings(wb, artifacts):
    ws = wb.create_sheet("Embeddings & Clusters")
    headers = ["Station", "Input File", "Cluster/Node ID", "Preview Text",
               "Connected To", "Edge Weight", "Cluster Size", "Vector Dim"]
    ws.append(headers)
    style_header(ws)

    for art in artifacts:
        data = art.get("data", {})
        station = art.get("station_name", "")
        src = art.get("input_file", "")

        # Clusters
        if "clusters" in data:
            for cl in data["clusters"][:20]:
                members = cl.get("members", [])
                ws.append([station, src, f"Cluster {cl.get('cluster_id', '')}",
                            trunc(cl.get("centroid_preview", ""), 200),
                            "; ".join(str(m.get("idx", "")) for m in members[:6]),
                            "", cl.get("size", ""), data.get("vector_dim", "")])

        # Graph edges
        if "edges" in data and "nodes" in data:
            for edge in data["edges"][:20]:
                src_label = edge.get("source_label") or edge.get("source", "")
                tgt_label = edge.get("target_label") or edge.get("target", "")
                ws.append([station, src, str(src_label),
                            trunc(str(src_label), 150),
                            str(tgt_label), edge.get("weight", ""),
                            "", data.get("vector_dim", "")])

        # Link index
        if "link_index" in data:
            for link in data["link_index"][:15]:
                ws.append([station, src, f"Link {link.get('vector_idx', '')}",
                            trunc(link.get("text", ""), 200), "", "", "", ""])

    add_freeze(ws)
    autofit(ws)
    alt_rows(ws)
    ws.sheet_properties.tabColor = "7030A0"
    return ws


# ── Sheet 7: NER & Metadata ───────────────────────────────────────────────────

def build_ner(wb, artifacts):
    ws = wb.create_sheet("NER & Metadata")
    headers = ["Station", "Input File", "Entity Text", "Entity Type",
               "Category", "Confidence", "Role", "Context"]
    ws.append(headers)
    style_header(ws)

    for art in artifacts:
        data = art.get("data", {})
        station = art.get("station_name", "")
        src = art.get("input_file", "")

        entities = []
        # Standard NER list
        if "entities" in data:
            entities = data["entities"]
        # Metadata block
        if "metadata" in data:
            meta = data["metadata"]
            for person in meta.get("authors", []):
                entities.append({"text": person, "label": "PER", "role": "author"})
            for org in meta.get("organizations", []):
                entities.append({"text": org, "label": "ORG"})
            for date in meta.get("dates", []):
                entities.append({"text": date, "label": "DATE"})
        # Key entities list
        if "key_entities" in data:
            for e in data["key_entities"]:
                entities.append({"text": e, "label": "KEY"})

        for ent in entities[:30]:
            ws.append([
                station, src,
                trunc(ent.get("text", str(ent)), 120),
                ent.get("label", ent.get("entity_type", "")),
                ent.get("category", ""),
                round(float(ent.get("score", ent.get("confidence", 0))), 4) if ent.get("score") or ent.get("confidence") else "",
                ent.get("role", ""),
                trunc(ent.get("context", ""), 150),
            ])

    add_freeze(ws)
    autofit(ws)
    alt_rows(ws)
    ws.sheet_properties.tabColor = "FF0000"
    return ws


# ── Sheet 8: All Raw ──────────────────────────────────────────────────────────

def build_all_raw(wb, artifacts):
    ws = wb.create_sheet("All Raw")
    headers = ["Station", "Input File", "Success", "Worker", "Processed At", "Key", "Value"]
    ws.append(headers)
    style_header(ws)

    for art in artifacts:
        station = art.get("station_name", "")
        src = art.get("input_file", "")
        success = "YES" if art.get("success", True) else "NO"
        worker = art.get("nlp_used", "")
        ts = art.get("processed_at", "")
        data = art.get("data", {})
        flat = flatten(data)
        if not flat:
            ws.append([station, src, success, worker, ts, "(empty data)", ""])
        else:
            for k, v in list(flat.items())[:40]:
                ws.append([station, src, success, worker, ts, k, trunc(str(v), 400)])

    add_freeze(ws)
    autofit(ws)
    ws.sheet_properties.tabColor = "808080"
    return ws


# ── Sheet 9: AAA website payload (compact one-row-per-section output) ────────

def build_aaa_payload(wb, artifacts):
    ws = wb.create_sheet("AAA_Website_Payload")
    headers = [
        "Article ID",
        "Station",
        "Input File",
        "Section ID",
        "Grade 8 Markdown",
        "Academic Markdown",
        "Website Payload JSON",
        "Website Payload Markdown",
    ]
    ws.append(headers)
    style_header(ws)

    for row in _iter_payload_rows(artifacts):
        ws.append(row)

    # wide JSON blobs can be long; keep row heights readable by wrapping.
    for r in ws.iter_rows(min_row=2, max_col=8):
        for c in r:
            c.alignment = WRAP_ALIGN
            c.border = THIN_BORDER

    if ws.max_row == 1:
        ws.append(["", "", "", "", "", "", json.dumps({"status": "no artifacts"}, ensure_ascii=False), ""])

    add_freeze(ws)
    autofit(ws, min_w=10, max_w=80)
    ws.sheet_properties.tabColor = "1F4E78"
    return ws


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("Loading artifacts from all station _outbox directories...")
    artifacts = load_artifacts()
    print(f"  Found {len(artifacts)} artifacts from {len(set(a.get('station_name','') for a in artifacts))} stations")

    if not artifacts:
        print("No artifacts found — run stations first.")
        return

    if AAA_TEMPLATE.exists():
        wb = openpyxl.load_workbook(AAA_TEMPLATE)
        # Keep template structure, then ensure the payload sheet is refreshed.
        for tmpl_sheet in ("AAA_Website_Payload",):
            if tmpl_sheet in wb.sheetnames:
                ws = wb[tmpl_sheet]
                if ws.max_row:
                    ws.delete_rows(1, ws.max_row)
    else:
        wb = openpyxl.Workbook()
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    print("Building sheets...")
    build_summary(wb, artifacts)
    print("  [1/8] Summary done")
    build_classifications(wb, artifacts)
    print("  [2/8] Classifications done")
    build_contradictions(wb, artifacts)
    print("  [3/8] Contradictions done")
    build_claims(wb, artifacts)
    print("  [4/8] Claims done")
    build_summaries(wb, artifacts)
    print("  [5/8] Summaries done")
    build_embeddings(wb, artifacts)
    print("  [6/8] Embeddings & Clusters done")
    build_ner(wb, artifacts)
    print("  [7/8] NER & Metadata done")
    build_all_raw(wb, artifacts)
    print("  [8/8] All Raw done")
    build_aaa_payload(wb, artifacts)
    print("  [9/9] AAA payload done")

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_name = f"BRAIN_EXPORT_{stamp}.xlsx"
    out_path = EXPORTS / out_name
    wb.save(str(out_path))

    aaa_copy = AAA_EXPORTS / out_name
    if out_path != aaa_copy:
        wb.save(str(aaa_copy))
    print(f"\nExport saved: {out_path}")
    if aaa_copy != out_path:
        print(f"  Also saved AAA copy: {aaa_copy}")
    print(f"  Artifacts processed: {len(artifacts)}")
    print(f"  Stations covered:    {len(set(a.get('station_name','') for a in artifacts))}")


if __name__ == "__main__":
    main()

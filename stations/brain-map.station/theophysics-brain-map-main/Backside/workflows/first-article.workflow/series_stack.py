from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SENTENCE_RE = re.compile(r"(?<=[.!?])\s+")
EQUATION_BLOCK_RE = re.compile(r"```text\s*(.*?)\s*```", re.DOTALL)
CLAIM_MARKER_RE = re.compile(r"\b(claim|therefore|thus|requires|implies|proves|shows|demonstrates|predicts|falsif|evidence|collapse|coherence|entropy|grace|truth|faith|trinity)\b", re.I)
NEGATION_RE = re.compile(r"\b(not|never|cannot|can't|no longer|fails|false|contradict|refute|collapse)\b", re.I)
AFFIRMATION_RE = re.compile(r"\b(is|are|does|can|must|requires|proves|shows|demonstrates|holds|confirms|supports)\b", re.I)
TERMS = [
    "grace",
    "entropy",
    "coherence",
    "collapse",
    "measurement",
    "observer",
    "time",
    "free will",
    "trinity",
    "truth",
    "faith",
    "decoherence",
    "law",
    "information",
    "genesis",
]


def read_text(path: Path) -> str:
    return path.read_text(encoding="utf-8-sig", errors="replace")


def write_text(path: Path, text: str) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text.rstrip() + "\n", encoding="utf-8")
    return path


def load_batch(batch_root: Path) -> list[dict[str, Any]]:
    csv_path = batch_root / "batch-results.csv"
    rows: list[dict[str, Any]] = []
    with csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            if row.get("export_path") and Path(row["export_path"]).exists():
                rows.append(row)
    rows.sort(key=lambda row: int(row.get("ordinal") or 0))
    return rows


def find_one(root: Path, name: str) -> Path | None:
    matches = list(root.rglob(name))
    return matches[0] if matches else None


def find_lossless_json(export_dir: Path) -> Path | None:
    lossless = export_dir / "lossless"
    matches = sorted(lossless.glob("*.json")) if lossless.exists() else []
    return matches[0] if matches else None


def compact_sentences(text: str, limit: int = 8) -> list[str]:
    compact = re.sub(r"\s+", " ", text).strip()
    sentences = [s.strip() for s in SENTENCE_RE.split(compact) if len(s.strip()) > 45]
    scored: list[tuple[int, str]] = []
    for sentence in sentences:
        score = 0
        score += 3 if CLAIM_MARKER_RE.search(sentence) else 0
        score += 2 if any(term in sentence.lower() for term in TERMS) else 0
        score += 1 if len(sentence) < 260 else 0
        if score:
            scored.append((score, sentence))
    scored.sort(key=lambda item: (-item[0], len(item[1])))
    return [sentence for _, sentence in scored[:limit]]


def article_records(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for row in rows:
        export_dir = Path(row["export_path"])
        manifest = json.loads((export_dir / "manifest.json").read_text(encoding="utf-8"))
        source_md = export_dir / "source.canonical.md"
        summary_path = find_one(export_dir / "stations", "executive-summary.md")
        math_path = find_one(export_dir / "stations", "math-layer.md")
        overview_path = find_one(export_dir / "stations", "overview.md")
        lossless_path = find_lossless_json(export_dir)
        source_text = read_text(source_md) if source_md.exists() else ""
        summary_text = read_text(summary_path) if summary_path else ""
        math_text = read_text(math_path) if math_path else ""
        overview_text = read_text(overview_path) if overview_path else ""
        lossless = json.loads(read_text(lossless_path)) if lossless_path else {}
        records.append(
            {
                "ordinal": int(row["ordinal"]),
                "name": row["name"],
                "export_dir": str(export_dir),
                "manifest": manifest,
                "address": manifest.get("address", row.get("address", "")),
                "vector": manifest.get("vector", row.get("vector", "")),
                "hash": manifest.get("hash", row.get("hash", "")),
                "master_equation_uuid": manifest.get("master_equation_uuid", lossless.get("master_equation_uuid", "")),
                "semantic_tag_count": manifest.get("semantic_tag_count", len(lossless.get("semantic_tags", []))),
                "source_text": source_text,
                "summary_text": summary_text,
                "math_text": math_text,
                "overview_text": overview_text,
                "lossless": lossless,
                "claim_sentences": compact_sentences(source_text, 10),
            }
        )
    return records


def render_simple_summary(records: list[dict[str, Any]]) -> str:
    lines = [
        "# GTQ Simple Series Summary",
        "",
        f"This stack covers {len(records)} Genesis-to-Quantum root articles.",
        "",
        "In plain terms: the series keeps returning to one core idea: Genesis is being read as a physics-facing structure, where collapse, coherence, time, grace, sin, observation, and restoration are not isolated metaphors but linked parts of one framework.",
        "",
        "## Article One-Liners",
        "",
    ]
    for record in records:
        title = title_from_record(record)
        one = record["claim_sentences"][0] if record["claim_sentences"] else first_nonempty_sentence(record["source_text"])
        lines.append(f"- **{record['ordinal']:02d}. {title}:** {one}")
    return "\n".join(lines)


def render_executive_summary(records: list[dict[str, Any]]) -> str:
    vector_counts = Counter(record["vector"] for record in records)
    term_counts = Counter()
    for record in records:
        low = record["source_text"].lower()
        for term in TERMS:
            if term in low:
                term_counts[term] += 1

    lines = [
        "# GTQ Cumulative Executive Summary",
        "",
        "## Decision-Grade Readout",
        "",
        f"The Genesis-to-Quantum root series produced {len(records)} successful workflow packets. Each packet includes canonical Markdown, executive summary, overview, math layer, image notes, lossless JSON/HTML, and a manifest.",
        "",
        "The dominant artifact shape is high-authority, procedural/mechanistic, knowledge-bearing, relational, faith/trust-bearing, and coherence/integration oriented. A second cluster adds `E=3`, which should be reviewed as either legitimate artifact-level disorder/collapse structure or possible topic drift.",
        "",
        "## Vector Distribution",
        "",
    ]
    for vector, count in vector_counts.most_common():
        lines.append(f"- `{vector}`: {count}")
    lines.extend(["", "## Recurring Terms", ""])
    for term, count in term_counts.most_common():
        lines.append(f"- {term}: {count}/{len(records)}")
    lines.extend(["", "## Cumulative Spine", ""])
    spine = cumulative_spine(records)
    lines.extend([f"{i}. {item}" for i, item in enumerate(spine, start=1)])
    lines.extend(
        [
            "",
            "## Immediate Review Targets",
            "",
            "- Review all `E=3` classifications for artifact-disorder vs subject-matter drift.",
            "- Promote repeated math candidates into the Math Translation Layer proper.",
            "- Run a stronger contradiction pass after claim extraction becomes its own station.",
            "- Compare GTQ-27 against the rest of the series; its vector is structurally different.",
        ]
    )
    return "\n".join(lines)


def cumulative_spine(records: list[dict[str, Any]]) -> list[str]:
    candidates: list[str] = []
    for record in records:
        if record["claim_sentences"]:
            candidates.append(record["claim_sentences"][0])
    return candidates[:12]


def render_lossless_summary(records: list[dict[str, Any]]) -> tuple[str, list[dict[str, Any]]]:
    rows = []
    lines = [
        "# GTQ Lossless Series Index",
        "",
        "This is not a replacement for each lossless artifact. It is the series-level navigation layer over the per-article reconstruction artifacts.",
        "",
        "| # | Article | Vector | Address | Lossless Claims | Equations | Open Threads |",
        "|---:|---|---|---|---:|---:|---:|",
    ]
    for record in records:
        lossless = record["lossless"]
        row = {
            "ordinal": record["ordinal"],
            "name": record["name"],
            "title": title_from_record(record),
            "address": record["address"],
            "vector": record["vector"],
            "master_equation_uuid": record.get("master_equation_uuid", ""),
            "semantic_tag_count": record.get("semantic_tag_count", 0),
            "claim_count": len(lossless.get("claim_arch", [])),
            "equation_count": len(lossless.get("eq_sem", [])),
            "open_threads": len(lossless.get("open_threads", [])),
            "export_dir": record["export_dir"],
        }
        rows.append(row)
        lines.append(
            f"| {row['ordinal']} | {row['title']} | `{row['vector']}` | `{row['address']}` | {row['claim_count']} | {row['equation_count']} | {row['open_threads']} |"
        )
    return "\n".join(lines), rows


def render_math_translation(records: list[dict[str, Any]]) -> tuple[str, list[dict[str, Any]]]:
    equations: list[dict[str, Any]] = []
    lines = [
        "# GTQ Cumulative Math Translation Layer",
        "",
        "This aggregates the deterministic per-article math-layer station output. It is a staging layer before the full Math-Translation-Layer runtime.",
        "",
    ]
    for record in records:
        found = [re.sub(r"\s+", " ", item).strip() for item in EQUATION_BLOCK_RE.findall(record["math_text"])]
        if not found:
            continue
        lines.extend(["", f"## {record['ordinal']:02d}. {title_from_record(record)}", ""])
        for index, eq in enumerate(found, start=1):
            translation = translate_math(eq)
            equations.append({"article": record["name"], "ordinal": record["ordinal"], "equation_index": index, "equation": eq, "translation": translation})
            lines.extend(
                [
                    f"### M{record['ordinal']:02d}.{index:03d}",
                    "",
                    "```text",
                    eq,
                    "```",
                    "",
                    f"Translation: {translation}",
                    "",
                ]
            )
    if not equations:
        lines.append("No math candidates detected.")
    return "\n".join(lines), equations


def translate_math(eq: str) -> str:
    low = eq.lower()
    if "chi" in low or "χ" in low:
        return "Master/coherence expression. Verify each factor is defined and whether the equation is symbolic, operational, or derived."
    if "entropy" in low or "decoherence" in low:
        return "Entropy/decoherence expression. Check whether disorder attenuates coherence rather than positively multiplying it."
    if "sigma" in low or "σ" in low or "p <" in low or "p<" in low:
        return "Statistical claim. Preserve dataset, baseline, test family, and replication status."
    if "grace" in low:
        return "Grace operator or grace-linked expression. Preserve whether grace is external input, force/operator, or interpretive label."
    if "trinity" in low:
        return "Trinity formalization candidate. Keep formal/proof behavior separate from theological interpretation."
    return "Formal/math-like relation. Define variables, role, derivation status, and computability before public use."


def render_contradictions(records: list[dict[str, Any]]) -> tuple[str, list[dict[str, Any]]]:
    term_sentences: dict[str, dict[str, list[dict[str, Any]]]] = {
        term: {"affirm": [], "negate": []} for term in TERMS
    }
    for record in records:
        for sentence in compact_sentences(record["source_text"], 80):
            low = sentence.lower()
            for term in TERMS:
                if term not in low:
                    continue
                bucket = "negate" if NEGATION_RE.search(sentence) else "affirm" if AFFIRMATION_RE.search(sentence) else ""
                if bucket:
                    term_sentences[term][bucket].append({"article": record["name"], "ordinal": record["ordinal"], "sentence": sentence})

    flags: list[dict[str, Any]] = []
    for term, buckets in term_sentences.items():
        if buckets["affirm"] and buckets["negate"]:
            flags.append(
                {
                    "term": term,
                    "severity": "REVIEW",
                    "affirm_count": len(buckets["affirm"]),
                    "negate_count": len(buckets["negate"]),
                    "affirm_examples": buckets["affirm"][:3],
                    "negate_examples": buckets["negate"][:3],
                }
            )

    lines = [
        "# GTQ Contradiction Scan",
        "",
        "This is a first heuristic contradiction station. It is expected to overflag. It looks for repeated terms that appear in both affirmative and negated/collapse/failure contexts.",
        "",
        f"- Flags: {len(flags)}",
        "",
    ]
    for flag in flags:
        lines.extend(
            [
                f"## {flag['term']}",
                "",
                f"- Affirm contexts: {flag['affirm_count']}",
                f"- Negation/failure contexts: {flag['negate_count']}",
                "",
                "### Affirm Examples",
                "",
            ]
        )
        for example in flag["affirm_examples"]:
            lines.append(f"- `{example['article']}`: {example['sentence']}")
        lines.extend(["", "### Negation / Failure Examples", ""])
        for example in flag["negate_examples"]:
            lines.append(f"- `{example['article']}`: {example['sentence']}")
        lines.extend(["", "Repair: classify whether this is true contradiction, regime distinction, rhetorical contrast, or sequence change.", ""])
    return "\n".join(lines), flags


def render_series_paper(records: list[dict[str, Any]], executive: str, math_layer: str, contradiction: str) -> str:
    return "\n\n".join(
        [
            "# Genesis-to-Quantum Series Stack Paper Draft",
            "",
            "## Abstract",
            f"This draft is generated from {len(records)} processed Genesis-to-Quantum workflow packets. It is not a final paper; it is a cumulative scaffold for turning the series into a coherent reviewable artifact.",
            "",
            executive,
            "",
            "## Math Translation Layer",
            "",
            math_layer[:18000] + ("\n\n[TRUNCATED IN PAPER DRAFT: see cumulative-math-translation.md]" if len(math_layer) > 18000 else ""),
            "",
            "## Contradiction / Regime Review",
            "",
            contradiction[:12000] + ("\n\n[TRUNCATED IN PAPER DRAFT: see contradiction-scan.md]" if len(contradiction) > 12000 else ""),
        ]
    )


def write_workbook(
    out: Path,
    records: list[dict[str, Any]],
    executive: str,
    simple: str,
    lossless_rows: list[dict[str, Any]],
    math_rows: list[dict[str, Any]],
    contradiction_flags: list[dict[str, Any]],
) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"
    vector_counts = Counter(record["vector"] for record in records)
    ws.append(["GTQ Series Stack Workbook"])
    ws.append(["Generated", datetime.now().isoformat(timespec="seconds")])
    ws.append(["Articles", len(records)])
    ws.append(["Math candidates", len(math_rows)])
    ws.append(["Contradiction flags", len(contradiction_flags)])
    ws.append([])
    ws.append(["Vector", "Count"])
    for vector, count in vector_counts.most_common():
        ws.append([vector, count])

    add_text_sheet(wb, "Simple Summary", simple)
    add_text_sheet(wb, "Executive Summary", executive)
    add_table_sheet(
        wb,
        "Articles",
        [
            {
                "ordinal": record["ordinal"],
                "name": record["name"],
                "title": title_from_record(record),
                "vector": record["vector"],
                "hash": record["hash"],
                "master_equation_uuid": record.get("master_equation_uuid", ""),
                "semantic_tag_count": record.get("semantic_tag_count", 0),
                "address": record["address"],
                "export_dir": record["export_dir"],
            }
            for record in records
        ],
    )
    add_table_sheet(wb, "Lossless Index", lossless_rows)
    add_table_sheet(wb, "Math Candidates", math_rows)
    add_table_sheet(
        wb,
        "Contradiction Flags",
        flatten_contradiction_flags(contradiction_flags),
    )
    add_table_sheet(
        wb,
        "Open Threads",
        build_open_thread_rows(records),
    )

    for sheet in wb.worksheets:
        style_sheet(sheet)
    path = out / "gtq-series-stack.xlsx"
    wb.save(path)
    return path


def add_text_sheet(wb: Workbook, title: str, text: str) -> None:
    ws = wb.create_sheet(title[:31])
    ws.append(["section", "text"])
    current = "Document"
    buffer: list[str] = []
    for line in text.splitlines():
        if line.startswith("#"):
            if buffer:
                ws.append([current, "\n".join(buffer).strip()])
                buffer = []
            current = line.strip("# ").strip() or current
        else:
            buffer.append(line)
    if buffer:
        ws.append([current, "\n".join(buffer).strip()])


def add_table_sheet(wb: Workbook, title: str, rows: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet(title[:31])
    if not rows:
        ws.append(["status"])
        ws.append(["no rows"])
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([safe_cell(row.get(header, "")) for header in headers])


def safe_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        text = value
    else:
        text = json.dumps(value, ensure_ascii=False)
    if isinstance(text, str) and len(text) > 32000:
        return text[:31900] + "\n[TRUNCATED]"
    return text


def flatten_contradiction_flags(flags: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for flag in flags:
        rows.append(
            {
                "term": flag.get("term", ""),
                "severity": flag.get("severity", ""),
                "affirm_count": flag.get("affirm_count", 0),
                "negate_count": flag.get("negate_count", 0),
                "affirm_examples": "; ".join(example.get("sentence", "") for example in flag.get("affirm_examples", [])),
                "negate_examples": "; ".join(example.get("sentence", "") for example in flag.get("negate_examples", [])),
                "repair": "classify as true contradiction, regime distinction, rhetorical contrast, or sequence change",
            }
        )
    return rows


def build_open_thread_rows(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for record in records:
        for thread in record["lossless"].get("open_threads", []):
            rows.append(
                {
                    "ordinal": record["ordinal"],
                    "article": record["name"],
                    "title": title_from_record(record),
                    "thread": thread,
                    "export_dir": record["export_dir"],
                }
            )
    return rows


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        max_len = 12
        for cell in column_cells[:80]:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(value), 80))
        ws.column_dimensions[letter].width = min(max_len + 2, 70)


def first_nonempty_sentence(text: str) -> str:
    for sentence in SENTENCE_RE.split(re.sub(r"\s+", " ", text).strip()):
        if len(sentence.strip()) > 45:
            return sentence.strip()
    return ""


def title_from_record(record: dict[str, Any]) -> str:
    address = record.get("address", "")
    if "/" in address:
        try:
            return address.split("/")[1].replace("-", " ").title()
        except Exception:
            pass
    name = Path(record["name"]).stem
    return name.replace("-", " ").title()


def main() -> int:
    parser = argparse.ArgumentParser(description="Stack a first-article workflow batch into cumulative series outputs.")
    parser.add_argument("--batch-root", type=Path, required=True)
    parser.add_argument("--out", type=Path)
    args = parser.parse_args()

    batch_root = args.batch_root
    out = args.out or (batch_root / "STACK")
    out.mkdir(parents=True, exist_ok=True)

    records = article_records(load_batch(batch_root))
    simple = render_simple_summary(records)
    executive = render_executive_summary(records)
    lossless_md, lossless_rows = render_lossless_summary(records)
    math_md, math_rows = render_math_translation(records)
    contradiction_md, contradiction_flags = render_contradictions(records)
    paper = render_series_paper(records, executive, math_md, contradiction_md)

    write_text(out / "simple-summary.md", simple)
    write_text(out / "cumulative-executive-summary.md", executive)
    write_text(out / "lossless-series-index.md", lossless_md)
    write_text(out / "cumulative-math-translation.md", math_md)
    write_text(out / "contradiction-scan.md", contradiction_md)
    write_text(out / "series-paper-draft.md", paper)
    (out / "lossless-series-index.json").write_text(json.dumps(lossless_rows, indent=2), encoding="utf-8")
    (out / "math-candidates.json").write_text(json.dumps(math_rows, indent=2), encoding="utf-8")
    (out / "contradiction-flags.json").write_text(json.dumps(contradiction_flags, indent=2), encoding="utf-8")
    workbook_path = write_workbook(out, records, executive, simple, lossless_rows, math_rows, contradiction_flags)
    manifest = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "batch_root": str(batch_root),
        "out": str(out),
        "articles": len(records),
        "math_candidates": len(math_rows),
        "contradiction_flags": len(contradiction_flags),
        "workbook": str(workbook_path),
        "files": [
            "simple-summary.md",
            "cumulative-executive-summary.md",
            "lossless-series-index.md",
            "cumulative-math-translation.md",
            "contradiction-scan.md",
            "series-paper-draft.md",
            "gtq-series-stack.xlsx",
        ],
    }
    (out / "stack-manifest.json").write_text(json.dumps(manifest, indent=2), encoding="utf-8")
    print(f"Stack written: {out}")
    print(f"Articles: {len(records)}")
    print(f"Math candidates: {len(math_rows)}")
    print(f"Contradiction flags: {len(contradiction_flags)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

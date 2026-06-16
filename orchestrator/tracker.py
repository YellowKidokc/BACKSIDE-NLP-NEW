from __future__ import annotations

import argparse
import importlib.util
import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[1]
WORKFLOWS = ROOT / "03_WORKFLOWS"
if not WORKFLOWS.exists():
    WORKFLOWS = ROOT / "workflows"
TRAVELERS = WORKFLOWS / "travelers"
TRACKER_XLSX = WORKFLOWS / "PIPELINE_TRACKER.xlsx"

CANONICAL_STATIONS = [
    "sbert-embedder",
    "classify-documents",
    "metadata-extractor",
    "claim-extractor",
    "7q-classifier",
    "paper-proof-grader",
    "reading-level-glossary",
    "readability-rewriter",
    "fruits-spirit-canon",
    "html-article",
]

WORKFLOW_STATIONS = {
    "first-article": CANONICAL_STATIONS,
    "quick-classify": ["sbert-embedder", "classify-documents", "metadata-extractor"],
    "mda-full": CANONICAL_STATIONS,
}

STATUS_COLUMNS = {
    "sbert": "sbert-embedder",
    "classify": "classify-documents",
    "7q": "7q-classifier",
    "claims": "claim-extractor",
    "proof": "paper-proof-grader",
    "glossary": "reading-level-glossary",
    "rewrite": "readability-rewriter",
    "html": "html-article",
}


def now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")


def slug_id(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9]+", "-", value).strip("-").upper()
    return cleaned[:60] or f"PAPER-{datetime.now():%Y%m%d%H%M%S}"


def traveler_path(paper_id: str) -> Path:
    return TRAVELERS / f"{paper_id}.json"


def load_traveler(paper_id: str) -> dict[str, Any]:
    path = traveler_path(paper_id)
    if not path.exists():
        raise FileNotFoundError(f"No traveler found for {paper_id}: {path}")
    return json.loads(path.read_text(encoding="utf-8-sig"))


def save_traveler(traveler: dict[str, Any]) -> Path:
    TRAVELERS.mkdir(parents=True, exist_ok=True)
    path = traveler_path(traveler["paper_id"])
    path.write_text(json.dumps(traveler, indent=2, ensure_ascii=False), encoding="utf-8")
    return path


def create_traveler(
    paper_id: str,
    title: str,
    workflow: str = "first-article",
    stations_required: list[str] | None = None,
) -> dict[str, Any]:
    required = stations_required or WORKFLOW_STATIONS.get(workflow, CANONICAL_STATIONS)
    traveler = {
        "paper_id": paper_id,
        "title": title,
        "entered_at": now_iso(),
        "workflow": workflow,
        "stations_required": list(required),
        "stations_completed": [],
        "stations_failed": [],
        "current_station": None,
        "artifacts": {},
        "status": "new",
    }
    update_status(traveler)
    save_traveler(traveler)
    return traveler


def update_status(traveler: dict[str, Any]) -> str:
    required = set(traveler.get("stations_required", []))
    completed = {entry.get("station") for entry in traveler.get("stations_completed", []) if entry.get("success")}
    failed = {entry.get("station") for entry in traveler.get("stations_failed", [])}
    if required and required <= completed:
        status = "complete"
    elif failed:
        status = "blocked"
    elif completed:
        status = "in_progress"
    else:
        status = "new"
    traveler["status"] = status
    remaining = [station for station in traveler.get("stations_required", []) if station not in completed]
    traveler["current_station"] = remaining[0] if remaining else None
    return status


def stamp_station(
    paper_id: str,
    station: str,
    success: bool,
    artifact: str = "",
    summary: str = "",
    timestamp: str | None = None,
) -> dict[str, Any]:
    traveler = load_traveler(paper_id)
    stamp = {
        "station": station,
        "completed_at": timestamp or now_iso(),
        "success": bool(success),
        "artifact": artifact,
        "summary": summary,
    }
    traveler.setdefault("stations_completed", []).append(stamp)
    if not success:
        traveler.setdefault("stations_failed", []).append(stamp)
    if artifact:
        traveler.setdefault("artifacts", {}).setdefault(station, []).append(artifact)
    update_status(traveler)
    save_traveler(traveler)
    return traveler


def iter_travelers() -> list[dict[str, Any]]:
    TRAVELERS.mkdir(parents=True, exist_ok=True)
    travelers = []
    for path in sorted(TRAVELERS.glob("*.json")):
        travelers.append(json.loads(path.read_text(encoding="utf-8-sig")))
    return travelers


def latest_station_entry(traveler: dict[str, Any], station: str) -> dict[str, Any] | None:
    matches = [entry for entry in traveler.get("stations_completed", []) if entry.get("station") == station]
    return matches[-1] if matches else None


def station_icon(traveler: dict[str, Any], station: str) -> str:
    entry = latest_station_entry(traveler, station)
    if entry:
        return "✅" if entry.get("success") else "❌"
    if traveler.get("current_station") == station:
        return "⏳"
    return "—"


def classification_data(traveler: dict[str, Any]) -> dict[str, Any]:
    artifact_paths = traveler.get("artifacts", {}).get("classify-documents", [])
    if not artifact_paths:
        return {}
    latest = Path(artifact_paths[-1])
    if not latest.is_absolute():
        latest = ROOT / latest
    if not latest.exists():
        return {}
    payload = json.loads(latest.read_text(encoding="utf-8-sig"))
    return payload.get("data", payload)


def write_excel(path: Path = TRACKER_XLSX) -> Path:
    if importlib.util.find_spec("openpyxl") is None:
        raise RuntimeError("openpyxl is required to write PIPELINE_TRACKER.xlsx")
    import openpyxl
    from openpyxl.styles import Font

    travelers = iter_travelers()
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Paper Status"
    headers = ["Paper ID", "Title", "Workflow", "Entered", *STATUS_COLUMNS.keys(), "Status"]
    ws.append(headers)
    for traveler in travelers:
        row = [
            traveler.get("paper_id", ""),
            traveler.get("title", ""),
            traveler.get("workflow", ""),
            traveler.get("entered_at", "")[:10],
        ]
        row.extend(station_icon(traveler, station) for station in STATUS_COLUMNS.values())
        row.append(str(traveler.get("status", "")).upper())
        ws.append(row)

    ws2 = wb.create_sheet("Station Output")
    ws2.append(["Paper ID", "Station", "Timestamp", "Success", "Artifact File", "Key Output"])
    for traveler in travelers:
        for entry in traveler.get("stations_completed", []):
            ws2.append([
                traveler.get("paper_id", ""),
                entry.get("station", ""),
                entry.get("completed_at", ""),
                "YES" if entry.get("success") else "NO",
                entry.get("artifact", ""),
                entry.get("summary", ""),
            ])

    ws3 = wb.create_sheet("Classification Data")
    ws3.append([
        "Paper ID", "doc_type", "classification", "tags", "physics_score", "theology_score",
        "consciousness_score", "evidence_status", "word_count", "reading_grade",
    ])
    for traveler in travelers:
        data = classification_data(traveler)
        spine = data.get("spine_mappings", {}) if isinstance(data, dict) else {}
        reading = data.get("reading_level", {}) if isinstance(data, dict) else {}
        ws3.append([
            traveler.get("paper_id", ""),
            data.get("doc_type", "") if isinstance(data, dict) else "",
            data.get("classification", "") if isinstance(data, dict) else "",
            ", ".join(data.get("tags", [])) if isinstance(data, dict) else "",
            spine.get("physics", {}).get("score", ""),
            spine.get("theology", {}).get("score", ""),
            spine.get("consciousness", {}).get("score", ""),
            spine.get("evidence", {}).get("status", ""),
            data.get("word_count", "") if isinstance(data, dict) else "",
            reading.get("flesch_kincaid_grade", ""),
        ])

    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        for column in sheet.columns:
            width = max(len(str(cell.value or "")) for cell in column[:50]) + 2
            sheet.column_dimensions[column[0].column_letter].width = min(max(width, 10), 60)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def print_status(paper_id: str | None = None) -> None:
    travelers = [load_traveler(paper_id)] if paper_id else iter_travelers()
    if not travelers:
        print("No travelers found.")
        return
    for traveler in travelers:
        print(f"{traveler.get('paper_id')} | {traveler.get('status', '').upper()} | {traveler.get('title')} | next={traveler.get('current_station') or '-'}")
        for entry in traveler.get("stations_completed", []):
            ok = "YES" if entry.get("success") else "NO"
            print(f"  - {entry.get('completed_at')} {entry.get('station')} success={ok} artifact={entry.get('artifact')} {entry.get('summary')}")


def main() -> int:
    parser = argparse.ArgumentParser(description="BACKSIDE workflow traveler tracker")
    parser.add_argument("--status", action="store_true", help="Show all traveler statuses")
    parser.add_argument("--paper", help="Show one paper/traveler by paper_id")
    parser.add_argument("--create", action="store_true", help="Create a traveler")
    parser.add_argument("--paper-id", help="Paper id for --create or --stamp")
    parser.add_argument("--title", default="Untitled", help="Paper title for --create")
    parser.add_argument("--workflow", default="first-article", help="Workflow name for --create")
    parser.add_argument("--stamp", action="store_true", help="Stamp a station execution")
    parser.add_argument("--station", help="Station name for --stamp")
    parser.add_argument("--success", action="store_true", help="Mark --stamp as successful")
    parser.add_argument("--artifact", default="", help="Artifact file for --stamp")
    parser.add_argument("--summary", default="", help="Summary for --stamp")
    parser.add_argument("--write-excel", action="store_true", help="Rebuild PIPELINE_TRACKER.xlsx")
    args = parser.parse_args()

    if args.create:
        paper_id = args.paper_id or slug_id(args.title)
        create_traveler(paper_id, args.title, args.workflow)
        print(f"Created traveler {traveler_path(paper_id)}")
    if args.stamp:
        if not args.paper_id or not args.station:
            parser.error("--stamp requires --paper-id and --station")
        stamp_station(args.paper_id, args.station, args.success, args.artifact, args.summary)
        print(f"Stamped {args.paper_id}: {args.station}")
    if args.write_excel:
        print(f"Wrote {write_excel()}")
    if args.status or args.paper:
        print_status(args.paper)
    if not any([args.create, args.stamp, args.write_excel, args.status, args.paper]):
        parser.print_help()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

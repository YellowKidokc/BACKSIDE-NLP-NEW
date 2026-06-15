from __future__ import annotations

import argparse
import csv
import json
from datetime import datetime
from html import escape
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:  # pragma: no cover - workbook export is optional at import time
    Workbook = None
    load_workbook = None


def resolve_brain_root() -> Path:
    for parent in Path(__file__).resolve().parents:
        if parent.name.lower() == "brain":
            return parent
    mapped = Path("X:\\")
    if mapped.exists():
        return mapped
    return Path(r"\\dlowenas\brain")


BRAIN_ROOT = resolve_brain_root()
DEFAULT_OUTPUT_ROOT = BRAIN_ROOT / "EXPORTS" / "semantic-snapshot.workflow"
BACKSIDE_EXPORTS = BRAIN_ROOT / "Backside" / "EXPORTS"

SURFACES: list[dict[str, str]] = [
    {
        "name": "first-article.workflow",
        "kind": "workflow",
        "path": str(BRAIN_ROOT / "Backside" / "workflows" / "first-article.workflow"),
        "role": "HTML/Markdown/image source -> canonical Markdown -> station outputs -> lossless packet.",
        "run": str(BRAIN_ROOT / "Backside" / "workflows" / "first-article.workflow" / "RUN.bat"),
    },
    {
        "name": "lossless_context_pipeline",
        "kind": "station-library",
        "path": str(BACKSIDE_EXPORTS / "lossless-context"),
        "role": "Nabla address, Master Equation UUID, semantic tags, JSON/HTML snapshot, Postgres append contract.",
        "run": "",
    },
    {
        "name": "chi-tagging.workflow",
        "kind": "workflow",
        "path": str(BRAIN_ROOT / "Backside" / "workflows" / "chi-tagging.workflow"),
        "role": "Canon source tag pass for chi variables and No-Drift reference terms.",
        "run": str(BRAIN_ROOT / "Backside" / "workflows" / "chi-tagging.workflow" / "RUN.bat"),
    },
    {
        "name": "paper-proof-grader.workflow",
        "kind": "workflow",
        "path": str(BRAIN_ROOT / "Backside" / "workflows" / "paper-proof-grader.workflow"),
        "role": "Claim grading, 7Q pressure, formal verification annotations, paper-grade HTML/JSON/XLSX.",
        "run": str(BRAIN_ROOT / "Backside" / "workflows" / "paper-proof-grader.workflow" / "RUN.bat"),
    },
    {
        "name": "axioms.workflow",
        "kind": "workflow",
        "path": str(BRAIN_ROOT / "Backside" / "workflows" / "axioms.workflow"),
        "role": "Axiom snapshot, rigor gates, canonical HTML reference, final review surfaces.",
        "run": str(BRAIN_ROOT / "Backside" / "workflows" / "axioms.workflow" / "RUN_AXIOMS_WORKFLOW.bat"),
    },
]

STATION_DIR = BRAIN_ROOT / "Backside" / "stations"
EXPECTED_STATIONS = [
    "master-equation-canon.station",
    "trinity-canon.station",
    "fruits-spirit-canon.station",
    "operators-canon.station",
]

KEY_OUTPUTS: list[dict[str, str]] = [
    {
        "name": "latest GTQ first-article batch",
        "path": str(BACKSIDE_EXPORTS / "first-article-workflow-series"),
        "pattern": "*/batch-results.csv",
    },
    {
        "name": "latest GTQ stack workbook",
        "path": str(BACKSIDE_EXPORTS / "first-article-workflow-series"),
        "pattern": "*/STACK/gtq-series-stack.xlsx",
    },
    {
        "name": "latest chi-tagging output",
        "path": str(BACKSIDE_EXPORTS / "chi-tagging"),
        "pattern": "*/canon-index.aggregate.json",
    },
    {
        "name": "axiom snapshot HTML",
        "path": str(BRAIN_ROOT / "Backside" / "workflows" / "axioms.workflow" / "papers" / "required_html_outputs_2026-05-11"),
        "pattern": "00_GENESIS-TO-QUANTUM-black-axiom-snapshot.html",
    },
    {
        "name": "paper grader output folder",
        "path": str(BRAIN_ROOT / "Backside" / "workflows" / "paper-proof-grader.workflow" / "OUTPUT"),
        "pattern": "*.paper-grade.json",
    },
]


def latest_match(root: Path, pattern: str) -> str:
    if not root.exists():
        return ""
    matches = sorted(root.glob(pattern), key=lambda item: item.stat().st_mtime if item.exists() else 0, reverse=True)
    return str(matches[0]) if matches else ""


def surface_row(surface: dict[str, str]) -> dict[str, Any]:
    path = Path(surface["path"])
    run = Path(surface["run"]) if surface.get("run") else None
    return {
        **surface,
        "exists": path.exists(),
        "run_exists": run.exists() if run else False,
        "file_count": sum(1 for item in path.rglob("*") if item.is_file()) if path.exists() else 0,
    }


def station_rows() -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for name in EXPECTED_STATIONS:
        path = STATION_DIR / name
        rows.append(
            {
                "name": name,
                "path": str(path),
                "exists": path.exists(),
                "kind": "independent-station",
            }
        )
    return rows


def output_rows() -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for item in KEY_OUTPUTS:
        rows.append({**item, "latest": latest_match(Path(item["path"]), item["pattern"])})
    return rows


def read_csv_rows(path: str, limit: int = 5000) -> list[dict[str, Any]]:
    if not path or not Path(path).exists():
        return []
    with Path(path).open("r", encoding="utf-8-sig", newline="") as handle:
        return [dict(row) for _, row in zip(range(limit), csv.DictReader(handle))]


def read_json(path: str) -> Any:
    if not path or not Path(path).exists():
        return None
    return json.loads(Path(path).read_text(encoding="utf-8"))


def latest_file(root: Path, pattern: str) -> str:
    return latest_match(root, pattern)


def collect_gtq_rows(outputs: list[dict[str, str]]) -> list[dict[str, Any]]:
    batch = next((row["latest"] for row in outputs if row["name"] == "latest GTQ first-article batch"), "")
    rows = read_csv_rows(batch)
    clean: list[dict[str, Any]] = []
    for row in rows:
        export_path = row.get("export_path") or row.get("export_dir") or ""
        manifest_path = str(Path(export_path) / "manifest.json") if export_path else ""
        manifest = read_json(manifest_path) or {}
        clean.append(
            {
                "ordinal": row.get("ordinal", ""),
                "name": row.get("name", ""),
                "status": row.get("status", ""),
                "vector": manifest.get("vector", row.get("vector", "")),
                "master_equation_uuid": manifest.get("master_equation_uuid", ""),
                "semantic_tag_count": manifest.get("semantic_tag_count", ""),
                "address": manifest.get("address", row.get("address", "")),
                "export_path": export_path,
            }
        )
    return clean


def collect_semantic_tag_rows(gtq_rows: list[dict[str, Any]], limit: int = 20000) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for article in gtq_rows:
        export_path = article.get("export_path")
        if not export_path:
            continue
        for tag_path in sorted((Path(export_path) / "lossless").glob("*.semantic-tags.json")):
            tags = read_json(str(tag_path)) or []
            for tag in tags:
                rows.append(
                    {
                        "article": article.get("name", ""),
                        "master_equation_uuid": tag.get("master_equation_uuid", article.get("master_equation_uuid", "")),
                        "tag_type": tag.get("tag_type", ""),
                        "tag_id": tag.get("tag_id", ""),
                        "block_id": tag.get("block_id", ""),
                        "chi_vars": ",".join(tag.get("chi_vars", [])),
                        "label": tag.get("label", ""),
                        "tag_file": str(tag_path),
                    }
                )
                if len(rows) >= limit:
                    return rows
    return rows


def collect_paper_grade_rows() -> list[dict[str, Any]]:
    output = BRAIN_ROOT / "Backside" / "workflows" / "paper-proof-grader.workflow" / "OUTPUT"
    rows: list[dict[str, Any]] = []
    if not output.exists():
        return rows
    for path in sorted(output.glob("*.paper-grade.json"), key=lambda item: item.stat().st_mtime, reverse=True)[:500]:
        data = read_json(str(path)) or {}
        stem = str(path).removesuffix(".paper-grade.json")
        rows.append(
            {
                "paper_id": data.get("paper_id", path.name.removesuffix(".paper-grade.json")),
                "title": data.get("title", ""),
                "claim_count": len(data.get("claims", [])) if isinstance(data.get("claims"), list) else data.get("claim_count", ""),
                "maturity": data.get("maturity", ""),
                "overall_score": data.get("overall_score", data.get("score", "")),
                "json": str(path),
                "html": stem + ".paper-grade.html",
                "xlsx": stem + ".paper-grade.xlsx",
            }
        )
    return rows


def collect_axiom_rows() -> list[dict[str, Any]]:
    root = BRAIN_ROOT / "Backside" / "workflows" / "axioms.workflow" / "01_OUTBOX_REPORTS"
    rows: list[dict[str, Any]] = []
    if not root.exists():
        return rows
    for path in sorted(root.glob("*.paper-grade.json"), key=lambda item: item.stat().st_mtime, reverse=True)[:1000]:
        data = read_json(str(path)) or {}
        stem = str(path).removesuffix(".paper-grade.json")
        rows.append(
            {
                "paper_id": data.get("paper_id", path.name.removesuffix(".paper-grade.json")),
                "title": data.get("title", ""),
                "claim_count": len(data.get("claims", [])) if isinstance(data.get("claims"), list) else data.get("claim_count", ""),
                "json": str(path),
                "html": stem + ".paper-grade.html",
                "xlsx": stem + ".paper-grade.xlsx",
            }
        )
    return rows


def write_url(path: Path, target: str) -> None:
    path.write_text(f"[InternetShortcut]\nURL=file:///{target.replace(chr(92), '/')}\n", encoding="utf-8")


def write_shortcuts(out_dir: Path, surfaces: list[dict[str, Any]], stations: list[dict[str, Any]]) -> None:
    shortcut_dir = out_dir / "00_STATION_SHORTCUTS"
    shortcut_dir.mkdir(parents=True, exist_ok=True)
    for row in surfaces + stations:
        safe = row["name"].replace("\\", "_").replace("/", "_").replace(":", "")
        write_url(shortcut_dir / f"{safe}.url", row["path"])


def add_table_sheet(wb: Any, title: str, rows: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet(title[:31])
    if not rows:
        ws.append(["status"])
        ws.append(["no rows found"])
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    style_sheet(ws)


def style_sheet(ws: Any) -> None:
    header_fill = PatternFill("solid", fgColor="EAF1F8")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="top")
    for column in ws.columns:
        values = [str(cell.value or "") for cell in column[:80]]
        width = min(max([len(value) for value in values] + [10]) + 2, 72)
        ws.column_dimensions[column[0].column_letter].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"


def write_workbook(out_dir: Path, payload: dict[str, Any]) -> str:
    if Workbook is None:
        return ""
    gtq_rows = collect_gtq_rows(payload["outputs"])
    semantic_tag_rows = collect_semantic_tag_rows(gtq_rows)
    paper_grade_rows = collect_paper_grade_rows()
    axiom_rows = collect_axiom_rows()

    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"
    ws.append(["Semantic Snapshot Workbook"])
    ws.append(["Generated", payload["generated_at"]])
    ws.append(["Shared surfaces", len(payload["surfaces"])])
    ws.append(["Independent stations", len(payload["stations"])])
    ws.append(["GTQ articles", len(gtq_rows)])
    ws.append(["Semantic tags", len(semantic_tag_rows)])
    ws.append(["Paper grader outputs", len(paper_grade_rows)])
    ws.append(["Axiom paper outputs", len(axiom_rows)])
    style_sheet(ws)

    add_table_sheet(wb, "Station Map", payload["surfaces"])
    add_table_sheet(wb, "Independent Stations", payload["stations"])
    add_table_sheet(wb, "Latest Outputs", payload["outputs"])
    add_table_sheet(wb, "GTQ Packets", gtq_rows)
    add_table_sheet(wb, "Semantic Tags", semantic_tag_rows)
    add_table_sheet(wb, "Paper Grader", paper_grade_rows)
    add_table_sheet(wb, "Axiom Outputs", axiom_rows)

    path = out_dir / "semantic-snapshot-workbook.xlsx"
    wb.save(path)
    return str(path)


def render_markdown(payload: dict[str, Any]) -> str:
    lines = [
        "# Semantic Snapshot Workflow Map",
        "",
        f"Generated: `{payload['generated_at']}`",
        "",
        "## Route",
        "",
        "```text",
        "HTML/Markdown source",
        "-> first-article.workflow",
        "-> source.canonical.md",
        "-> lossless JSON/HTML",
        "-> master_equation_uuid + semantic-tags.md/json",
        "-> chi/canon tag reference",
        "-> paper-proof-grader.workflow",
        "-> axioms.workflow snapshot / rigor gates",
        "-> Postgres append-only audit memory",
        "```",
        "",
        "## Shared Surfaces",
        "",
        "| Name | Exists | Run Exists | Role | Path |",
        "|---|---:|---:|---|---|",
    ]
    for row in payload["surfaces"]:
        lines.append(f"| {row['name']} | {row['exists']} | {row['run_exists']} | {row['role']} | `{row['path']}` |")
    lines.extend(["", "## Independent Stations", "", "| Station | Exists | Path |", "|---|---:|---|"])
    for row in payload["stations"]:
        lines.append(f"| {row['name']} | {row['exists']} | `{row['path']}` |")
    lines.extend(["", "## Latest Outputs", "", "| Output | Latest Match |", "|---|---|"])
    for row in payload["outputs"]:
        lines.append(f"| {row['name']} | `{row['latest'] or 'MISSING'}` |")
    lines.extend(["", "## Workbook", "", f"`{payload.get('workbook') or 'openpyxl unavailable; workbook not written'}`"])
    lines.extend(
        [
            "",
            "## Operating Decision",
            "",
            "The paper grader remains a shared workflow, not a copied subfolder inside first-article.workflow. The semantic snapshot workflow links to it and records its outputs so the grader can serve GTQ, axioms, canonical papers, and future corpora without duplication.",
        ]
    )
    return "\n".join(lines)


def render_html(markdown: str, payload: dict[str, Any]) -> str:
    surface_rows = "\n".join(
        f"<tr><td>{escape(row['name'])}</td><td>{row['exists']}</td><td>{row['run_exists']}</td><td>{escape(row['role'])}</td><td><code>{escape(row['path'])}</code></td></tr>"
        for row in payload["surfaces"]
    )
    station_rows_html = "\n".join(
        f"<tr><td>{escape(row['name'])}</td><td>{row['exists']}</td><td><code>{escape(row['path'])}</code></td></tr>"
        for row in payload["stations"]
    )
    output_rows_html = "\n".join(
        f"<tr><td>{escape(row['name'])}</td><td><code>{escape(row['latest'] or 'MISSING')}</code></td></tr>"
        for row in payload["outputs"]
    )
    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <title>Semantic Snapshot Workflow Map</title>
  <style>
    body {{ font-family: system-ui, sans-serif; margin: 0; background: #f7f8fb; color: #18212f; }}
    main {{ max-width: 1180px; margin: 0 auto; padding: 32px 24px; }}
    table {{ width: 100%; border-collapse: collapse; background: white; margin: 12px 0 28px; }}
    th, td {{ border: 1px solid #d7deea; padding: 8px; text-align: left; vertical-align: top; }}
    th {{ background: #edf2f8; }}
    code, pre {{ background: #e9eef6; padding: 2px 5px; border-radius: 4px; }}
    pre {{ padding: 14px; overflow: auto; }}
  </style>
</head>
<body>
<main>
  <h1>Semantic Snapshot Workflow Map</h1>
  <p>Generated: <code>{escape(payload['generated_at'])}</code></p>
  <h2>Route</h2>
  <pre>HTML/Markdown source
-> first-article.workflow
-> source.canonical.md
-> lossless JSON/HTML
-> master_equation_uuid + semantic-tags.md/json
-> chi/canon tag reference
-> paper-proof-grader.workflow
-> axioms.workflow snapshot / rigor gates
-> Postgres append-only audit memory</pre>
  <h2>Shared Surfaces</h2>
  <table><tr><th>Name</th><th>Exists</th><th>Run Exists</th><th>Role</th><th>Path</th></tr>{surface_rows}</table>
  <h2>Independent Stations</h2>
  <table><tr><th>Station</th><th>Exists</th><th>Path</th></tr>{station_rows_html}</table>
  <h2>Latest Outputs</h2>
  <table><tr><th>Output</th><th>Latest Match</th></tr>{output_rows_html}</table>
  <h2>Workbook</h2>
  <p><code>{escape(payload.get('workbook') or 'openpyxl unavailable; workbook not written')}</code></p>
  <h2>Operating Decision</h2>
  <p>The paper grader remains a shared workflow, not a copied subfolder inside first-article.workflow. The semantic snapshot workflow links to it and records its outputs so the grader can serve GTQ, axioms, canonical papers, and future corpora without duplication.</p>
</main>
</body>
</html>
"""


def run(output_root: Path = DEFAULT_OUTPUT_ROOT) -> Path:
    run_id = datetime.now().strftime("%Y%m%d-%H%M%S")
    out_dir = output_root / run_id
    out_dir.mkdir(parents=True, exist_ok=True)
    payload = {
        "workflow": "semantic-snapshot.workflow",
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "surfaces": [surface_row(row) for row in SURFACES],
        "stations": station_rows(),
        "outputs": output_rows(),
    }
    workbook_path = write_workbook(out_dir, payload)
    payload["workbook"] = workbook_path
    markdown = render_markdown(payload)
    (out_dir / "semantic-snapshot-map.json").write_text(json.dumps(payload, indent=2), encoding="utf-8")
    (out_dir / "semantic-snapshot-map.md").write_text(markdown, encoding="utf-8")
    (out_dir / "semantic-snapshot-map.html").write_text(render_html(markdown, payload), encoding="utf-8")
    write_shortcuts(out_dir, payload["surfaces"], payload["stations"])
    latest = output_root / "LATEST"
    latest.write_text(str(out_dir), encoding="utf-8")
    return out_dir


def main() -> None:
    parser = argparse.ArgumentParser(description="Build the semantic snapshot workflow map.")
    parser.add_argument("--output-root", default=str(DEFAULT_OUTPUT_ROOT))
    args = parser.parse_args()
    out_dir = run(Path(args.output_root))
    print(f"Semantic snapshot map written: {out_dir}")


if __name__ == "__main__":
    main()

